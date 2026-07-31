"""轉檔器驗收 —— 選用的外部交叉驗證。

參照檔是月報 11401–11412 重組出來的，可直接當標準答案。逐格相同即代表
期間對應、機構對應、合計列取值三者同時正確。缺檔時 skip。
"""

import os
from pathlib import Path

import pytest
from openpyxl import load_workbook

from tools.ingest_fsc import METRICS, convert, discover, norm

REPO_ROOT = Path(__file__).resolve().parent.parent
SRC = REPO_ROOT / "fixtures" / "data" / "金融業務資訊揭露"


def _find_reference_xlsx() -> Path | None:
    configured = os.getenv("SLIDEGEN_XLSX")
    candidates = [
        Path(configured).expanduser() if configured else None,
        REPO_ROOT / "source" / "附件四_預期修正參照資料.xlsx",
        REPO_ROOT / "fixtures" / "data" / "附件四_預期修正參照資料.xlsx",
    ]
    return next((path for path in candidates if path and path.is_file()), None)


REF = _find_reference_xlsx()

PERIODS_114 = [f"114{m:02d}" for m in range(1, 13)]

# 轉出的指標檔 → 附件四對應的工作表
REF_SHEETS = {
    "流通卡數": "P.5預期修正_流通卡數",
    "當月簽帳金額": "P.5預期修正_當月簽帳金額",
}

pytestmark = pytest.mark.skipif(
    not SRC.exists() or REF is None,
    reason="缺少金管會月報或附件四",
)


@pytest.fixture(scope="module")
def converted(tmp_path_factory):
    return convert(SRC, tmp_path_factory.mktemp("fsc_114"), PERIODS_114)


def test_all_24_months_discovered():
    """11301–11412 共 24 個月，少一個就代表 discover 漏檔。"""
    periods = [p.parent.name[:5] for p in discover(SRC)]
    assert len(periods) == 24
    assert periods[0] == "11301" and periods[-1] == "11412"


@pytest.mark.parametrize("metric,ref_sheet", REF_SHEETS.items())
def test_matches_reference_cell_by_cell(converted, metric, ref_sheet):
    """轉出的寬表必須與附件四逐格相同。"""
    got = load_workbook(converted[metric], data_only=True).worksheets[0]
    ref = load_workbook(REF, data_only=True)[ref_sheet]

    # 附件四：第 1 列標題、第 2–34 列資料（合計在第 34 列末列）
    ref_rows = {
        norm(ref.cell(r, 1).value): [ref.cell(r, c).value for c in range(2, 14)]
        for r in range(2, 35)
    }
    got_rows = {
        norm(got.cell(r, 1).value): [got.cell(r, c).value for c in range(2, 14)]
        for r in range(2, got.max_row + 1)
    }

    assert set(got_rows) == set(ref_rows), "機構名單與附件四不符"

    for name, ref_vals in ref_rows.items():
        assert got_rows[name] == ref_vals, f"{name} 的逐月數值與附件四不符"


def test_period_headers_are_roc_yyymm(converted):
    ws = load_workbook(converted["流通卡數"], data_only=True).worksheets[0]
    headers = [ws.cell(1, c).value for c in range(2, 14)]
    assert headers == [int(p) for p in PERIODS_114]


def test_derived_metrics_are_not_materialised():
    """衍生量不得落地——有兩個真相來源就必然有一天會不一致。"""
    for banned in ("有效卡率", "平均每卡簽帳金額", "市占率", "年增率", "月增率"):
        assert banned not in METRICS, f"{banned} 是衍生量，應由 engine 計算而非轉檔產出"
