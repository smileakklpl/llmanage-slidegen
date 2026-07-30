"""兩種輸入版型必須產出同形狀的 payload。

同一份資料在現實中有兩種擺法：主辦方給的附件四是**單檔多工作表**，
金管會官網下載下來的是**多檔單表**。版型是輸入端的細節，不該外溢成
`dataset_loader` / `metric_engine` 的分支邏輯——所以這裡驗的是
「兩條路收斂到同一個形狀」，而不只是各自跑得動。

另外驗 dataset_id 的可讀性：backend 的 dataset_id 是內容雜湊 UUID5，
而 metric_engine 拿它組 metric_key。不覆寫的話指標會叫
`f47ac10b_58cc_....value`，LLM 只看得到指標目錄，無從判斷該引用哪個。
"""

import pytest
from openpyxl import Workbook

from ppt_generation.data import backend_bridge


PERIODS = [11401, 11402, 11403, 11404, 11405, 11406]
BANKS = ["臺灣銀行", "中國信託", "台新銀行", "玉山銀行"]
METRICS = ["流通卡數", "有效卡數"]


def _rows(seed: int):
    for index, bank in enumerate(BANKS, start=1):
        yield [bank] + [seed * 1000 + index * 100 + month for month in PERIODS]


def _fill(worksheet, metric: str, seed: int) -> None:
    worksheet.title = metric
    worksheet.append(["金融機構名稱"] + PERIODS)

    for row in _rows(seed):
        worksheet.append(row)


def _single_file(directory, name="workbook.xlsx"):
    """版型 A：一個 .xlsx，每張工作表一個指標。"""
    directory.mkdir(parents=True, exist_ok=True)
    path = directory / name
    workbook = Workbook()

    for index, metric in enumerate(METRICS):
        worksheet = (
            workbook.active if index == 0 else workbook.create_sheet()
        )
        _fill(worksheet, metric, index + 1)

    workbook.save(path)
    return path


def _file_per_metric(directory):
    """版型 B：一個目錄，每個 .xlsx 一張工作表一個指標。"""
    directory.mkdir(parents=True, exist_ok=True)

    for index, metric in enumerate(METRICS):
        workbook = Workbook()
        _fill(workbook.active, metric, index + 1)
        workbook.save(directory / f"{metric}.xlsx")

    return directory


def _dataset_values(payload):
    """把 payload 攤平成 {dataset_id: {(entity, 欄名): 值}} 供逐格比對。"""
    flattened = {}

    for dataset in payload["datasets"]:
        cells = {}

        for record in dataset.get("records") or []:
            values = record.get("values") or {}
            row_key = tuple(
                str(item.get("value"))
                for item in values.values()
                if isinstance(item.get("value"), str)
            )

            for column, item in values.items():
                cells[(row_key, column)] = item.get("value")

        flattened[dataset["dataset_id"]] = cells

    return flattened


# ---------------------------------------------------------------------------
# discover_excel_files
# ---------------------------------------------------------------------------
def test_discover_single_file(tmp_path):
    path = _single_file(tmp_path)

    assert backend_bridge.discover_excel_files(path) == [path]


def test_discover_directory_is_sorted(tmp_path):
    directory = _file_per_metric(tmp_path / "fsc")

    found = backend_bridge.discover_excel_files(directory)

    assert [path.name for path in found] == sorted(
        path.name for path in found
    )
    assert len(found) == len(METRICS)


def test_discover_skips_excel_lock_files(tmp_path):
    """Excel 開著檔案時會留下 ~$ 暫存鎖檔，送進 backend 會炸。"""
    directory = _file_per_metric(tmp_path / "fsc")
    (directory / "~$流通卡數.xlsx").write_bytes(b"lock")

    found = backend_bridge.discover_excel_files(directory)

    assert all(not path.name.startswith("~$") for path in found)
    assert len(found) == len(METRICS)


def test_discover_rejects_missing_path(tmp_path):
    with pytest.raises(backend_bridge.NoExcelInputError):
        backend_bridge.discover_excel_files(tmp_path / "not_there.xlsx")


def test_discover_rejects_non_excel(tmp_path):
    path = tmp_path / "data.csv"
    path.write_text("a,b\n1,2\n", encoding="utf-8")

    with pytest.raises(backend_bridge.NoExcelInputError):
        backend_bridge.discover_excel_files(path)


def test_discover_rejects_empty_directory(tmp_path):
    directory = tmp_path / "empty"
    directory.mkdir()

    with pytest.raises(backend_bridge.NoExcelInputError):
        backend_bridge.discover_excel_files(directory)


# ---------------------------------------------------------------------------
# ingest_excel：兩種版型
# ---------------------------------------------------------------------------
def test_layout_a_single_file_multi_sheet(tmp_path):
    payload = backend_bridge.ingest_excel(_single_file(tmp_path))

    assert len(payload["datasets"]) == len(METRICS)
    assert payload["source_files"] == ["workbook.xlsx"]


def test_layout_b_directory_of_files(tmp_path):
    payload = backend_bridge.ingest_excel(_file_per_metric(tmp_path / "fsc"))

    assert len(payload["datasets"]) == len(METRICS)
    assert len(payload["source_files"]) == len(METRICS)


def test_both_layouts_produce_the_same_values(tmp_path):
    """同一份資料、兩種擺法，逐格數值必須相同。"""
    payload_a = backend_bridge.ingest_excel(_single_file(tmp_path / "a"))
    payload_b = backend_bridge.ingest_excel(_file_per_metric(tmp_path / "b"))

    assert _dataset_values(payload_a) == _dataset_values(payload_b)


def test_both_layouts_share_the_same_payload_shape(tmp_path):
    """下游只認 payload 的形狀；少一個鍵就得在 loader 裡長出分支。"""
    payload_a = backend_bridge.ingest_excel(_single_file(tmp_path / "a"))
    payload_b = backend_bridge.ingest_excel(_file_per_metric(tmp_path / "b"))

    assert set(payload_a) == set(payload_b)
    assert {
        "filename",
        "pipeline_status",
        "source_files",
        "datasets",
        "warnings",
        "errors",
    } <= set(payload_a)


# ---------------------------------------------------------------------------
# dataset_id
# ---------------------------------------------------------------------------
def test_dataset_ids_are_readable_metric_names(tmp_path):
    payload = backend_bridge.ingest_excel(_single_file(tmp_path))

    assert sorted(
        dataset["dataset_id"] for dataset in payload["datasets"]
    ) == sorted(METRICS)


def test_backend_hash_id_is_preserved(tmp_path):
    """覆寫成可讀 id 之後，原本的雜湊 id 仍要留著，追溯才不會斷。"""
    payload = backend_bridge.ingest_excel(_single_file(tmp_path))

    for dataset in payload["datasets"]:
        assert dataset["backend_dataset_id"]
        assert dataset["backend_dataset_id"] != dataset["dataset_id"]


def test_dataset_ids_stay_unique_across_files_with_same_sheet_name(tmp_path):
    """兩個檔案有同名工作表時，靠檔名消歧，不可互相覆蓋。"""
    directory = tmp_path / "dup"
    directory.mkdir()

    for name, seed in (("113年", 1), ("114年", 2)):
        workbook = Workbook()
        _fill(workbook.active, "流通卡數", seed)
        workbook.save(directory / f"{name}.xlsx")

    payload = backend_bridge.ingest_excel(directory)
    ids = [dataset["dataset_id"] for dataset in payload["datasets"]]

    assert len(ids) == 2
    assert len(set(ids)) == 2


def test_filename_is_pinned_to_the_real_file(tmp_path):
    """backend 內部可能用暫存檔跑，filename 要釘回真實檔名供稽核。"""
    directory = _file_per_metric(tmp_path / "fsc")

    payload = backend_bridge.ingest_excel(directory)
    filenames = {dataset["filename"] for dataset in payload["datasets"]}

    assert filenames == {f"{metric}.xlsx" for metric in METRICS}


# ---------------------------------------------------------------------------
# sheet_name
# ---------------------------------------------------------------------------
def test_sheet_name_selects_one_sheet(tmp_path):
    payload = backend_bridge.ingest_excel(
        _single_file(tmp_path),
        sheet_name="有效卡數",
    )

    assert [dataset["dataset_id"] for dataset in payload["datasets"]] == [
        "有效卡數"
    ]


def test_sheet_name_is_rejected_for_directory_input(tmp_path):
    """目錄輸入下「哪一張工作表」沒有唯一解，寧可擋掉也不要猜。"""
    directory = _file_per_metric(tmp_path / "fsc")

    with pytest.raises(ValueError):
        backend_bridge.ingest_excel(directory, sheet_name="流通卡數")


# ---------------------------------------------------------------------------
# 抽不出資料
# ---------------------------------------------------------------------------
def test_empty_workbook_raises(tmp_path):
    path = tmp_path / "empty.xlsx"
    Workbook().save(path)

    with pytest.raises(backend_bridge.NoExcelInputError):
        backend_bridge.ingest_excel(path)
