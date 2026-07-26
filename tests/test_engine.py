"""engine 實作正確性 — T1 的核心：算出來的值必須逐格等於附件四。

tests/test_metric_definitions.py 驗的是**定義**對不對（用 pandas 直接算）。
這支驗的是**實作**對不對（走 reader + build_store 這條真的管線）。
兩者缺一不可：定義對但實作寫錯，一樣會產出錯的簡報。
"""

import json
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from contracts.sheet_map import SheetMap
from engine.metrics import build_store, entity_slug, metric_slug
from engine.reader import read_sheet
from paths import find_xlsx

XLSX = find_xlsx()
from paths import GOLDEN as _G
GOLDEN = _G / "sheet_map.json"

pytestmark = pytest.mark.skipif(
    XLSX is None, reason="找不到附件四，請放進 fixtures/data/ 或設 B_LLM_XLSX"
)

P7 = ["P.7預期修正_流通卡數", "P.7預期修正_當月簽帳金額"]


def _spec(name):
    smap = SheetMap.model_validate(json.loads(GOLDEN.read_text(encoding="utf-8")))
    return smap.get(name)


@pytest.mark.parametrize("sheet", P7)
def test_share_matches_source_derived_column(sheet):
    """市佔率必須逐格等於附件四的衍生欄——這是 T1 的硬斷言。"""
    spec = _spec(sheet)
    sd = read_sheet(XLSX, spec)
    store = build_store([sd])
    slug, _ = metric_slug(sheet)

    ws = load_workbook(XLSX, data_only=True)[sheet]
    ecol = column_index_from_string(spec.entity_col)
    dcol = column_index_from_string(spec.derived_cols[0])
    names = {str(ws.cell(r, ecol).value).strip(): r
             for r in range(spec.first_data_row, spec.last_data_row + 1)}

    for i, name in enumerate(sd.rows, 1):
        got = store.value_of(f"{entity_slug(name, i)}_{slug}_share")
        expected = ws.cell(names[name], dcol).value
        assert abs(got - expected) < 1e-12, f"{name} 市佔率不符"


@pytest.mark.parametrize("sheet", P7)
def test_total_row_excluded_from_entities(sheet):
    """合計列絕不可出現在 rows 裡，否則它會佔據第 1 名並使其後名次全部位移。"""
    sd = read_sheet(XLSX, sheet_spec := _spec(sheet))
    assert "總計" not in sd.rows
    assert sd.total, "合計列沒被讀到"
    # 合計必須大於任何單一機構，這是它是合計的算術證據
    assert sd.total_sum() > max(sum(v.values()) for v in sd.rows.values())


@pytest.mark.parametrize("sheet", P7)
def test_rank_starts_at_one_and_is_dense(sheet):
    sd = read_sheet(XLSX, _spec(sheet))
    store = build_store([sd])
    slug, _ = metric_slug(sheet)
    ranks = sorted(
        m.value for k, m in store.metrics.items() if k.endswith(f"_{slug}_rank")
    )
    assert ranks[0] == 1
    assert ranks == list(range(1, len(ranks) + 1)), "名次有跳號或重複"


def test_yoy_is_not_computable():
    """附件四只有 114 年，任何 YoY 都必須 computable=false（附件三錯誤 #1）。

    key 形如 market_total_cards_yoy_11412——帶期間，因為每一期各自判定基期是否存在。
    """
    store = build_store([read_sheet(XLSX, _spec(P7[0]))])
    yoy = [k for k in store.metrics if "_yoy_" in k]
    assert yoy, "應該要產出 YoY 指標（即使不可計算也要列出來）"
    for k in yoy:
        assert not store.get(k).computable, f"{k} 不該可計算，附件四沒有 113 年基期"
        assert store.value_of(k) is None
        assert store.get(k).reason, "computable=false 必須說明原因"
