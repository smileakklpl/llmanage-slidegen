"""金管會信用卡月報 → entity_by_period 寬表。

    python -m tools.ingest_fsc --out fixtures/data/fsc_114 --periods 11401,…,11412

月報每月一檔，形狀是 entity_by_metric（列＝機構，欄＝13 個指標）。
下游要的是 entity_by_period（列＝機構，欄＝期間，一指標一檔）。
32 機構 × 24 月 × 6 指標 = 4608 格，不適合手動整理。

## 來源檔的四個坑（實測 24 個檔）

1. **標題列在第 4 列**，第 1 列是合併標題、第 3 列是單位與資料月份。
2. **第 38 列之後是註釋**（揭露項目認定標準），共 14 行說明文字。
   openpyxl 的 max_row 會算進去，資料其實只到第 37 列。
3. **工作表名稱不統一**：'11301'…'揭露'…'揭露.' 共三種，
   所以一律取 worksheets[0]，不靠名稱。
4. **標題含全形空白**：'流通卡數　　　'，比對前必須正規化。

好消息是形狀本身完全一致：24 個檔都是標題列 4、總計列 37、14 欄，
32 家機構名單與標題列組合各只有一種變體。
"""

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook, load_workbook

HEADER_ROW = 4
FIRST_DATA_ROW = 5
TOTAL_ROW = 37
ENTITY_COL = 1

_PERIOD = re.compile(r"(\d{5})")


def norm(s) -> str:
    """去掉半形與全形空白。來源檔的標題帶尾隨全形空格。"""
    return re.sub(r"[\s　]+", "", str(s or ""))


# 轉出的檔名 → 來源標題（正規化後比對）。
# 只收原始量：有效卡率、平均每卡簽帳金額、月增率、年增率、市占率變化
# 一律由 engine 計算，不在這裡轉出——衍生值若也落地就會有兩個真相來源。
METRICS: Dict[str, str] = {
    "流通卡數": "流通卡數",
    "有效卡數": "有效卡數",
    "當月簽帳金額": "當月簽帳金額",
    "循環信用餘額": "循環信用餘額",
    "未到期分期付款餘額": "未到期分期付款餘額",
    "當月轉銷呆帳金額": "當月轉銷呆帳金額",
}


@dataclass
class MonthTable:
    period: str
    entities: List[str]
    values: Dict[str, Dict[str, float]]  # {指標標題: {機構: 值}}
    total: Dict[str, float]  # {指標標題: 合計值}


def _num(v) -> Optional[float]:
    return float(v) if isinstance(v, (int, float)) else None


def read_month(path: Path) -> MonthTable:
    period_src = _PERIOD.search(path.parent.name) or _PERIOD.search(path.name)
    if period_src is None:
        raise ValueError(f"無法從路徑推出期間：{path}")
    period = period_src.group(1)

    ws = load_workbook(path, data_only=True).worksheets[0]

    if norm(ws.cell(HEADER_ROW, ENTITY_COL).value) != "金融機構名稱":
        raise ValueError(
            f"{path.name} 第 {HEADER_ROW} 列不是標題列（實際為 "
            f"{ws.cell(HEADER_ROW, ENTITY_COL).value!r}）。來源格式可能改版，拒絕繼續。"
        )
    if norm(ws.cell(TOTAL_ROW, ENTITY_COL).value) != "總計":
        raise ValueError(
            f"{path.name} 第 {TOTAL_ROW} 列不是合計列。合計列位置錯誤會讓市佔率分母失準，拒絕繼續。"
        )

    headers = {norm(ws.cell(HEADER_ROW, c).value): c for c in range(2, 15)}
    entities = [
        norm(ws.cell(r, ENTITY_COL).value) for r in range(FIRST_DATA_ROW, TOTAL_ROW)
    ]

    values: Dict[str, Dict[str, float]] = {}
    total: Dict[str, float] = {}
    for header, col in headers.items():
        if not header:
            continue
        values[header] = {
            norm(ws.cell(r, ENTITY_COL).value): v
            for r in range(FIRST_DATA_ROW, TOTAL_ROW)
            if (v := _num(ws.cell(r, col).value)) is not None
        }
        if (t := _num(ws.cell(TOTAL_ROW, col).value)) is not None:
            total[header] = t

    return MonthTable(period=period, entities=entities, values=values, total=total)


def discover(root: Path) -> List[Path]:
    """找出所有月報 xlsx，依期間排序。.ods 是同一份資料的另一種格式，跳過。"""
    return sorted(root.glob("*/*.xlsx"), key=lambda p: _PERIOD.search(p.parent.name).group(1))


def build_wide(months: List[MonthTable], source_header: str) -> Workbook:
    """把某個指標的 N 個月組成 entity_by_period 寬表，合計列放末列。"""
    periods = [m.period for m in months]
    entities = months[0].entities

    wb = Workbook()
    ws = wb.active
    ws.title = source_header[:31]  # Excel 工作表名稱上限
    ws.append(["金融機構名稱"] + [int(p) for p in periods])
    for e in entities:
        ws.append([e] + [m.values.get(source_header, {}).get(e) for m in months])
    ws.append(["總計"] + [m.total.get(source_header) for m in months])
    return wb


def convert(
    src_root: Path, out_dir: Path, periods: Optional[List[str]] = None
) -> Dict[str, Path]:
    """把來源目錄轉成一指標一檔的 xlsx，回傳 {檔名: 路徑}。

    periods 指定要包含哪些期間；None 表示全部。
    做成參數是為了同一支程式產出兩份資料集：
      114 年單獨一份供 T2 斷言「無基期不得產出 YoY」，
      113+114 一份證明基期存在時算得出來。
    """
    months = [read_month(p) for p in discover(src_root)]
    if periods:
        want = set(periods)
        months = [m for m in months if m.period in want]
    if not months:
        raise ValueError(f"{src_root} 底下找不到符合的月報")

    # 機構名單必須跨月完全一致，否則寬表會出現空洞而不自知
    base = months[0].entities
    for m in months[1:]:
        if m.entities != base:
            diff = set(m.entities) ^ set(base)
            raise ValueError(f"{m.period} 的機構名單與 {months[0].period} 不同：{sorted(diff)}")

    out_dir.mkdir(parents=True, exist_ok=True)
    written: Dict[str, Path] = {}
    for out_name, header in METRICS.items():
        if header not in months[0].values:
            raise ValueError(f"來源檔沒有欄位 {header!r}，可用的有：{sorted(months[0].values)}")
        path = out_dir / f"{out_name}.xlsx"
        build_wide(months, header).save(path)
        written[out_name] = path
    return written


def main() -> None:
    import argparse

    repo_root = Path(__file__).resolve().parent.parent
    default_source = repo_root / "fixtures" / "data" / "金融業務資訊揭露"
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", default=str(default_source))
    ap.add_argument("--out", required=True, help="輸出目錄，如 fixtures/data/fsc_114")
    ap.add_argument("--periods", default=None, help="逗號分隔，如 11401,11402,…；省略為全部")
    args = ap.parse_args()

    periods = [p.strip() for p in args.periods.split(",")] if args.periods else None
    written = convert(Path(args.src), Path(args.out), periods)
    for name, path in written.items():
        print(f"  {name:<12s} → {path}")
    print(f"\n共 {len(written)} 個指標檔")


if __name__ == "__main__":
    main()
