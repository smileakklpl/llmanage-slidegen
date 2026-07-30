"""金管會信用卡月報 → 單一多工作表 xlsx（附件四版型）。

    python -m tools.build_fsc_workbook --out fixtures/data/fsc_114_workbook.xlsx \
        --periods 11401,11402,11403,11404,11405,11406,11407,11408,11409,11410,11411,11412

`tools/ingest_fsc.py` 把月報轉成「一指標一檔」（`fixtures/data/fsc_114/` 底下
6 個 xlsx）。這支轉成另一種版型：**一個檔案、多個工作表**，形狀對齊
`source/附件四_預期修正參照資料.xlsx`——每張工作表是一個指標的
`金融機構名稱 × 民國年月` 交叉表。

兩種版型都要存在，因為它們是 backend ingestion 的兩條輸入路徑：

    fixtures/data/fsc_114/           多檔單表（每檔 1 個 dataset）
    fixtures/data/fsc_114_workbook.xlsx  單檔多表（1 檔 6 個 dataset）

資料同源——都走 `ingest_fsc.read_month()`，所以兩種版型的數值必然一致，
不會出現「換個版型數字就不一樣」這種最難查的問題。

## 與附件四刻意不同的兩點

1. **不寫入市佔率欄。** 附件四的 P.7 工作表有一欄命題方自算的市佔率。
   這裡只輸出原始量，市佔率與排名一律由 engine 即時計算。理由見
   `fixtures/README.md`：衍生量一旦落地就有兩個真相來源，兩邊算法有出入時
   沒有辦法判斷誰對。附件四那一欄的用途是反過來驗我們的公式，不是拿來當輸入。

2. **工作表名稱不加 `P.5預期修正_` 前綴。** 那是命題方對「附件三第幾頁該修正」
   的頁次標記，不是資料本身的屬性。這裡直接用指標名稱當工作表名稱。
   （`P.{頁碼}_{指標名稱}` 是 renderer **輸出**稽核 Excel 時的命名規則，
   見 steering 的命名慣例，不適用於輸入資料。）

## 合計列

來源月報第 37 列是「總計」，`build_wide()` 會把它放在寬表末列。這裡沿用同樣
的處理：合計列留在資料裡，由 engine 的 reader 在計算市佔率時排除
（`metric_definitions.json` 的 `ranking` 規則：合計列若未排除會佔據第 1 名）。
"""

import argparse
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook

from tools.ingest_fsc import METRICS, MonthTable, discover, read_month

#: 工作表順序。依「規模 → 使用 → 風險」排，讓人打開檔案時的閱讀動線合理，
#: 也對齊簡報章節的推進順序（市場概況 → 活躍度 → 獲利 → 風險）。
SHEET_ORDER: List[str] = [
    "流通卡數",
    "有效卡數",
    "當月簽帳金額",
    "循環信用餘額",
    "未到期分期付款餘額",
    "當月轉銷呆帳金額",
]


def _load_months(
    src_root: Path,
    periods: Optional[List[str]] = None,
) -> List[MonthTable]:
    """讀取來源月報並檢查機構名單跨月一致。"""
    months = [read_month(path) for path in discover(src_root)]

    if periods:
        want = set(periods)
        months = [month for month in months if month.period in want]

    if not months:
        raise ValueError(f"{src_root} 底下找不到符合的月報")

    # 與 ingest_fsc.convert() 相同的檢查：機構名單必須跨月完全一致，
    # 否則交叉表會出現空洞而不自知。
    base = months[0].entities

    for month in months[1:]:
        if month.entities != base:
            diff = set(month.entities) ^ set(base)
            raise ValueError(
                f"{month.period} 的機構名單與 {months[0].period} 不同："
                f"{sorted(diff)}"
            )

    return months


def build_workbook(months: List[MonthTable]) -> Workbook:
    """把多個指標組成單一多工作表 Workbook。"""
    periods = [month.period for month in months]
    entities = months[0].entities

    workbook = Workbook()
    # 預設會有一張空白工作表，等第一個指標寫進去時改名沿用。
    workbook.remove(workbook.active)

    for metric_name in SHEET_ORDER:
        header = METRICS[metric_name]

        if header not in months[0].values:
            raise ValueError(
                f"來源檔沒有欄位 {header!r}，"
                f"可用的有：{sorted(months[0].values)}"
            )

        worksheet = workbook.create_sheet(title=metric_name[:31])
        worksheet.append(
            ["金融機構名稱"] + [int(period) for period in periods]
        )

        for entity in entities:
            worksheet.append(
                [entity]
                + [
                    month.values.get(header, {}).get(entity)
                    for month in months
                ]
            )

        worksheet.append(
            ["總計"]
            + [month.total.get(header) for month in months]
        )

        # 凍結首列與首欄：33 列 × 13 欄捲動時看不到自己在哪一格。
        worksheet.freeze_panes = "B2"

    return workbook


def build(
    src_root: Path,
    out_path: Path,
    periods: Optional[List[str]] = None,
) -> Path:
    """讀月報 → 建立多工作表 xlsx → 存檔，回傳輸出路徑。"""
    months = _load_months(src_root, periods)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    build_workbook(months).save(out_path)

    return out_path


def main() -> None:
    from paths import FSC_RAW

    parser = argparse.ArgumentParser(
        description="把金管會月報整理成單一多工作表 xlsx（附件四版型）",
    )
    parser.add_argument("--src", default=str(FSC_RAW))
    parser.add_argument(
        "--out",
        required=True,
        help="輸出檔案，如 fixtures/data/fsc_114_workbook.xlsx",
    )
    parser.add_argument(
        "--periods",
        default=None,
        help="逗號分隔，如 11401,11402,…；省略為全部",
    )

    args = parser.parse_args()

    periods = (
        [period.strip() for period in args.periods.split(",")]
        if args.periods
        else None
    )

    out_path = build(Path(args.src), Path(args.out), periods)

    from openpyxl import load_workbook

    workbook = load_workbook(out_path, read_only=True)

    print(f"  {out_path}")
    print(f"  {len(workbook.sheetnames)} 個工作表：")

    for name in workbook.sheetnames:
        sheet = workbook[name]
        print(f"    {name:<12s} {sheet.max_row} 列 × {sheet.max_column} 欄")

    workbook.close()


if __name__ == "__main__":
    main()
