from pathlib import Path

from openpyxl import Workbook

from app.ingestion.extractor import (
    extract_excel_tables,
)
from app.ingestion.schemas import (
    QualitySeverity,
    QualityStatus,
)
from app.ingestion.validator import (
    validate_workbook_extraction,
)


def _issue_codes(report) -> set[str]:
    return {
        issue.code
        for table in report.tables
        for issue in table.issues
    }


def test_clean_table_passes(
    tmp_path: Path,
) -> None:
    file_path = tmp_path / "clean.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "銷售資料"

    worksheet.append(
        ["月份", "營收", "訂單數"]
    )

    worksheet.append(
        ["2026-01", 100000, 20]
    )

    worksheet.append(
        ["2026-02", 120000, 25]
    )

    worksheet.append(
        ["2026-03", 135000, 27]
    )

    workbook.save(file_path)

    extraction = extract_excel_tables(
        file_path
    )

    report = validate_workbook_extraction(
        extraction
    )

    assert report.status == QualityStatus.PASS
    assert report.error_count == 0


def test_duplicate_rows_warning(
    tmp_path: Path,
) -> None:
    file_path = tmp_path / "duplicate.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "資料"

    worksheet.append(
        ["名稱", "數值"]
    )

    worksheet.append(["A", 100])
    worksheet.append(["A", 100])
    worksheet.append(["B", 200])

    workbook.save(file_path)

    extraction = extract_excel_tables(
        file_path
    )

    report = validate_workbook_extraction(
        extraction
    )

    assert "DUPLICATE_ROWS" in (
        _issue_codes(report)
    )

    assert (
        report.status
        == QualityStatus.WARNING
    )


def test_missing_value_warning(
    tmp_path: Path,
) -> None:
    file_path = tmp_path / "missing.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "資料"

    worksheet.append(
        ["名稱", "數值"]
    )

    worksheet.append(["A", None])
    worksheet.append(["B", None])
    worksheet.append(["C", 300])

    workbook.save(file_path)

    extraction = extract_excel_tables(
        file_path
    )

    report = validate_workbook_extraction(
        extraction
    )

    assert "HIGH_MISSING_RATE" in (
        _issue_codes(report)
    )


def test_type_mismatch_error(
    tmp_path: Path,
) -> None:
    file_path = tmp_path / "type.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "資料"

    worksheet.append(
        ["名稱", "數值"]
    )

    worksheet.append(["A", 100])
    worksheet.append(["B", 200])
    worksheet.append(["C", 300])

    workbook.save(file_path)

    extraction = extract_excel_tables(
        file_path
    )

    table = extraction.tables[0]

    # 模擬資料在抽取後遭到異常修改。
    table.rows[1].cells["數值"].value = (
        "不是數字"
    )

    report = validate_workbook_extraction(
        extraction
    )

    assert "COLUMN_TYPE_MISMATCH" in (
        _issue_codes(report)
    )

    assert report.status == QualityStatus.FAIL


def test_balanced_balance_sheet(
    tmp_path: Path,
) -> None:
    file_path = tmp_path / "balanced.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "資產負債表"

    worksheet["A1"] = (
        "ABC 公司資產負債表"
    )

    worksheet["A2"] = (
        "單位：新臺幣千元"
    )

    worksheet.append(
        ["會計科目", "2025年", "2026年"]
    )

    worksheet.append(
        ["流動資產", 500000, 550000]
    )

    worksheet.append(
        ["非流動資產", 300000, 310000]
    )

    worksheet.append(
        ["資產總計", 800000, 860000]
    )

    worksheet.append(
        ["流動負債", 200000, 220000]
    )

    worksheet.append(
        ["非流動負債", 150000, 160000]
    )

    worksheet.append(
        ["負債總計", 350000, 380000]
    )

    worksheet.append(
        ["權益總計", 450000, 480000]
    )

    worksheet.append(
        ["負債及權益總計", 800000, 860000]
    )

    workbook.save(file_path)

    extraction = extract_excel_tables(
        file_path
    )

    report = validate_workbook_extraction(
        extraction
    )

    assert (
        "BALANCE_SHEET_UNBALANCED"
        not in _issue_codes(report)
    )


def test_unbalanced_balance_sheet(
    tmp_path: Path,
) -> None:
    file_path = tmp_path / "unbalanced.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "資產負債表"

    worksheet["A1"] = (
        "ABC 公司資產負債表"
    )

    worksheet["A2"] = (
        "單位：新臺幣千元"
    )

    worksheet.append(
        ["會計科目", "2025年", "2026年"]
    )

    worksheet.append(
        ["流動資產", 500000, 550000]
    )

    worksheet.append(
        ["非流動資產", 300000, 310000]
    )

    worksheet.append(
        ["資產總計", 800000, 900000]
    )

    worksheet.append(
        ["流動負債", 200000, 220000]
    )

    worksheet.append(
        ["非流動負債", 150000, 160000]
    )

    worksheet.append(
        ["負債總計", 350000, 380000]
    )

    worksheet.append(
        ["權益總計", 450000, 480000]
    )

    worksheet.append(
        ["負債及權益總計", 800000, 860000]
    )

    workbook.save(file_path)

    extraction = extract_excel_tables(
        file_path
    )

    report = validate_workbook_extraction(
        extraction
    )

    assert (
        "BALANCE_SHEET_UNBALANCED"
        in _issue_codes(report)
    )

    assert report.status == QualityStatus.FAIL