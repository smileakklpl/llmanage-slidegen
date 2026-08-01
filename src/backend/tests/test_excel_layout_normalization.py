"""Regression matrix for deterministic Excel layout normalization."""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table

from app.ingestion.cell_tokenizer import (
    TokenKind,
    assemble_accounting_values,
    tokenize_sheet,
)
from app.ingestion.extractor import extract_excel_tables
from app.ingestion.layout_analyzer import analyze_sheet_layouts
from app.ingestion.layout_graph import (
    build_grid_relationship_graph,
)
from app.ingestion.workbook_ir import build_workbook_ir
from app.ingestion.normalization_spec import (
    LayoutStrategy,
    NormalizationSpec,
    OutputColumnSpec,
    TableRegionSpec,
)
from app.ingestion.pipeline import run_ingestion_pipeline


REPOSITORY_ROOT = Path(__file__).resolve().parents[3]
AMKR_FIXTURE = REPOSITORY_ROOT / "fixtures" / "data" / "amkr.xlsx"
ASE_FIXTURE = REPOSITORY_ROOT / "fixtures" / "data" / "ASE.xlsx"


def _save(workbook: Workbook, path: Path) -> Path:
    workbook.save(path)
    workbook.close()
    return path


def test_period_axis_marker_pairs_preserve_value_cells_and_contract(
    tmp_path: Path,
) -> None:
    path = tmp_path / "period_markers.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Operations"
    worksheet["A1"] = "Operating metrics"
    worksheet["A2"] = "Unit: USD thousands"
    worksheet["D3"] = 2025
    worksheet["H3"] = 2024
    worksheet["L3"] = 2023

    worksheet["A4"] = "Revenue"
    worksheet["D4"] = "$"
    worksheet["E4"] = "1,234"
    worksheet["H4"] = "$"
    worksheet["I4"] = "(1,100)"
    worksheet["L4"] = "$"
    worksheet["M4"] = 980

    worksheet["A5"] = "Cost"
    worksheet["D5"] = 800
    worksheet["H5"] = 700
    worksheet["L5"] = 650

    worksheet["A6"] = "Profit"
    worksheet["D6"] = 434
    worksheet["H6"] = "—"
    worksheet["L6"] = 330
    _save(workbook, path)

    extraction = extract_excel_tables(path)

    assert extraction.table_count == 1
    table = extraction.tables[0]
    assert table.layout_strategy == LayoutStrategy.MARKER_VALUE_PAIR.value
    assert [column.label for column in table.columns] == [
        "項目",
        "2025",
        "2024",
        "2023",
    ]
    assert table.metadata.unit == "USD thousands"
    assert table.columns[0].unit is None
    assert table.columns[1].unit == "USD thousands"

    revenue = table.rows[0].cells
    assert revenue["2025"].value == 1234
    assert revenue["2025"].source.cell == "E4"
    assert revenue["2024"].value == -1100
    assert revenue["2024"].source.cell == "I4"
    assert revenue["2023"].source.cell == "M4"

    profit = table.rows[2].cells
    assert profit["2024"].raw_value == "—"
    assert profit["2024"].value is None
    assert profit["2024"].source.cell == "H6"

    assert table.normalization_spec is not None
    serialized = json.dumps(
        table.normalization_spec,
        ensure_ascii=False,
        sort_keys=True,
    )
    restored = NormalizationSpec.model_validate_json(serialized)
    assert restored.contract_version == "1.0"
    assert restored.strategy == LayoutStrategy.MARKER_VALUE_PAIR

    result = run_ingestion_pipeline(path)
    dataset = result.datasets[0]
    assert dataset.normalization_spec == table.normalization_spec
    assert dataset.layout_confidence == table.layout_confidence
    assert dataset.requires_human_review is False


def test_plain_period_cross_table_uses_period_axis_strategy(
    tmp_path: Path,
) -> None:
    path = tmp_path / "cross_table.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Monthly"
    worksheet.append(["Entity", 2024, 2025])
    worksheet.append(["North", 100, 120])
    worksheet.append(["South", 80, 95])
    worksheet.append(["East", 70, 90])
    _save(workbook, path)

    table = extract_excel_tables(path).tables[0]

    assert table.layout_strategy == LayoutStrategy.PERIOD_AXIS.value
    assert [column.label for column in table.columns] == [
        "Entity",
        "2024",
        "2025",
    ]
    assert table.rows[0].cells["2024"].source.cell == "B2"
    assert table.rows[0].cells["2025"].source.cell == "C2"


def test_merged_multi_row_header_is_flattened_without_losing_sources(
    tmp_path: Path,
) -> None:
    path = tmp_path / "merged_header.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Regional"
    worksheet["A1"] = "Regional performance"
    worksheet.merge_cells("A1:C1")
    worksheet["A2"] = "Region"
    worksheet["B2"] = 2025
    worksheet.merge_cells("B2:C2")
    worksheet["B3"] = "Revenue"
    worksheet["C3"] = "Profit"
    worksheet.append(["North", 120, 30])
    worksheet.append(["South", 100, 20])
    worksheet.append(["East", 90, 18])
    _save(workbook, path)

    table = extract_excel_tables(path).tables[0]

    assert table.layout_strategy == LayoutStrategy.MULTI_ROW_HEADER.value
    assert [column.label for column in table.columns] == [
        "Region",
        "2025 Revenue",
        "2025 Profit",
    ]
    assert table.header_row == 2
    assert table.rows[0].excel_row == 4
    assert table.rows[0].cells["2025_revenue"].source.cell == "B4"
    assert table.normalization_spec is not None
    assert table.normalization_spec["header_rows"] == [2, 3]


def test_vertical_and_side_by_side_tables_become_separate_regions(
    tmp_path: Path,
) -> None:
    path = tmp_path / "multiple_tables.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Multiple"

    for column, header in (("A", "Product"), ("B", "Sales")):
        worksheet[f"{column}1"] = header
    for row, values in enumerate((("A", 10), ("B", 20), ("C", 30)), 2):
        worksheet.cell(row, 1, values[0])
        worksheet.cell(row, 2, values[1])

    worksheet["D1"] = "Region"
    worksheet["E1"] = "Orders"
    for row, values in enumerate((("N", 3), ("S", 4), ("E", 5)), 2):
        worksheet.cell(row, 4, values[0])
        worksheet.cell(row, 5, values[1])

    worksheet["A7"] = "Month"
    worksheet["B7"] = "Cost"
    for row, values in enumerate((("Jan", 7), ("Feb", 8), ("Mar", 9)), 8):
        worksheet.cell(row, 1, values[0])
        worksheet.cell(row, 2, values[1])
    _save(workbook, path)

    extraction = extract_excel_tables(path)

    assert extraction.table_count == 3
    assert {table.full_range for table in extraction.tables} == {
        "Multiple!A1:B4",
        "Multiple!D1:E4",
        "Multiple!A7:B10",
    }
    assert all(
        table.layout_strategy == LayoutStrategy.FLAT_TABLE.value
        for table in extraction.tables
    )


def test_ambiguous_layout_is_extracted_only_for_human_review(
    tmp_path: Path,
) -> None:
    path = tmp_path / "ambiguous.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Ambiguous"
    worksheet.append(["Label", 10])
    worksheet.append(["A", 1])
    worksheet.append(["B", 2])
    worksheet.append(["C", 3])
    _save(workbook, path)

    result = run_ingestion_pipeline(path)

    assert len(result.datasets) == 1
    dataset = result.datasets[0]
    assert dataset.layout_confidence is not None
    assert 0.0 <= dataset.layout_confidence < 0.75
    assert dataset.requires_human_review is True
    assert any(
        "版面正規化信心分數" in reason
        or "資料集信心分數低於" in reason
        for reason in dataset.review_reasons
    )


def test_normalization_spec_below_gate_is_not_extractable() -> None:
    spec = NormalizationSpec(
        sheet_name="Sheet1",
        region=TableRegionSpec(
            region_id="region_1",
            min_row=1,
            max_row=3,
            min_column=1,
            max_column=2,
            discovery_method="test",
        ),
        strategy=LayoutStrategy.FLAT_TABLE,
        header_rows=[1],
        data_rows=[2, 3],
        output_columns=[
            OutputColumnSpec(
                label="Name",
                source_columns=[1],
                semantic_role="dimension",
            ),
            OutputColumnSpec(
                label="Value",
                source_columns=[2],
            ),
        ],
        confidence=0.74,
    )

    assert spec.is_extractable is False
    assert spec.requires_human_review is True


@pytest.mark.skipif(
    not AMKR_FIXTURE.exists(),
    reason="local Amkor acceptance workbook is not available",
)
def test_local_amkr_workbook_uses_year_columns_and_real_value_cells() -> None:
    extraction = extract_excel_tables(AMKR_FIXTURE)

    assert extraction.table_count >= 3
    income = next(
        table
        for table in extraction.tables
        if table.sheet_name == "amkr income"
    )
    assert [column.label for column in income.columns] == [
        "項目",
        "2025",
        "2024",
        "2023",
    ]
    assert all(
        "$" not in column.label
        and not column.label.startswith("未命名")
        for column in income.columns
    )

    net_sales = next(
        row
        for row in income.rows
        if row.cells["項目"].value == "Net sales"
    )
    assert net_sales.cells["2025"].value == 6_707_981
    assert net_sales.cells["2025"].source.cell == "E6"
    assert net_sales.cells["2024"].value == 6_317_692
    assert net_sales.cells["2024"].source.cell == "K6"
    assert net_sales.cells["2023"].value == 6_503_065
    assert net_sales.cells["2023"].source.cell == "Q6"

    for table in extraction.tables:
        assert table.normalization_spec is not None
        NormalizationSpec.model_validate(table.normalization_spec)


def test_formula_layout_scoring_uses_cached_values() -> None:
    formula_workbook = Workbook()
    formula_sheet = formula_workbook.active
    formula_sheet.title = "Formula"
    formula_sheet.append(["Item", "Value"])
    formula_sheet.append(["A", "=1+1"])
    formula_sheet.append(["B", "=2+2"])

    value_workbook = Workbook()
    value_sheet = value_workbook.active
    value_sheet.title = "Formula"
    value_sheet.append(["Item", "Value"])
    value_sheet.append(["A", 2])
    value_sheet.append(["B", 4])

    workbook_ir = build_workbook_ir(
        formula_workbook,
        value_workbook,
    )
    specs = analyze_sheet_layouts(workbook_ir.sheets["Formula"])

    assert len(specs) == 1
    assert specs[0].strategy == LayoutStrategy.FLAT_TABLE
    assert specs[0].data_rows == [2, 3]
    assert workbook_ir.sheets["Formula"].cell_at(2, 2).formula == "=1+1"
    assert workbook_ir.sheets["Formula"].value_at(2, 2) == 2
    formula_workbook.close()
    value_workbook.close()


def test_explicit_excel_table_does_not_hide_uncovered_table(
    tmp_path: Path,
) -> None:
    path = tmp_path / "explicit_and_ordinary.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Mixed"
    worksheet.append(["Product", "Sales", None, "Region", "Orders"])
    worksheet.append(["A", 10, None, "N", 3])
    worksheet.append(["B", 20, None, "S", 4])
    worksheet.append(["C", 30, None, "E", 5])
    worksheet.add_table(Table(displayName="SalesTable", ref="A1:B4"))
    _save(workbook, path)

    extraction = extract_excel_tables(path)

    assert extraction.table_count == 2
    assert {table.full_range for table in extraction.tables} == {
        "Mixed!A1:B4",
        "Mixed!D1:E4",
    }


def test_period_table_does_not_absorb_side_by_side_flat_table(
    tmp_path: Path,
) -> None:
    path = tmp_path / "period_beside_flat.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "SideBySide"
    worksheet.append(["Entity", 2024, 2025, None, "Category", "Value"])
    worksheet.append(["North", 100, 120, None, "A", 1])
    worksheet.append(["South", 80, 95, None, "B", 2])
    worksheet.append(["East", 70, 90, None, "C", 3])
    _save(workbook, path)

    extraction = extract_excel_tables(path)

    assert extraction.table_count == 2
    assert {table.full_range for table in extraction.tables} == {
        "SideBySide!A1:C4",
        "SideBySide!E1:F4",
    }
    assert {
        table.layout_strategy for table in extraction.tables
    } == {
        LayoutStrategy.PERIOD_AXIS.value,
        LayoutStrategy.FLAT_TABLE.value,
    }


def test_repeated_periods_are_resolved_as_multi_row_header(
    tmp_path: Path,
) -> None:
    path = tmp_path / "repeated_periods.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Hierarchical"
    worksheet.append(["Region", 2025, 2025])
    worksheet.append([None, "Revenue", "Profit"])
    worksheet.append(["North", 120, 30])
    worksheet.append(["South", 100, 20])
    worksheet.append(["East", 90, 18])
    _save(workbook, path)

    table = extract_excel_tables(path).tables[0]

    assert table.layout_strategy == LayoutStrategy.MULTI_ROW_HEADER.value
    assert [column.label for column in table.columns] == [
        "Region",
        "2025 Revenue",
        "2025 Profit",
    ]
    assert [
        source.cell for source in table.columns[1].header_sources
    ] == ["B1", "B2"]


def test_text_only_flat_table_requires_review(
    tmp_path: Path,
) -> None:
    path = tmp_path / "text_only.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Notes"
    worksheet.append(["Topic", "Description"])
    worksheet.append(["A", "First note"])
    worksheet.append(["B", "Second note"])
    worksheet.append(["C", "Third note"])
    _save(workbook, path)

    result = run_ingestion_pipeline(path)

    assert len(result.datasets) == 1
    assert result.datasets[0].layout_confidence == 0.89
    assert result.datasets[0].requires_human_review is True


def test_persisted_normalization_specs_can_be_replayed(
    tmp_path: Path,
) -> None:
    path = tmp_path / "refresh.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Refresh"
    worksheet.append(["Entity", 2024, 2025])
    worksheet.append(["North", 100, 120])
    worksheet.append(["South", 80, 95])
    worksheet.append(["East", 70, 90])
    _save(workbook, path)

    first = extract_excel_tables(path)
    plans = [table.normalization_spec for table in first.tables]
    replayed = extract_excel_tables(
        path,
        normalization_specs=plans,
    )

    assert replayed.model_dump(mode="json") == first.model_dump(mode="json")


def test_accounting_fragment_assembler_is_deterministic() -> None:
    assembled = assemble_accounting_values(
        ["$", "(1,860,511", ")\u00a0"]
    )

    assert assembled.value == -1_860_511
    assert assembled.raw_value == "(1,860,511)"
    assert assembled.source_indexes == (1, 2)
    assert assembled.transformations == (
        "join_accounting_fragments",
        "remove_thousands_separator",
        "apply_parentheses_negative",
    )


def test_grid_graph_merges_vertical_pages_and_keeps_currency_axis(
    tmp_path: Path,
) -> None:
    path = tmp_path / "paged_multicurrency.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Statement"
    worksheet["A1"] = "Consolidated statement"
    worksheet["A2"] = "Amounts in Thousands"

    for header_row, currency_row in ((3, 4), (9, 10)):
        worksheet.cell(header_row, 3, 2023)
        worksheet.cell(header_row, 7, 2024)
        worksheet.cell(header_row, 11, 2025)
        worksheet.cell(currency_row, 3, "NT$")
        worksheet.cell(currency_row, 7, "NT$")
        worksheet.cell(currency_row, 11, "NT$")
        worksheet.cell(currency_row, 15, "US$")

    worksheet["A5"] = "Loss"
    worksheet["D5"] = "(1,234"
    worksheet["E5"] = ")"
    worksheet["H5"] = 200
    worksheet["L5"] = 300
    worksheet["P5"] = 10
    worksheet["A6"] = "Profit"
    worksheet["D6"] = 500
    worksheet["H6"] = 600
    worksheet["L6"] = 700
    worksheet["P6"] = 23

    worksheet["A11"] = "Assets"
    worksheet["D11"] = 800
    worksheet["H11"] = 900
    worksheet["L11"] = 1000
    worksheet["P11"] = 33
    worksheet["A12"] = "Liabilities"
    worksheet["D12"] = 300
    worksheet["H12"] = 400
    worksheet["L12"] = 500
    worksheet["P12"] = 17
    _save(workbook, path)

    extraction = extract_excel_tables(path)

    assert extraction.table_count == 1
    table = extraction.tables[0]
    assert table.row_count == 4
    assert table.normalization_spec is not None
    assert table.normalization_spec["invariants_passed"] is True
    assert table.normalization_spec["explained_numeric_ratio"] == 1.0
    assert "merge_vertical_continuation" in (
        table.normalization_spec["transformations"]
    )
    assert [column.label for column in table.columns] == [
        "項目",
        "2023 TWD",
        "2024 TWD",
        "2025 TWD",
        "2025 USD",
    ]

    loss = next(
        row for row in table.rows if row.cells["項目"].value == "Loss"
    )
    value = loss.cells["2023_twd"]
    assert value.value == -1234
    assert [source.cell for source in value.sources] == ["D5", "E5"]

    result = run_ingestion_pipeline(path)
    normalized_loss = next(
        record
        for record in result.datasets[0].records
        if record.source_row == 5
    )
    assert {
        evidence.cell
        for evidence in normalized_loss.values["2023_twd"].evidence
    } == {"D5", "E5"}


def test_token_graph_contains_row_and_column_relationships(
    tmp_path: Path,
) -> None:
    path = tmp_path / "graph.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Graph"
    worksheet.append(["Entity", 2024, 2025])
    worksheet.append(["North", 100, 120])
    worksheet.append(["South", 80, 95])
    _save(workbook, path)

    formula = load_workbook(path, data_only=False)
    values = load_workbook(path, data_only=True)
    sheet = build_workbook_ir(formula, values).sheets["Graph"]
    tokenized = tokenize_sheet(sheet)
    graph = build_grid_relationship_graph(sheet)

    assert any(
        token.kind == TokenKind.PERIOD
        and token.value == "2024"
        for token in tokenized.tokens
    )
    assert graph.horizontal_edges
    assert graph.vertical_edges
    formula.close()
    values.close()


@pytest.mark.skipif(
    not ASE_FIXTURE.exists(),
    reason="local ASE acceptance workbook is not available",
)
def test_local_ase_workbook_converges_to_four_traceable_tables() -> None:
    extraction = extract_excel_tables(ASE_FIXTURE)

    assert extraction.table_count == 4
    tables = {table.sheet_name: table for table in extraction.tables}
    assert set(tables) == {
        "ASE cashflow",
        "ASE income",
        "ASE balance 2025",
        "ASE balance 2024",
    }
    cashflow = tables["ASE cashflow"]
    assert [column.label for column in cashflow.columns] == [
        "項目",
        "2023 TWD",
        "2024 TWD",
        "2025 TWD",
        "2025 USD",
    ]
    assert cashflow.row_count == 75
    assert cashflow.normalization_spec is not None
    assert cashflow.normalization_spec["invariants_passed"] is True
    assert cashflow.normalization_spec["explained_numeric_ratio"] == 1.0

    loss = next(row for row in cashflow.rows if row.excel_row == 16)
    value = loss.cells["2023_twd"]
    assert value.value == -1_860_511
    assert [source.cell for source in value.sources] == ["D16", "E16"]

    income = tables["ASE income"]
    assert income.metadata.unit == (
        "Amounts in Thousands Except Earnings Per Share"
    )
    assert any("單位說明包含例外" in warning for warning in income.warnings)


def test_quality_warning_is_scoped_to_matching_table(
    tmp_path: Path,
) -> None:
    path = tmp_path / "scoped_quality.xlsx"
    workbook = Workbook()
    clean = workbook.active
    clean.title = "Clean"
    clean.append(["Name", "Value"])
    clean.append(["A", 10])
    clean.append(["B", 20])
    clean.append(["C", 30])

    duplicate = workbook.create_sheet("Duplicate")
    duplicate.append(["Name", "Value"])
    duplicate.append(["X", 1])
    duplicate.append(["X", 1])
    duplicate.append(["X", 1])
    _save(workbook, path)

    result = run_ingestion_pipeline(path)
    datasets = {
        dataset.name: dataset
        for dataset in result.datasets
    }

    assert datasets["Clean"].requires_human_review is False
    assert datasets["Clean"].review_reasons == []
    assert datasets["Duplicate"].requires_human_review is True
    assert "資料品質驗證產生警告" not in (
        datasets["Duplicate"].review_reasons
    )
    assert (
        datasets["Duplicate"].confidence
        < datasets["Clean"].confidence
    )
