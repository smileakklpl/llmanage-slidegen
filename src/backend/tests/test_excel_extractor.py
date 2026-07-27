from pathlib import Path

from openpyxl import Workbook

from app.ingestion.extractor import (
    extract_excel_tables,
)
from app.ingestion.schemas import (
    ColumnDataType,
    SheetContentType,
)


def test_extract_basic_table(
    tmp_path: Path,
) -> None:
    file_path = tmp_path / "sales.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "銷售資料"

    worksheet.append(["月份", "營收", "訂單數"])
    worksheet.append(["2026-01", 100000, 20])
    worksheet.append(["2026-02", 120000, 25])
    worksheet.append(["2026-03", 135000, 27])

    workbook.save(file_path)

    result = extract_excel_tables(file_path)

    assert result.table_count == 1

    table = result.tables[0]

    assert table.sheet_name == "銷售資料"
    assert table.header_row == 1
    assert table.row_count == 3
    assert table.column_count == 3
    assert table.full_range == "銷售資料!A1:C4"

    assert table.columns[0].label == "月份"
    assert table.columns[1].label == "營收"

    assert (
        table.columns[1].data_type
        == ColumnDataType.INTEGER
    )

    assert (
        table.rows[0]
        .cells["營收"]
        .value
        == 100000
    )

    assert (
        table.rows[0]
        .cells["營收"]
        .source
        .cell
        == "B2"
    )


def test_extract_title_and_unit(
    tmp_path: Path,
) -> None:
    file_path = tmp_path / "revenue.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "營收"

    worksheet["A1"] = "ABC 公司營收表"
    worksheet["A2"] = "單位：新臺幣千元"

    worksheet.append(
        ["年度", "營業收入", "營業利益"]
    )

    worksheet.append(
        [2025, 500000, 80000]
    )

    worksheet.append(
        [2026, 620000, 105000]
    )

    workbook.save(file_path)

    result = extract_excel_tables(file_path)
    table = result.tables[0]

    assert (
        table.metadata.title
        == "ABC 公司營收表"
    )

    assert (
        table.metadata.unit
        == "新臺幣千元"
    )

    assert table.header_row == 3

    assert (
        table.columns[1].unit
        == "新臺幣千元"
    )


def test_extract_balance_sheet(
    tmp_path: Path,
) -> None:
    file_path = tmp_path / "balance.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "資產負債表"

    worksheet["A1"] = (
        "ABC 公司資產負債表"
    )

    worksheet["A2"] = (
        "單位：新臺幣千元"
    )

    worksheet.append(
        ["會計科目", "2025年", "2026年"]
    )

    worksheet.append(
        ["流動資產", 500000, 550000]
    )

    worksheet.append(
        ["非流動資產", 300000, 310000]
    )

    worksheet.append(
        ["資產總計", 800000, 860000]
    )

    worksheet.append(
        ["流動負債", 200000, 220000]
    )

    worksheet.append(
        ["非流動負債", 150000, 160000]
    )

    worksheet.append(
        ["負債總計", 350000, 380000]
    )

    worksheet.append(
        ["權益總計", 450000, 480000]
    )

    worksheet.append(
        ["負債及權益總計", 800000, 860000]
    )

    workbook.save(file_path)

    result = extract_excel_tables(file_path)
    table = result.tables[0]

    assert (
        table.table_kind
        == SheetContentType.FINANCIAL_STATEMENT
    )

    assert table.header_row == 3
    assert table.row_count == 8

    assert (
        table.rows[2]
        .cells["會計科目"]
        .value
        == "資產總計"
    )

    assert (
        table.rows[2]
        .cells["2026年"]
        .value
        == 860000
    )


def test_preserve_formula(
    tmp_path: Path,
) -> None:
    file_path = tmp_path / "formula.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "計算資料"

    worksheet.append(
        ["產品", "單價", "數量", "總額"]
    )

    worksheet.append(
        ["產品 A", 100, 3, None]
    )

    worksheet["D2"] = "=B2*C2"

    worksheet.append(
        ["產品 B", 200, 2, None]
    )

    worksheet["D3"] = "=B3*C3"

    workbook.save(file_path)

    result = extract_excel_tables(file_path)
    table = result.tables[0]

    assert (
        table.rows[0]
        .cells["總額"]
        .formula
        == "=B2*C2"
    )

    assert (
        table.rows[0]
        .cells["總額"]
        .source
        .cell
        == "D2"
    )


def test_skip_non_table_sheet(
    tmp_path: Path,
) -> None:
    file_path = tmp_path / "mixed.xlsx"

    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "資料"
    worksheet.append(["名稱", "數值"])
    worksheet.append(["A", 100])
    worksheet.append(["B", 200])
    worksheet.append(["C", 300])

    note_sheet = workbook.create_sheet("備註")
    note_sheet["A1"] = "這是一段說明文字"
    note_sheet["A2"] = "沒有表格結構"

    workbook.save(file_path)

    result = extract_excel_tables(file_path)

    assert result.table_count == 1
    assert result.tables[0].sheet_name == "資料"
    assert "備註" in result.skipped_sheets