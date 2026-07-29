from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

from app.ingestion.classifier import inspect_excel_content
from app.ingestion.schemas import (
    FinancialStatementSubtype,
    SheetContentType,
)


def test_empty_excel_sheet(tmp_path: Path) -> None:
    file_path = tmp_path / "empty.xlsx"

    workbook = Workbook()
    workbook.save(file_path)

    result = inspect_excel_content(file_path)

    assert result.sheet_count == 1
    assert (
        result.sheets[0].primary_content_type
        == SheetContentType.EMPTY
    )


def test_structured_table(tmp_path: Path) -> None:
    file_path = tmp_path / "table.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "銷售資料"

    worksheet.append(["月份", "北區", "南區"])
    worksheet.append(["2026-01", 100, 120])
    worksheet.append(["2026-02", 110, 135])
    worksheet.append(["2026-03", 125, 140])

    workbook.save(file_path)

    result = inspect_excel_content(file_path)
    sheet = result.sheets[0]

    assert (
        sheet.primary_content_type
        == SheetContentType.STRUCTURED_TABLE
    )
    assert sheet.detected_header_row == 1
    assert sheet.numeric_cells == 6


def test_balance_sheet(tmp_path: Path) -> None:
    file_path = tmp_path / "balance_sheet.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "資產負債表"

    worksheet.append(["ABC 公司資產負債表"])
    worksheet.append(["單位：新臺幣千元"])
    worksheet.append(["會計科目", "2025年", "2026年"])
    worksheet.append(["流動資產", 500000, 550000])
    worksheet.append(["非流動資產", 300000, 310000])
    worksheet.append(["資產總計", 800000, 860000])
    worksheet.append(["流動負債", 200000, 220000])
    worksheet.append(["非流動負債", 150000, 160000])
    worksheet.append(["負債總計", 350000, 380000])
    worksheet.append(["權益總計", 450000, 480000])
    worksheet.append(
        ["負債及權益總計", 800000, 860000]
    )

    workbook.save(file_path)

    result = inspect_excel_content(file_path)
    sheet = result.sheets[0]

    assert (
        sheet.primary_content_type
        == SheetContentType.FINANCIAL_STATEMENT
    )
    assert (
        sheet.financial_statement_subtype
        == FinancialStatementSubtype.BALANCE_SHEET
    )
    assert sheet.financial_statement_score >= 0.8


def test_native_chart(tmp_path: Path) -> None:
    file_path = tmp_path / "chart.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "圖表資料"

    worksheet.append(["區域", "營收"])
    worksheet.append(["北區", 100])
    worksheet.append(["中區", 120])
    worksheet.append(["南區", 150])

    chart = BarChart()
    chart.title = "各區營收"

    chart_data = Reference(
        worksheet,
        min_col=2,
        min_row=1,
        max_row=4,
    )

    chart_categories = Reference(
        worksheet,
        min_col=1,
        min_row=2,
        max_row=4,
    )

    chart.add_data(
        chart_data,
        titles_from_data=True,
    )
    chart.set_categories(chart_categories)

    worksheet.add_chart(chart, "D2")

    workbook.save(file_path)

    result = inspect_excel_content(file_path)
    sheet = result.sheets[0]

    assert (
        sheet.primary_content_type
        == SheetContentType.MIXED_CONTENT
    )
    assert (
        SheetContentType.STRUCTURED_TABLE
        in sheet.components
    )
    assert (
        SheetContentType.NATIVE_CHART
        in sheet.components
    )
    assert sheet.chart_count == 1


def test_multiple_sheet_types(tmp_path: Path) -> None:
    file_path = tmp_path / "multiple.xlsx"

    workbook = Workbook()

    table_sheet = workbook.active
    table_sheet.title = "一般資料"
    table_sheet.append(["產品", "銷售額"])
    table_sheet.append(["A", 100])
    table_sheet.append(["B", 200])
    table_sheet.append(["C", 300])

    empty_sheet = workbook.create_sheet("空白頁")

    workbook.save(file_path)

    result = inspect_excel_content(file_path)

    assert result.sheet_count == 2
    assert (
        result.sheets[0].primary_content_type
        == SheetContentType.STRUCTURED_TABLE
    )
    assert (
        result.sheets[1].primary_content_type
        == SheetContentType.EMPTY
    )