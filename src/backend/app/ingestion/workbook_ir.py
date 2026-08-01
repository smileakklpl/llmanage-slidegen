"""Deterministic intermediate representation for Excel layout analysis."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Iterable

from openpyxl.cell.cell import MergedCell
from openpyxl.utils.cell import range_boundaries
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.ingestion.classifier import is_period_like_header


MAX_LAYOUT_ROWS = 10_000
MAX_LAYOUT_COLUMNS = 200


@dataclass(frozen=True)
class CellIR:
    """A non-empty source cell retained with its Excel coordinates."""

    row: int
    column: int
    coordinate: str
    value: Any
    number_format: str | None
    formula: str | None


@dataclass(frozen=True)
class TableRegion:
    """A rectangular table candidate inside one worksheet."""

    region_id: str
    min_row: int
    max_row: int
    min_column: int
    max_column: int
    discovery_method: str = "non_empty_bands"

    def contains(self, row: int, column: int) -> bool:
        return (
            self.min_row <= row <= self.max_row
            and self.min_column <= column <= self.max_column
        )


@dataclass
class SheetIR:
    """Sparse worksheet representation used by all layout strategies."""

    title: str
    max_row: int
    max_column: int
    cells: dict[tuple[int, int], CellIR]
    merged_ranges: tuple[str, ...] = ()
    merged_anchors: dict[tuple[int, int], tuple[int, int]] = field(
        default_factory=dict
    )
    explicit_regions: tuple[TableRegion, ...] = ()
    regions: tuple[TableRegion, ...] = ()

    def cell_at(
        self,
        row: int,
        column: int,
        *,
        resolve_merged: bool = True,
    ) -> CellIR | None:
        key = (row, column)
        if resolve_merged:
            key = self.merged_anchors.get(key, key)
        return self.cells.get(key)

    def value_at(
        self,
        row: int,
        column: int,
        *,
        resolve_merged: bool = True,
    ) -> Any:
        cell = self.cell_at(
            row,
            column,
            resolve_merged=resolve_merged,
        )
        return None if cell is None else cell.value


@dataclass(frozen=True)
class WorkbookIR:
    """Workbook-level layout IR, including named-range evidence."""

    sheets: dict[str, SheetIR]


def _is_empty(value: Any) -> bool:
    return value is None or (
        isinstance(value, str) and not value.strip()
    )


def _is_period_value(value: Any) -> bool:
    if is_period_like_header(value):
        return True
    if isinstance(value, (date, datetime)):
        return True
    if not isinstance(value, str):
        return False
    compact = value.strip().replace("年", "").replace("月", "")
    if not compact.isdigit():
        return False
    try:
        return is_period_like_header(int(compact))
    except ValueError:
        return False


def _axis_bands(
    occupied: Iterable[int],
    *,
    allowed_blank_gap: int,
) -> list[tuple[int, int]]:
    values = sorted(set(occupied))
    if not values:
        return []

    bands: list[tuple[int, int]] = []
    start = values[0]
    previous = values[0]

    for value in values[1:]:
        blank_gap = value - previous - 1
        if blank_gap > allowed_blank_gap:
            bands.append((start, previous))
            start = value
        previous = value

    bands.append((start, previous))
    return bands


def _has_period_axis(
    sheet: SheetIR,
    min_row: int,
    max_row: int,
    min_column: int,
    max_column: int,
) -> bool:
    for row in range(min_row, max_row + 1):
        count = sum(
            _is_period_value(
                sheet.value_at(
                    row,
                    column,
                    resolve_merged=False,
                )
            )
            for column in range(min_column, max_column + 1)
        )
        if count >= 2:
            return True
    return False


def _period_aware_column_bands(
    sheet: SheetIR,
    min_row: int,
    max_row: int,
    column_bands: list[tuple[int, int]],
    active_cells: set[tuple[int, int]],
) -> list[tuple[int, int]]:
    """Merge only bands that participate in the same sparse period axis."""
    best_anchors: list[int] = []
    for row in range(min_row, max_row + 1):
        anchors = [
            column
            for column in range(
                column_bands[0][0],
                column_bands[-1][1] + 1,
            )
            if (row, column) in active_cells
            and _is_period_value(
                sheet.value_at(
                    row,
                    column,
                    resolve_merged=False,
                )
            )
        ]
        if len(anchors) > len(best_anchors):
            best_anchors = anchors

    if len(best_anchors) < 2:
        return column_bands

    anchor_band_indexes = sorted(
        {
            index
            for index, (band_min, band_max) in enumerate(column_bands)
            if any(band_min <= anchor <= band_max for anchor in best_anchors)
        }
    )
    if len(anchor_band_indexes) <= 1:
        return column_bands

    first_anchor_band = anchor_band_indexes[0]
    last_anchor_band = anchor_band_indexes[-1]
    merge_start = max(0, first_anchor_band - 1)
    merged = (
        column_bands[merge_start][0],
        column_bands[last_anchor_band][1],
    )
    return (
        column_bands[:merge_start]
        + [merged]
        + column_bands[last_anchor_band + 1 :]
    )


def detect_table_regions(sheet: SheetIR) -> tuple[TableRegion, ...]:
    """Split a sheet into explicit and heuristic table candidates.

    Excel tables and named ranges are authoritative seeds, but do not hide
    ordinary tables elsewhere. Two blank rows split vertical regions. One
    wholly blank column splits side-by-side tables unless period anchors span
    multiple sparse bands, as in marker/value financial layouts.
    """
    explicit_regions = list(sheet.explicit_regions)
    active_cells = {
        (row, column)
        for row, column in sheet.cells
        if not any(
            region.contains(row, column)
            for region in explicit_regions
        )
    }
    occupied_rows = [row for row, _ in active_cells]
    row_bands = _axis_bands(
        occupied_rows,
        allowed_blank_gap=1,
    )
    regions: list[TableRegion] = list(explicit_regions)

    for min_row, max_row in row_bands:
        occupied_columns = [
            column
            for row, column in active_cells
            if min_row <= row <= max_row
        ]
        if not occupied_columns:
            continue

        column_bands = _axis_bands(
            occupied_columns,
            allowed_blank_gap=0,
        )
        column_bands = _period_aware_column_bands(
            sheet,
            min_row,
            max_row,
            column_bands,
            active_cells,
        )

        for min_column, max_column in column_bands:
            cell_count = sum(
                min_row <= row <= max_row
                and min_column <= column <= max_column
                for row, column in active_cells
            )
            if cell_count < 2:
                continue
            region_number = len(regions) + 1
            regions.append(
                TableRegion(
                    region_id=f"region_{region_number}",
                    min_row=min_row,
                    max_row=max_row,
                    min_column=min_column,
                    max_column=max_column,
                )
            )

    return tuple(regions)


def _merged_anchor_map(
    worksheet: Worksheet,
) -> tuple[tuple[str, ...], dict[tuple[int, int], tuple[int, int]]]:
    ranges: list[str] = []
    anchors: dict[tuple[int, int], tuple[int, int]] = {}

    for merged_range in worksheet.merged_cells.ranges:
        ranges.append(str(merged_range))
        min_column, min_row, max_column, max_row = range_boundaries(
            str(merged_range)
        )
        anchor = (min_row, min_column)
        for row in range(min_row, max_row + 1):
            for column in range(min_column, max_column + 1):
                anchors[(row, column)] = anchor

    return tuple(ranges), anchors


def _worksheet_table_regions(
    worksheet: Worksheet,
) -> list[TableRegion]:
    regions: list[TableRegion] = []
    for index, table_name in enumerate(worksheet.tables, start=1):
        table = worksheet.tables[table_name]
        min_column, min_row, max_column, max_row = range_boundaries(
            table.ref
        )
        regions.append(
            TableRegion(
                region_id=f"excel_table_{index}",
                min_row=min_row,
                max_row=max_row,
                min_column=min_column,
                max_column=max_column,
                discovery_method="excel_table",
            )
        )
    return regions


def build_sheet_ir(
    worksheet: Worksheet,
    *,
    value_worksheet: Worksheet | None = None,
    named_regions: Iterable[TableRegion] = (),
) -> SheetIR:
    """Build a bounded sparse IR using cached values for formula scoring."""
    max_row = min(worksheet.max_row, MAX_LAYOUT_ROWS)
    max_column = min(worksheet.max_column, MAX_LAYOUT_COLUMNS)
    cells: dict[tuple[int, int], CellIR] = {}

    for row in worksheet.iter_rows(
        min_row=1,
        max_row=max_row,
        min_col=1,
        max_col=max_column,
    ):
        for cell in row:
            if isinstance(cell, MergedCell) or _is_empty(cell.value):
                continue
            formula = (
                cell.value
                if isinstance(cell.value, str)
                and cell.value.startswith("=")
                else None
            )
            analysis_value = cell.value
            if formula is not None and value_worksheet is not None:
                cached_value = value_worksheet.cell(
                    row=cell.row,
                    column=cell.column,
                ).value
                if not _is_empty(cached_value):
                    analysis_value = cached_value
            cells[(cell.row, cell.column)] = CellIR(
                row=cell.row,
                column=cell.column,
                coordinate=cell.coordinate,
                value=analysis_value,
                number_format=cell.number_format,
                formula=formula,
            )

    merged_ranges, merged_anchors = _merged_anchor_map(worksheet)
    explicit = _worksheet_table_regions(worksheet)
    explicit.extend(named_regions)
    sheet = SheetIR(
        title=worksheet.title,
        max_row=max_row,
        max_column=max_column,
        cells=cells,
        merged_ranges=merged_ranges,
        merged_anchors=merged_anchors,
        explicit_regions=tuple(explicit),
    )
    sheet.regions = detect_table_regions(sheet)
    return sheet


def _named_regions_by_sheet(
    workbook: Workbook,
) -> dict[str, list[TableRegion]]:
    result: dict[str, list[TableRegion]] = {}
    for defined_name in workbook.defined_names.values():
        if defined_name.name.startswith("_xlnm."):
            continue
        try:
            destinations = list(defined_name.destinations)
        except (AttributeError, TypeError):
            continue
        for index, (sheet_name, reference) in enumerate(
            destinations,
            start=1,
        ):
            try:
                min_column, min_row, max_column, max_row = (
                    range_boundaries(reference.replace("$", ""))
                )
            except ValueError:
                continue
            result.setdefault(sheet_name, []).append(
                TableRegion(
                    region_id=f"named_range_{defined_name.name}_{index}",
                    min_row=min_row,
                    max_row=max_row,
                    min_column=min_column,
                    max_column=max_column,
                    discovery_method="named_range",
                )
            )
    return result


def build_workbook_ir(
    workbook: Workbook,
    value_workbook: Workbook | None = None,
) -> WorkbookIR:
    """Create layout IR, using cached formula values when available."""
    named_regions = _named_regions_by_sheet(workbook)
    return WorkbookIR(
        sheets={
            worksheet.title: build_sheet_ir(
                worksheet,
                value_worksheet=(
                    value_workbook[worksheet.title]
                    if value_workbook is not None
                    and worksheet.title in value_workbook.sheetnames
                    else None
                ),
                named_regions=named_regions.get(worksheet.title, ()),
            )
            for worksheet in workbook.worksheets
        }
    )
