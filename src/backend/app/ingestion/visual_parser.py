import json
import os
import re
import tempfile
import threading
import time
from collections import Counter
from datetime import date, datetime
from html import escape
from pathlib import Path
from statistics import median
from typing import Any

import pymupdf
from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter
from PIL import Image

from app.ingestion.schemas import (
    ColumnDataType,
    ExtractedCell,
    ExtractedTableRow,
    SheetContentInspection,
    SheetContentType,
    SourceCell,
    TableColumnSpec,
    TableDatasetSpec,
    TableMetadata,
    VisualDocumentContent,
    VisualPageContent,
    VisualTableContent,
    VisualTextBlock,
    WorkbookContentInspection,
    WorkbookTableExtraction,
)


MAX_VISUAL_PAGES = 20
PDF_RENDER_DPI = 200
HUMAN_REVIEW_CONFIDENCE = 0.80

_ENGINE_CACHE: dict[bool, Any] = {}
_ENGINE_CACHE_LOCK = threading.Lock()
_ENGINE_INFERENCE_LOCKS: dict[bool, threading.Lock] = {}


def _env_flag(
    name: str,
    default: bool = False,
) -> bool:
    value = os.getenv(
        name,
        "1" if default else "0",
    )

    return value.strip().lower() in {
        "1",
        "true",
        "yes",
        "on",
    }


def _get_engine(
    enable_chart_recognition: bool,
) -> Any:
    """
    延遲載入 OCR 模型，並保護 process-wide cache 初始化。

    PPStructureV3 的 predict 不保證可由多個 Job 同時安全呼叫；每種模型設定
    另有一把 inference lock，在允許多 Job 時仍維持單一 engine 序列推論。
    """
    with _ENGINE_CACHE_LOCK:
        cached = _ENGINE_CACHE.get(enable_chart_recognition)
        if cached is not None:
            return cached

        try:
            from paddleocr import PPStructureV3
        except ImportError as error:
            raise RuntimeError(
                "尚未安裝 PaddleOCR，請先執行 "
                'python -m pip install "paddleocr[all]"'
            ) from error

        engine = PPStructureV3(
            use_doc_orientation_classify=True,
            use_doc_unwarping=False,
            use_textline_orientation=True,
            use_formula_recognition=False,
            use_chart_recognition=enable_chart_recognition,
            format_block_content=True,
            device=os.getenv("OCR_DEVICE", "cpu"),
        )
        _ENGINE_CACHE[enable_chart_recognition] = engine
        _ENGINE_INFERENCE_LOCKS.setdefault(
            enable_chart_recognition,
            threading.Lock(),
        )
        return engine


def _result_to_dict(
    result: Any,
) -> dict[str, Any]:
    data = getattr(result, "json", {})

    if callable(data):
        data = data()

    if isinstance(data, str):
        data = json.loads(data)

    if not isinstance(data, dict):
        raise ValueError(
            "PaddleOCR 回傳格式不是 JSON object"
        )

    print(
        "[PaddleOCR DEBUG] top-level keys:",
        list(data.keys()),
    )

    nested_result = data.get("res")

    if isinstance(nested_result, dict):
        print(
            "[PaddleOCR DEBUG] res keys:",
            list(nested_result.keys()),
        )

        if any(
            key in nested_result
            for key in (
                "overall_ocr_res",
                "parsing_res_list",
                "table_res_list",
            )
        ):
            data = nested_result

    print(
        "[PaddleOCR DEBUG] final keys:",
        list(data.keys()),
    )

    return data

def _bbox_from_polygon(
    polygon: Any,
) -> list[float]:
    if polygon is None:
        return []

    try:
        points = list(polygon)

        xs = [
            float(point[0])
            for point in points
        ]

        ys = [
            float(point[1])
            for point in points
        ]

        return [
            min(xs),
            min(ys),
            max(xs),
            max(ys),
        ]

    except Exception:
        return []


def _normalize_scalar(
    value: Any,
) -> Any:
    if not isinstance(value, str):
        return value

    text = re.sub(
        r"\s+",
        " ",
        value.strip(),
    )

    if not text:
        return None

    if re.fullmatch(
        r"\([+-]?\d[\d,]*(?:\.\d+)?\)",
        text,
    ):
        number = float(
            text[1:-1].replace(",", "")
        )

        return (
            -int(number)
            if number.is_integer()
            else -number
        )

    if re.fullmatch(
        r"[+-]?\d[\d,]*(?:\.\d+)?",
        text,
    ):
        number = float(
            text.replace(",", "")
        )

        return (
            int(number)
            if number.is_integer()
            else number
        )

    if re.fullmatch(
        r"\d{4}-\d{1,2}-\d{1,2}",
        text,
    ):
        try:
            return date.fromisoformat(text)
        except ValueError:
            pass

    return text


def _infer_data_type(
    values: list[Any],
) -> ColumnDataType:
    non_empty_values = [
        value
        for value in values
        if value is not None
    ]

    if not non_empty_values:
        return ColumnDataType.EMPTY

    types: set[ColumnDataType] = set()

    for value in non_empty_values:
        if isinstance(value, bool):
            types.add(ColumnDataType.BOOLEAN)

        elif isinstance(value, datetime):
            types.add(ColumnDataType.DATETIME)

        elif isinstance(value, date):
            types.add(ColumnDataType.DATE)

        elif (
            isinstance(value, int)
            and not isinstance(value, bool)
        ):
            types.add(ColumnDataType.INTEGER)

        elif isinstance(value, float):
            types.add(ColumnDataType.NUMBER)

        else:
            types.add(ColumnDataType.STRING)

    if types == {ColumnDataType.INTEGER}:
        return ColumnDataType.INTEGER

    if types.issubset({
        ColumnDataType.INTEGER,
        ColumnDataType.NUMBER,
    }):
        return ColumnDataType.NUMBER

    if len(types) == 1:
        return next(iter(types))

    return ColumnDataType.MIXED


def _create_column_key(
    label: str,
    index: int,
    existing_keys: set[str],
) -> str:
    key = re.sub(
        r"[^\w]+",
        "_",
        label.strip().lower(),
        flags=re.UNICODE,
    ).strip("_")

    if not key:
        key = f"column_{index + 1}"

    candidate = key
    duplicate_index = 2

    while candidate in existing_keys:
        candidate = (
            f"{key}_{duplicate_index}"
        )

        duplicate_index += 1

    return candidate


def _html_table_rows(
    html: str,
) -> list[list[str]]:
    soup = BeautifulSoup(
        html,
        "html.parser",
    )

    table = soup.find("table")

    if table is None:
        return []

    rows: list[list[str]] = []

    for row_element in table.find_all("tr"):
        cells = row_element.find_all(
            ["th", "td"],
            recursive=False,
        )

        row = [
            re.sub(
                r"\s+",
                " ",
                cell.get_text(
                    " ",
                    strip=True,
                ),
            )
            for cell in cells
        ]

        if row:
            rows.append(row)

    if not rows:
        return []

    width = max(
        len(row)
        for row in rows
    )

    for row in rows:
        row.extend(
            [""] * (
                width - len(row)
            )
        )

    return rows



def _text_blocks_to_table_html(
    text_blocks: list[VisualTextBlock],
) -> tuple[str, float] | None:
    """
    當 PPStructureV3 沒有產生 table_res_list / pred_html 時，
    嘗試依 OCR 文字框的 Y/X 座標推測表格欄列。

    這是低信任度 fallback：只建立候選資料集，後續必須人工確認。
    """
    valid_blocks = [
        block
        for block in text_blocks
        if (
            block.text.strip()
            and len(block.bbox) == 4
        )
    ]

    if len(valid_blocks) < 4:
        return None

    heights = [
        max(
            1.0,
            float(block.bbox[3])
            - float(block.bbox[1]),
        )
        for block in valid_blocks
    ]

    # 同一列 OCR block 的中心 Y 容許誤差。
    y_tolerance = max(
        10.0,
        median(heights) * 0.7,
    )

    sorted_blocks = sorted(
        valid_blocks,
        key=lambda block: (
            (
                float(block.bbox[1])
                + float(block.bbox[3])
            )
            / 2,
            float(block.bbox[0]),
        ),
    )

    grouped_rows: list[
        tuple[float, list[VisualTextBlock]]
    ] = []

    # 1. 依 Y 座標分列。
    for block in sorted_blocks:
        center_y = (
            float(block.bbox[1])
            + float(block.bbox[3])
        ) / 2

        if not grouped_rows:
            grouped_rows.append(
                (
                    center_y,
                    [block],
                )
            )
            continue

        previous_y, previous_blocks = (
            grouped_rows[-1]
        )

        if (
            abs(center_y - previous_y)
            <= y_tolerance
        ):
            previous_blocks.append(block)

            new_y = sum(
                (
                    float(item.bbox[1])
                    + float(item.bbox[3])
                )
                / 2
                for item in previous_blocks
            ) / len(previous_blocks)

            grouped_rows[-1] = (
                new_y,
                previous_blocks,
            )
        else:
            grouped_rows.append(
                (
                    center_y,
                    [block],
                )
            )

    rows = [
        sorted(
            blocks,
            key=lambda block: float(
                block.bbox[0]
            ),
        )
        for _, blocks in grouped_rows
    ]

    if len(rows) < 2:
        return None

    # 2. 用最常出現的欄位數推測主要表格寬度。
    column_counts = [
        len(row)
        for row in rows
        if len(row) >= 2
    ]

    if len(column_counts) < 2:
        return None

    most_common_columns, most_common_count = (
        Counter(column_counts)
        .most_common(1)[0]
    )

    if most_common_columns < 2:
        return None

    structure_consistency = (
        most_common_count
        / len(column_counts)
    )

    # 規律性太低時，不要把一般段落文字誤判為表格。
    if structure_consistency < 0.5:
        return None

    reference_rows = [
        row
        for row in rows
        if len(row) == most_common_columns
    ]

    if len(reference_rows) < 2:
        return None

    # 3. 依規律列推測各欄的 X 中心位置。
    column_centers: list[float] = []

    for column_index in range(
        most_common_columns
    ):
        centers = [
            (
                float(row[column_index].bbox[0])
                + float(row[column_index].bbox[2])
            )
            / 2
            for row in reference_rows
        ]

        column_centers.append(
            float(median(centers))
        )

    # 4. 每個 OCR block 分配到最近的欄。
    normalized_rows: list[list[str]] = []

    for row in rows:
        cells: list[list[str]] = [
            []
            for _ in range(
                most_common_columns
            )
        ]

        for block in row:
            center_x = (
                float(block.bbox[0])
                + float(block.bbox[2])
            ) / 2

            nearest_column = min(
                range(
                    most_common_columns
                ),
                key=lambda index: abs(
                    center_x
                    - column_centers[index]
                ),
            )

            cells[nearest_column].append(
                block.text.strip()
            )

        normalized_row = [
            " ".join(cell).strip()
            for cell in cells
        ]

        # 至少兩欄有值，才視為資料列。
        if sum(
            bool(cell)
            for cell in normalized_row
        ) >= 2:
            normalized_rows.append(
                normalized_row
            )

    if len(normalized_rows) < 2:
        return None

    # 5. 轉成 HTML，直接重用既有 _html_to_dataset()。
    html_parts = ["<table>"]

    for row_index, row in enumerate(
        normalized_rows
    ):
        html_parts.append("<tr>")

        tag = (
            "th"
            if row_index == 0
            else "td"
        )

        for value in row:
            html_parts.append(
                f"<{tag}>"
                f"{escape(value)}"
                f"</{tag}>"
            )

        html_parts.append("</tr>")

    html_parts.append("</table>")

    confidences = [
        block.confidence
        for block in valid_blocks
        if block.confidence > 0
    ]

    average_ocr_confidence = (
        sum(confidences)
        / len(confidences)
        if confidences
        else 0.0
    )

    # 結構越規律、OCR 越可信，候選信心越高；
    # 但 fallback 永遠低於人工審核門檻。
    confidence = min(
        average_ocr_confidence
        * structure_consistency,
        HUMAN_REVIEW_CONFIDENCE - 0.01,
    )

    return (
        "".join(html_parts),
        max(0.0, confidence),
    )


def _html_to_dataset(
    filename: str,
    page_number: int,
    table_index: int,
    html: str,
    confidence: float,
) -> TableDatasetSpec | None:
    rows = _html_table_rows(html)

    if len(rows) < 2:
        return None

    header = rows[0]
    data_rows = rows[1:]

    virtual_sheet = (
        f"visual_page_{page_number}_"
        f"table_{table_index}"
    )

    existing_keys: set[str] = set()
    definitions: list[
        tuple[int, str, str]
    ] = []

    for column_index, raw_label in enumerate(
        header
    ):
        label = raw_label.strip()

        if not label:
            label = (
                f"未命名欄位 "
                f"{column_index + 1}"
            )

        key = _create_column_key(
            label,
            column_index,
            existing_keys,
        )

        existing_keys.add(key)

        definitions.append(
            (
                column_index,
                key,
                label,
            )
        )

    values_by_column: dict[
        str,
        list[Any],
    ] = {
        key: []
        for _, key, _ in definitions
    }

    empty_counts = {
        key: 0
        for _, key, _ in definitions
    }

    extracted_rows: list[
        ExtractedTableRow
    ] = []

    for row_index, raw_row in enumerate(
        data_rows,
        start=2,
    ):
        row_cells: dict[
            str,
            ExtractedCell,
        ] = {}

        row_has_value = False

        for (
            column_index,
            key,
            _,
        ) in definitions:
            raw_value = (
                raw_row[column_index]
                if column_index < len(raw_row)
                else ""
            )

            value = _normalize_scalar(
                raw_value
            )

            if value is None:
                empty_counts[key] += 1
            else:
                row_has_value = True

            values_by_column[key].append(
                value
            )

            coordinate = (
                f"{get_column_letter(column_index + 1)}"
                f"{row_index}"
            )

            row_cells[key] = ExtractedCell(
                raw_value=raw_value,
                value=value,
                formula=None,
                number_format=None,
                source=SourceCell(
                    sheet=virtual_sheet,
                    cell=coordinate,
                    row=row_index,
                    column=column_index + 1,
                ),
            )

        if row_has_value:
            extracted_rows.append(
                ExtractedTableRow(
                    excel_row=row_index,
                    cells=row_cells,
                )
            )

    columns: list[
        TableColumnSpec
    ] = []

    for column_index, key, label in definitions:
        coordinate = (
            f"{get_column_letter(column_index + 1)}1"
        )

        columns.append(
            TableColumnSpec(
                key=key,
                label=label,
                index=column_index,
                data_type=_infer_data_type(
                    values_by_column[key]
                ),
                unit=(
                    "%"
                    if "%" in label
                    else None
                ),
                nullable=(
                    empty_counts[key] > 0
                ),
                header_source=SourceCell(
                    sheet=virtual_sheet,
                    cell=coordinate,
                    row=1,
                    column=column_index + 1,
                ),
            )
        )

    last_column = get_column_letter(
        len(columns)
    )

    last_row = max(
        1,
        len(extracted_rows) + 1,
    )

    return TableDatasetSpec(
        filename=filename,
        sheet_name=virtual_sheet,
        table_kind=(
            SheetContentType
            .STRUCTURED_TABLE
        ),
        financial_statement_subtype=None,
        metadata=TableMetadata(
            title=None,
            entity=None,
            unit=None,
            notes=[],
        ),
        header_row=1,
        header_range=(
            f"{virtual_sheet}!A1:"
            f"{last_column}1"
        ),
        data_range=(
            f"{virtual_sheet}!A2:"
            f"{last_column}{last_row}"
            if extracted_rows
            else None
        ),
        full_range=(
            f"{virtual_sheet}!A1:"
            f"{last_column}{last_row}"
        ),
        row_count=len(extracted_rows),
        column_count=len(columns),
        columns=columns,
        rows=extracted_rows,
        warnings=[
            "此表格由圖片 OCR 與版面模型產生",
            (
                "視覺表格平均信心分數："
                f"{confidence:.2f}"
            ),
            "合併儲存格可能需要人工確認",
        ],
    )


def _prepare_page_images(
    file_path: Path,
    output_directory: Path,
    *,
    deadline_monotonic: float | None = None,
    max_pages: int = MAX_VISUAL_PAGES,
) -> tuple[
    list[tuple[int, Path]],
    int,
    list[str],
]:
    warnings: list[str] = []
    page_limit = max(1, min(MAX_VISUAL_PAGES, int(max_pages)))

    def expired() -> bool:
        return (
            deadline_monotonic is not None
            and time.monotonic() >= deadline_monotonic
        )

    if file_path.suffix.lower() == ".pdf":
        document = pymupdf.open(str(file_path))

        try:
            total_pages = len(document)
            planned_pages = min(total_pages, page_limit)

            if total_pages > page_limit:
                warnings.append(
                    f"掃描 PDF 共 {total_pages} 頁，只進行前 {page_limit} 頁"
                    "的視覺辨識"
                )

            page_images: list[tuple[int, Path]] = []
            for page_index in range(planned_pages):
                if expired():
                    warnings.append(
                        "OCR 整體時間上限已到；"
                        f"已完成 {len(page_images)}/{total_pages} 頁影像轉換"
                    )
                    break

                page = document[page_index]
                pixmap = page.get_pixmap(dpi=PDF_RENDER_DPI, alpha=False)
                image_path = output_directory / f"page_{page_index + 1}.png"
                pixmap.save(str(image_path))
                page_images.append((page_index + 1, image_path))

            return page_images, total_pages, warnings
        finally:
            document.close()

    try:
        with Image.open(file_path) as image:
            image.verify()
    except Exception as error:
        raise ValueError(f"圖片無法開啟：{error}") from error

    if expired():
        warnings.append("OCR 整體時間上限已到；圖片尚未進行視覺辨識")
        return [], 1, warnings

    return [(1, file_path)], 1, warnings


def _classify_page(
    has_text: bool,
    has_table: bool,
    has_chart: bool,
    has_image: bool,
) -> tuple[
    SheetContentType,
    list[SheetContentType],
]:
    components: list[
        SheetContentType
    ] = []

    if has_text:
        components.append(
            SheetContentType.DOCUMENT_TEXT
        )

    if has_table:
        components.append(
            SheetContentType.STRUCTURED_TABLE
        )

    if has_chart:
        components.append(
            SheetContentType.CHART_IMAGE
        )

    if has_image:
        components.append(
            SheetContentType.EMBEDDED_IMAGE
        )

    if has_chart and has_table:
        primary = (
            SheetContentType.MIXED_CONTENT
        )

    elif has_chart:
        primary = (
            SheetContentType.CHART_IMAGE
        )

    elif has_table:
        primary = (
            SheetContentType
            .STRUCTURED_TABLE
        )

    elif has_text:
        primary = (
            SheetContentType.DOCUMENT_TEXT
        )

    elif has_image:
        primary = (
            SheetContentType.EMBEDDED_IMAGE
        )

    else:
        primary = SheetContentType.UNKNOWN

    return primary, components


def inspect_visual_input(
    file_path: str | Path,
    original_filename: str | None = None,
    enable_chart_recognition: (
        bool | None
    ) = None,
    engine: Any | None = None,
    *,
    deadline_monotonic: float | None = None,
    max_pages: int = MAX_VISUAL_PAGES,
) -> tuple[
    WorkbookContentInspection,
    VisualDocumentContent,
    WorkbookTableExtraction,
]:
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(
            f"找不到視覺檔案：{path}"
        )

    filename = (
        original_filename or path.name
    )

    if enable_chart_recognition is None:
        enable_chart_recognition = _env_flag(
            "ENABLE_CHART_RECOGNITION",
            default=False,
        )

    shared_engine = engine is None

    visual_pages: list[
        VisualPageContent
    ] = []

    classifications: list[
        SheetContentInspection
    ] = []

    extracted_tables: list[
        TableDatasetSpec
    ] = []

    skipped_pages: list[str] = []
    warnings: list[str] = []

    with tempfile.TemporaryDirectory() as temp:
        page_images, total_pages, render_warnings = (
            _prepare_page_images(
                path,
                Path(temp),
                deadline_monotonic=deadline_monotonic,
                max_pages=max_pages,
            )
        )

        warnings.extend(render_warnings)
        if page_images and engine is None:
            engine = _get_engine(enable_chart_recognition)

        inference_lock = (
            _ENGINE_INFERENCE_LOCKS.get(enable_chart_recognition)
            if shared_engine
            else None
        )

        for page_number, image_path in page_images:
            if (
                deadline_monotonic is not None
                and time.monotonic() >= deadline_monotonic
            ):
                warnings.append(
                    "OCR 整體時間上限已到；"
                    f"已完成 {len(visual_pages)}/{total_pages} 頁視覺辨識"
                )
                break

            assert engine is not None
            if inference_lock is None:
                prediction = engine.predict(str(image_path))
            else:
                with inference_lock:
                    prediction = engine.predict(str(image_path))

            result_items = list(
                prediction
            )

            if not result_items:
                warnings.append(
                    f"第 {page_number} 頁"
                    "沒有取得視覺辨識結果"
                )

                skipped_pages.append(
                    f"page_{page_number}"
                )

                continue

            payload = _result_to_dict(
                result_items[0]
            )

            width = int(
                payload.get("width", 0) or 0
            )

            height = int(
                payload.get("height", 0) or 0
            )

            ocr_result = (
                payload.get(
                    "overall_ocr_res",
                    {},
                )
                or {}
            )

            texts = list(
                ocr_result.get(
                    "rec_texts",
                    [],
                )
                or []
            )

            scores = [
                float(score)
                for score in (
                    ocr_result.get(
                        "rec_scores",
                        [],
                    )
                    or []
                )
            ]

            polygons = list(
                ocr_result.get(
                    "rec_polys",
                    [],
                )
                or []
            )

            text_blocks: list[
                VisualTextBlock
            ] = []

            for index, text in enumerate(texts):
                confidence = (
                    scores[index]
                    if index < len(scores)
                    else 0.0
                )

                polygon = (
                    polygons[index]
                    if index < len(polygons)
                    else None
                )

                text_blocks.append(
                    VisualTextBlock(
                        text=str(text),
                        confidence=confidence,
                        bbox=_bbox_from_polygon(
                            polygon
                        ),
                        label="text",
                    )
                )

            average_confidence = (
                sum(scores) / len(scores)
                if scores
                else 0.0
            )

            parsing_blocks = list(
                payload.get(
                    "parsing_res_list",
                    [],
                )
                or []
            )

            labels = [
                str(
                    block.get(
                        "block_label",
                        "",
                    )
                ).lower()
                for block in parsing_blocks
            ]

            chart_blocks = [
                block
                for block in parsing_blocks
                if "chart" in str(
                    block.get(
                        "block_label",
                        "",
                    )
                ).lower()
            ]

            chart_contents = [
                str(
                    block.get(
                        "block_content",
                        "",
                    )
                ).strip()
                for block in chart_blocks
                if str(
                    block.get(
                        "block_content",
                        "",
                    )
                ).strip()
            ]

            table_results = list(
                payload.get(
                    "table_res_list",
                    [],
                )
                or []
            )

            visual_tables: list[
                VisualTableContent
            ] = []

            extracted_on_page = 0
            fallback_used = False

            for table_index, table_result in enumerate(
                table_results,
                start=1,
            ):
                html = table_result.get(
                    "pred_html"
                )

                table_ocr = (
                    table_result.get(
                        "table_ocr_pred",
                        {},
                    )
                    or {}
                )

                table_scores = [
                    float(score)
                    for score in (
                        table_ocr.get(
                            "rec_scores",
                            [],
                        )
                        or []
                    )
                ]

                table_confidence = (
                    sum(table_scores)
                    / len(table_scores)
                    if table_scores
                    else average_confidence
                )

                table_rows = (
                    _html_table_rows(html)
                    if html
                    else []
                )

                visual_tables.append(
                    VisualTableContent(
                        page_number=page_number,
                        table_index=table_index,
                        html=html,
                        row_count=len(
                            table_rows
                        ),
                        column_count=(
                            max(
                                (
                                    len(row)
                                    for row
                                    in table_rows
                                ),
                                default=0,
                            )
                        ),
                        confidence=round(
                            table_confidence,
                            4,
                        ),
                        bbox=[],
                    )
                )

                if html:
                    dataset = _html_to_dataset(
                        filename=filename,
                        page_number=page_number,
                        table_index=table_index,
                        html=html,
                        confidence=(
                            table_confidence
                        ),
                    )

                    if dataset is not None:
                        extracted_tables.append(
                            dataset
                        )

                        extracted_on_page += 1

            # PPStructureV3 沒有產出可用 pred_html 時，
            # 從 OCR 文字框座標建立低信任度候選表格。
            if (
                extracted_on_page == 0
                and text_blocks
            ):
                fallback_result = (
                    _text_blocks_to_table_html(
                        text_blocks
                    )
                )

                if fallback_result is not None:
                    (
                        fallback_html,
                        fallback_confidence,
                    ) = fallback_result

                    fallback_table_index = (
                        len(table_results) + 1
                    )

                    fallback_dataset = (
                        _html_to_dataset(
                            filename=filename,
                            page_number=page_number,
                            table_index=(
                                fallback_table_index
                            ),
                            html=fallback_html,
                            confidence=(
                                fallback_confidence
                            ),
                        )
                    )

                    if fallback_dataset is not None:
                        fallback_dataset.warnings.append(
                            "此資料表由 OCR 文字座標"
                            "推測欄列結構，必須經人工確認"
                        )

                        extracted_tables.append(
                            fallback_dataset
                        )

                        visual_rows = (
                            _html_table_rows(
                                fallback_html
                            )
                        )

                        visual_tables.append(
                            VisualTableContent(
                                page_number=page_number,
                                table_index=(
                                    fallback_table_index
                                ),
                                html=fallback_html,
                                row_count=len(
                                    visual_rows
                                ),
                                column_count=max(
                                    (
                                        len(row)
                                        for row
                                        in visual_rows
                                    ),
                                    default=0,
                                ),
                                confidence=round(
                                    fallback_confidence,
                                    4,
                                ),
                                bbox=[],
                            )
                        )

                        extracted_on_page += 1
                        fallback_used = True

            has_text = bool(
                "\n".join(texts).strip()
            )

            has_table = (
                bool(table_results)
                or "table" in labels
                or fallback_used
            )

            has_chart = bool(chart_blocks)

            has_image = any(
                label in {
                    "image",
                    "figure",
                    "picture",
                }
                for label in labels
            )

            (
                primary_type,
                components,
            ) = _classify_page(
                has_text=has_text,
                has_table=has_table,
                has_chart=has_chart,
                has_image=has_image,
            )

            page_warnings: list[str] = []

            requires_human_review = (
                average_confidence
                < HUMAN_REVIEW_CONFIDENCE
                or has_chart
                or fallback_used
                or (
                    has_table
                    and extracted_on_page == 0
                )
            )

            if (
                average_confidence
                < HUMAN_REVIEW_CONFIDENCE
            ):
                page_warnings.append(
                    "OCR 平均信心分數低於 0.80"
                )

            if has_chart:
                page_warnings.append(
                    "圖表截圖的數值需人工確認，"
                    "不得只依長條高度或折線位置"
                    "直接視為正確數值"
                )

            if fallback_used:
                page_warnings.append(
                    "PPStructureV3 未直接產生表格，"
                    "目前資料由 OCR 文字座標推測欄列結構，"
                    "需人工確認"
                )

            if (
                has_table
                and extracted_on_page == 0
            ):
                page_warnings.append(
                    "偵測到表格區域，"
                    "但無法建立結構化資料表"
                )

            visual_pages.append(
                VisualPageContent(
                    page_number=page_number,
                    width=width,
                    height=height,
                    primary_content_type=(
                        primary_type
                    ),
                    components=components,
                    ocr_text="\n".join(texts),
                    average_ocr_confidence=round(
                        average_confidence,
                        4,
                    ),
                    text_blocks=text_blocks,
                    tables=visual_tables,
                    chart_count=len(
                        chart_blocks
                    ),
                    chart_contents=(
                        chart_contents
                    ),
                    likely_scanned=True,
                    requires_human_review=(
                        requires_human_review
                    ),
                    warnings=page_warnings,
                )
            )

            classifications.append(
                SheetContentInspection(
                    sheet_name=(
                        f"page_{page_number}"
                    ),
                    sheet_state="visible",
                    primary_content_type=(
                        primary_type
                    ),
                    components=components,
                    financial_statement_subtype=None,
                    confidence=round(
                        average_confidence,
                        4,
                    ),
                    max_row=len(texts),
                    max_column=1,
                    non_empty_cells=len(texts),
                    text_cells=len(texts),
                    numeric_cells=0,
                    formula_cells=0,
                    date_cells=0,
                    merged_range_count=0,
                    chart_count=len(
                        chart_blocks
                    ),
                    image_count=(
                        1 if has_image else 0
                    ),
                    detected_header_row=None,
                    structured_table_score=(
                        0.9
                        if has_table
                        else 0.0
                    ),
                    financial_statement_score=0.0,
                    evidence=[
                        f"OCR 文字區塊：{len(texts)}",
                        (
                            "OCR 平均信心："
                            f"{average_confidence:.2f}"
                        ),
                        (
                            "視覺表格數："
                            f"{len(table_results)}"
                        ),
                        (
                            "圖表區塊數："
                            f"{len(chart_blocks)}"
                        ),
                    ],
                    warnings=page_warnings,
                )
            )

            if extracted_on_page == 0:
                skipped_pages.append(
                    f"page_{page_number}"
                )

    meaningful_types = [
        result.primary_content_type
        for result in classifications
    ]

    if not meaningful_types:
        overall_type = (
            SheetContentType.UNKNOWN
        )

    elif len(set(meaningful_types)) == 1:
        overall_type = meaningful_types[0]

    else:
        overall_type = (
            SheetContentType.MIXED_CONTENT
        )

    confidence_values = [
        page.average_ocr_confidence
        for page in visual_pages
    ]

    document_confidence = (
        sum(confidence_values)
        / len(confidence_values)
        if confidence_values
        else 0.0
    )

    requires_review = any(
        page.requires_human_review
        for page in visual_pages
    )

    visual_document = VisualDocumentContent(
        filename=filename,
        engine="PaddleOCR PPStructureV3",
        page_count=total_pages,
        processed_page_count=len(
            visual_pages
        ),
        average_ocr_confidence=round(
            document_confidence,
            4,
        ),
        chart_recognition_enabled=(
            enable_chart_recognition
        ),
        requires_human_review=(
            requires_review
        ),
        pages=visual_pages,
        warnings=warnings.copy(),
    )

    classification = (
        WorkbookContentInspection(
            filename=filename,
            sheet_count=len(
                classifications
            ),
            overall_content_type=(
                overall_type
            ),
            confidence=round(
                document_confidence,
                4,
            ),
            sheets=classifications,
            warnings=warnings.copy(),
        )
    )

    extraction = WorkbookTableExtraction(
        filename=filename,
        table_count=len(
            extracted_tables
        ),
        tables=extracted_tables,
        skipped_sheets=skipped_pages,
        warnings=warnings.copy(),
    )

    return (
        classification,
        visual_document,
        extraction,
    )