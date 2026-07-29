import csv
import io
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl.utils import get_column_letter

from app.ingestion.schemas import (
    ColumnDataType,
    ExtractedCell,
    ExtractedTableRow,
    SheetContentInspection,
    SheetContentType,
    SourceCell,
    TableColumnSpec,
    TableDatasetSpec,
    TableMetadata,
    WorkbookContentInspection,
    WorkbookTableExtraction,
)


SUPPORTED_ENCODINGS = (
    "utf-8-sig",
    "utf-8",
    "cp950",
    "big5",
    "utf-16",
)

SUPPORTED_DELIMITERS = ",\t;|"

HEADER_SCAN_ROWS = 20
BLANK_ROWS_TO_STOP = 3
MAX_TEXT_ROWS = 100_000
MAX_TEXT_COLUMNS = 500


UNIT_PATTERN = re.compile(
    r"(?:單位|unit)\s*[:：]\s*(.+)",
    flags=re.IGNORECASE,
)

NUMBER_PATTERN = re.compile(
    r"^[+-]?\d[\d,]*(?:\.\d+)?$"
)

NEGATIVE_NUMBER_PATTERN = re.compile(
    r"^\(([+-]?\d[\d,]*(?:\.\d+)?)\)$"
)

ISO_DATE_PATTERN = re.compile(
    r"^\d{4}-\d{1,2}-\d{1,2}$"
)

ISO_DATETIME_PATTERN = re.compile(
    r"^\d{4}-\d{1,2}-\d{1,2}"
    r"[T ]\d{1,2}:\d{2}"
    r"(?::\d{2})?$"
)


def _is_empty(value: Any) -> bool:
    return value is None or (
        isinstance(value, str)
        and not value.strip()
    )


def _decode_text_file(
    file_path: str | Path,
) -> tuple[str, str]:
    """
    嘗試使用常見中文與 Unicode 編碼讀取文字檔。

    回傳：
    - 解碼後文字
    - 使用的編碼
    """
    path = Path(file_path)
    data = path.read_bytes()

    if not data:
        raise ValueError("文字檔案是空的")

    for encoding in SUPPORTED_ENCODINGS:
        try:
            text = data.decode(encoding)
            return text, encoding
        except UnicodeDecodeError:
            continue

    raise ValueError(
        "無法辨識文字編碼，目前支援 "
        "UTF-8、UTF-16、CP950 與 Big5"
    )


def _detect_delimiter(text: str) -> str:
    """
    偵測逗號、Tab、分號或直線分隔符號。
    """
    sample = text[:65_536]

    try:
        dialect = csv.Sniffer().sniff(
            sample,
            delimiters=SUPPORTED_DELIMITERS,
        )

        return dialect.delimiter

    except csv.Error:
        lines = [
            line
            for line in sample.splitlines()[:20]
            if line.strip()
        ]

        if not lines:
            raise ValueError("文字內容沒有可讀資料")

        delimiter_scores = {
            delimiter: sum(
                line.count(delimiter)
                for line in lines
            )
            for delimiter in SUPPORTED_DELIMITERS
        }

        delimiter = max(
            delimiter_scores,
            key=delimiter_scores.get,
        )

        if delimiter_scores[delimiter] == 0:
            raise ValueError(
                "找不到逗號、Tab、分號或 | 分隔符號"
            )

        return delimiter


def _read_delimited_rows(
    file_path: str | Path,
) -> tuple[list[list[str]], str, str]:
    """
    讀取分隔文字檔。

    回傳：
    - 二維資料列
    - 編碼
    - 分隔符號
    """
    text, encoding = _decode_text_file(file_path)
    delimiter = _detect_delimiter(text)

    reader = csv.reader(
        io.StringIO(text),
        delimiter=delimiter,
    )

    rows: list[list[str]] = []

    for row_number, row in enumerate(
        reader,
        start=1,
    ):
        if row_number > MAX_TEXT_ROWS:
            raise ValueError(
                f"文字表格超過 {MAX_TEXT_ROWS} 列限制"
            )

        if len(row) > MAX_TEXT_COLUMNS:
            raise ValueError(
                f"第 {row_number} 列超過 "
                f"{MAX_TEXT_COLUMNS} 欄限制"
            )

        cleaned_row = [
            value.strip()
            for value in row
        ]

        while (
            cleaned_row
            and not cleaned_row[-1]
        ):
            cleaned_row.pop()

        rows.append(cleaned_row)

    while rows and not any(rows[-1]):
        rows.pop()

    return rows, encoding, delimiter


def _normalize_scalar(value: Any) -> Any:
    """
    將文字轉成適合的 Python 型態。

    例如：
    1,200       → 1200
    (1,200)     → -1200
    2026-07-26  → date
    true        → True
    """
    if not isinstance(value, str):
        return value

    text = value.strip()

    if not text:
        return None

    lowered = text.lower()

    if lowered in {"true", "yes", "是"}:
        return True

    if lowered in {"false", "no", "否"}:
        return False

    negative_match = (
        NEGATIVE_NUMBER_PATTERN.fullmatch(text)
    )

    if negative_match:
        numeric_text = (
            negative_match.group(1)
            .replace(",", "")
        )

        number = float(numeric_text)

        if number.is_integer():
            return -int(number)

        return -number

    if NUMBER_PATTERN.fullmatch(text):
        number = float(
            text.replace(",", "")
        )

        if number.is_integer():
            return int(number)

        return number

    if ISO_DATETIME_PATTERN.fullmatch(text):
        try:
            return datetime.fromisoformat(
                text.replace(" ", "T")
            )
        except ValueError:
            pass

    if ISO_DATE_PATTERN.fullmatch(text):
        try:
            return date.fromisoformat(text)
        except ValueError:
            pass

    return text


def _count_non_empty(
    row: list[Any],
) -> int:
    return sum(
        not _is_empty(value)
        for value in row
    )


def _get_row_value(
    rows: list[list[str]],
    row_index: int,
    column_index: int,
) -> str:
    if row_index >= len(rows):
        return ""

    row = rows[row_index]

    if column_index >= len(row):
        return ""

    return row[column_index]


def _detect_header_row(
    rows: list[list[str]],
) -> int | None:
    """
    找出最可能的表頭列。

    回傳 0-based index。
    """
    candidates: list[
        tuple[int, float]
    ] = []

    scan_end = min(
        len(rows),
        HEADER_SCAN_ROWS,
    )

    for row_index in range(scan_end):
        row = rows[row_index]

        non_empty_values = [
            value
            for value in row
            if not _is_empty(value)
        ]

        if len(non_empty_values) < 2:
            continue

        normalized_values = [
            _normalize_scalar(value)
            for value in non_empty_values
        ]

        text_count = sum(
            isinstance(value, str)
            for value in normalized_values
        )

        text_ratio = (
            text_count
            / len(normalized_values)
        )

        unique_ratio = (
            len(set(non_empty_values))
            / len(non_empty_values)
        )

        minimum_data_cells = max(
            1,
            (len(non_empty_values) + 1) // 2,
        )

        following_rows = 0

        for data_index in range(
            row_index + 1,
            min(row_index + 5, len(rows)),
        ):
            if (
                _count_non_empty(rows[data_index])
                >= minimum_data_cells
            ):
                following_rows += 1

        if following_rows < 2:
            continue

        score = (
            text_ratio * 0.40
            + unique_ratio * 0.20
            + min(
                following_rows / 3,
                1,
            ) * 0.30
            + max(
                0,
                0.10 - row_index * 0.01,
            )
        )

        candidates.append(
            (row_index, score)
        )

    if not candidates:
        return None

    candidates.sort(
        key=lambda item: item[1],
        reverse=True,
    )

    return candidates[0][0]


def _calculate_table_score(
    rows: list[list[str]],
    header_index: int | None,
) -> tuple[float, int]:
    if header_index is None:
        return 0.0, 0

    header = rows[header_index]

    header_width = _count_non_empty(
        header
    )

    if header_width < 2:
        return 0.0, 0

    minimum_data_cells = max(
        1,
        (header_width + 1) // 2,
    )

    data_like_rows = 0
    comparable_width_rows = 0

    for row in rows[
        header_index + 1:
        header_index + 201
    ]:
        non_empty_count = _count_non_empty(
            row
        )

        if non_empty_count >= minimum_data_cells:
            data_like_rows += 1

        if (
            abs(non_empty_count - header_width)
            <= max(1, header_width // 3)
        ):
            comparable_width_rows += 1

    score = 0.40

    if data_like_rows >= 2:
        score += 0.30
    elif data_like_rows == 1:
        score += 0.10

    if comparable_width_rows >= 2:
        score += 0.20

    numeric_count = sum(
        isinstance(
            _normalize_scalar(value),
            (int, float),
        )
        and not isinstance(
            _normalize_scalar(value),
            bool,
        )
        for row in rows[
            header_index + 1:
            header_index + 21
        ]
        for value in row
        if not _is_empty(value)
    )

    if numeric_count > 0:
        score += 0.10

    return min(score, 1.0), data_like_rows


def _delimiter_name(
    delimiter: str,
) -> str:
    names = {
        ",": "comma",
        "\t": "tab",
        ";": "semicolon",
        "|": "pipe",
    }

    return names.get(
        delimiter,
        repr(delimiter),
    )


def inspect_delimited_content(
    file_path: str | Path,
) -> WorkbookContentInspection:
    """
    將 CSV／TSV／TXT 當作只有一張工作表的資料來源。
    """
    path = Path(file_path)

    rows, encoding, delimiter = (
        _read_delimited_rows(path)
    )

    max_column = max(
        (len(row) for row in rows),
        default=0,
    )

    non_empty_cells = 0
    text_cells = 0
    numeric_cells = 0
    date_cells = 0

    for row in rows:
        for raw_value in row:
            if _is_empty(raw_value):
                continue

            non_empty_cells += 1
            value = _normalize_scalar(
                raw_value
            )

            if isinstance(value, datetime):
                date_cells += 1
            elif isinstance(value, date):
                date_cells += 1
            elif (
                isinstance(value, (int, float))
                and not isinstance(value, bool)
            ):
                numeric_cells += 1
            else:
                text_cells += 1

    header_index = _detect_header_row(rows)

    (
        structured_score,
        data_like_rows,
    ) = _calculate_table_score(
        rows,
        header_index,
    )

    evidence = [
        f"文字編碼：{encoding}",
        (
            "分隔符號："
            f"{_delimiter_name(delimiter)}"
        ),
    ]

    warnings: list[str] = []
    components: list[SheetContentType] = []

    if not rows or non_empty_cells == 0:
        primary_type = SheetContentType.EMPTY
        confidence = 1.0

    elif (
        header_index is not None
        and structured_score >= 0.60
    ):
        primary_type = (
            SheetContentType.STRUCTURED_TABLE
        )

        components.append(
            SheetContentType.STRUCTURED_TABLE
        )

        confidence = structured_score

        evidence.append(
            f"第 {header_index + 1} 列可能是表頭"
        )

        evidence.append(
            f"偵測到約 {data_like_rows} 列資料"
        )

    else:
        primary_type = SheetContentType.UNKNOWN
        confidence = 0.30

        warnings.append(
            "文字內容存在，但找不到可信的表格結構"
        )

    sheet_result = SheetContentInspection(
        sheet_name="data",
        sheet_state="visible",
        primary_content_type=primary_type,
        components=components,
        financial_statement_subtype=None,
        confidence=round(confidence, 4),
        max_row=len(rows),
        max_column=max_column,
        non_empty_cells=non_empty_cells,
        text_cells=text_cells,
        numeric_cells=numeric_cells,
        formula_cells=0,
        date_cells=date_cells,
        merged_range_count=0,
        chart_count=0,
        image_count=0,
        detected_header_row=(
            header_index + 1
            if header_index is not None
            else None
        ),
        structured_table_score=round(
            structured_score,
            4,
        ),
        financial_statement_score=0.0,
        evidence=evidence,
        warnings=warnings,
    )

    return WorkbookContentInspection(
        filename=path.name,
        sheet_count=1,
        overall_content_type=primary_type,
        confidence=round(confidence, 4),
        sheets=[sheet_result],
        warnings=warnings.copy(),
    )


def _create_column_key(
    label: str,
    index: int,
    existing_keys: set[str],
) -> str:
    normalized = re.sub(
        r"[^\w]+",
        "_",
        label.strip().lower(),
        flags=re.UNICODE,
    ).strip("_")

    if not normalized:
        normalized = f"column_{index + 1}"

    candidate = normalized
    duplicate_number = 2

    while candidate in existing_keys:
        candidate = (
            f"{normalized}_{duplicate_number}"
        )

        duplicate_number += 1

    return candidate


def _infer_data_type(
    values: list[Any],
) -> ColumnDataType:
    non_empty_values = [
        value
        for value in values
        if value is not None
    ]

    if not non_empty_values:
        return ColumnDataType.EMPTY

    detected_types: set[
        ColumnDataType
    ] = set()

    for value in non_empty_values:
        if isinstance(value, bool):
            detected_types.add(
                ColumnDataType.BOOLEAN
            )

        elif isinstance(value, datetime):
            detected_types.add(
                ColumnDataType.DATETIME
            )

        elif isinstance(value, date):
            detected_types.add(
                ColumnDataType.DATE
            )

        elif (
            isinstance(value, int)
            and not isinstance(value, bool)
        ):
            detected_types.add(
                ColumnDataType.INTEGER
            )

        elif isinstance(value, float):
            detected_types.add(
                ColumnDataType.NUMBER
            )

        else:
            detected_types.add(
                ColumnDataType.STRING
            )

    if detected_types == {
        ColumnDataType.INTEGER
    }:
        return ColumnDataType.INTEGER

    if detected_types.issubset({
        ColumnDataType.INTEGER,
        ColumnDataType.NUMBER,
    }):
        return ColumnDataType.NUMBER

    if len(detected_types) == 1:
        return next(iter(detected_types))

    return ColumnDataType.MIXED


def _extract_unit_from_header(
    label: str,
) -> str | None:
    match = re.search(
        r"[（(]\s*([^（）()]+?)\s*[）)]",
        label,
    )

    if match:
        return match.group(1).strip()

    if "%" in label:
        return "%"

    return None


def _extract_metadata(
    rows: list[list[str]],
    header_index: int,
    min_column: int,
    max_column: int,
) -> TableMetadata:
    title: str | None = None
    unit: str | None = None
    notes: list[str] = []

    for row in rows[:header_index]:
        texts = [
            value.strip()
            for value in row[
                min_column:
                max_column + 1
            ]
            if value.strip()
        ]

        if not texts:
            continue

        combined_text = " ".join(texts)

        unit_match = UNIT_PATTERN.search(
            combined_text
        )

        if unit_match and unit is None:
            unit = unit_match.group(1).strip()
            continue

        if title is None:
            title = combined_text
        else:
            notes.append(combined_text)

    return TableMetadata(
        title=title,
        entity=None,
        unit=unit,
        notes=notes,
    )


def extract_delimited_table(
    file_path: str | Path,
    sheet_name: str | None = None,
) -> WorkbookTableExtraction:
    """
    抽取 CSV／TSV／TXT 表格。

    為了和 Excel 共用 Schema，
    虛擬工作表名稱固定為 data。
    """
    path = Path(file_path)

    if (
        sheet_name is not None
        and sheet_name != "data"
    ):
        raise ValueError(
            "分隔文字檔只有虛擬工作表 data"
        )

    inspection = inspect_delimited_content(
        path
    )

    sheet_inspection = inspection.sheets[0]

    if (
        sheet_inspection.primary_content_type
        != SheetContentType.STRUCTURED_TABLE
    ):
        return WorkbookTableExtraction(
            filename=path.name,
            table_count=0,
            tables=[],
            skipped_sheets=["data"],
            warnings=[
                "文字檔無法辨識為結構化表格"
            ],
        )

    rows, encoding, delimiter = (
        _read_delimited_rows(path)
    )

    header_index = (
        sheet_inspection.detected_header_row
    )

    if header_index is None:
        raise ValueError("找不到表頭列")

    header_index -= 1

    header_row = rows[header_index]

    non_empty_header_columns = [
        index
        for index, value in enumerate(
            header_row
        )
        if not _is_empty(value)
    ]

    if not non_empty_header_columns:
        raise ValueError("表頭列沒有欄位")

    min_column = min(
        non_empty_header_columns
    )

    max_column = max(
        non_empty_header_columns
    )

    metadata = _extract_metadata(
        rows=rows,
        header_index=header_index,
        min_column=min_column,
        max_column=max_column,
    )

    existing_keys: set[str] = set()

    header_definitions: list[
        tuple[int, str, str, str | None]
    ] = []

    for relative_index, column_index in enumerate(
        range(
            min_column,
            max_column + 1,
        )
    ):
        label = _get_row_value(
            rows,
            header_index,
            column_index,
        ).strip()

        if not label:
            label = (
                f"未命名欄位 "
                f"{relative_index + 1}"
            )

        key = _create_column_key(
            label=label,
            index=relative_index,
            existing_keys=existing_keys,
        )

        existing_keys.add(key)

        unit = (
            _extract_unit_from_header(label)
            or metadata.unit
        )

        header_definitions.append(
            (
                column_index,
                key,
                label,
                unit,
            )
        )

    extracted_rows: list[
        ExtractedTableRow
    ] = []

    column_values: dict[
        str,
        list[Any],
    ] = {
        key: []
        for _, key, _, _
        in header_definitions
    }

    column_empty_counts = {
        key: 0
        for _, key, _, _
        in header_definitions
    }

    blank_rows = 0

    for row_index in range(
        header_index + 1,
        len(rows),
    ):
        raw_row = rows[row_index]

        selected_values = [
            _get_row_value(
                rows,
                row_index,
                column_index,
            )
            for column_index, _, _, _
            in header_definitions
        ]

        if _count_non_empty(
            selected_values
        ) == 0:
            blank_rows += 1

            if blank_rows >= BLANK_ROWS_TO_STOP:
                break

            continue

        blank_rows = 0

        row_cells: dict[
            str,
            ExtractedCell,
        ] = {}

        for (
            column_index,
            key,
            _,
            _,
        ) in header_definitions:
            raw_value = _get_row_value(
                rows,
                row_index,
                column_index,
            )

            normalized_value = (
                _normalize_scalar(raw_value)
            )

            if normalized_value is None:
                column_empty_counts[key] += 1

            column_values[key].append(
                normalized_value
            )

            excel_column = get_column_letter(
                column_index + 1
            )

            coordinate = (
                f"{excel_column}"
                f"{row_index + 1}"
            )

            row_cells[key] = ExtractedCell(
                raw_value=raw_value,
                value=normalized_value,
                formula=None,
                number_format=None,
                source=SourceCell(
                    sheet="data",
                    cell=coordinate,
                    row=row_index + 1,
                    column=column_index + 1,
                ),
            )

        extracted_rows.append(
            ExtractedTableRow(
                excel_row=row_index + 1,
                cells=row_cells,
            )
        )

    columns: list[
        TableColumnSpec
    ] = []

    for relative_index, (
        column_index,
        key,
        label,
        unit,
    ) in enumerate(header_definitions):
        coordinate = (
            f"{get_column_letter(column_index + 1)}"
            f"{header_index + 1}"
        )

        columns.append(
            TableColumnSpec(
                key=key,
                label=label,
                index=relative_index,
                data_type=_infer_data_type(
                    column_values[key]
                ),
                unit=unit,
                nullable=(
                    column_empty_counts[key] > 0
                ),
                header_source=SourceCell(
                    sheet="data",
                    cell=coordinate,
                    row=header_index + 1,
                    column=column_index + 1,
                ),
            )
        )

    first_column_letter = get_column_letter(
        min_column + 1
    )

    last_column_letter = get_column_letter(
        max_column + 1
    )

    header_excel_row = header_index + 1

    if extracted_rows:
        last_data_row = (
            extracted_rows[-1].excel_row
        )

        data_range = (
            f"data!{first_column_letter}"
            f"{header_excel_row + 1}:"
            f"{last_column_letter}"
            f"{last_data_row}"
        )

        full_end_row = last_data_row
    else:
        data_range = None
        full_end_row = header_excel_row

    table = TableDatasetSpec(
        filename=path.name,
        sheet_name="data",
        table_kind=(
            SheetContentType.STRUCTURED_TABLE
        ),
        financial_statement_subtype=None,
        metadata=metadata,
        header_row=header_excel_row,
        header_range=(
            f"data!{first_column_letter}"
            f"{header_excel_row}:"
            f"{last_column_letter}"
            f"{header_excel_row}"
        ),
        data_range=data_range,
        full_range=(
            f"data!{first_column_letter}"
            f"{header_excel_row}:"
            f"{last_column_letter}"
            f"{full_end_row}"
        ),
        row_count=len(extracted_rows),
        column_count=len(columns),
        columns=columns,
        rows=extracted_rows,
        warnings=[
            f"文字編碼：{encoding}",
            (
                "分隔符號："
                f"{_delimiter_name(delimiter)}"
            ),
        ],
    )

    return WorkbookTableExtraction(
        filename=path.name,
        table_count=1,
        tables=[table],
        skipped_sheets=[],
        warnings=[],
    )