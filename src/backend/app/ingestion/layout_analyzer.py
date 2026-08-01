"""Deterministic Excel layout strategies and confidence scoring."""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any, Callable

from openpyxl.utils import get_column_letter

from app.ingestion.classifier import is_period_like_header
from app.ingestion.layout_graph import (
    build_logical_table_hypotheses,
    optimize_logical_plan,
)
from app.ingestion.normalization_spec import (
    LayoutStrategy,
    NormalizationSpec,
    NormalizedMetadataSpec,
    OutputColumnSpec,
    PeriodValueGroup,
    SourceCoordinateSpec,
    TableRegionSpec,
)
from app.ingestion.normalization_validator import (
    validate_normalization_plan,
)
from app.ingestion.workbook_ir import SheetIR, TableRegion


NUMBER_PATTERN = re.compile(r"^[+-]?\d[\d,]*(?:\.\d+)?$")
NEGATIVE_PATTERN = re.compile(r"^\([+-]?\d[\d,]*(?:\.\d+)?\)$")
UNIT_PATTERN = re.compile(
    r"(?:單位|unit)\s*[:：]\s*(.+)",
    flags=re.IGNORECASE,
)
THOUSANDS_PATTERN = re.compile(
    r"\bin\s+thousands\b",
    flags=re.IGNORECASE,
)
CURRENCY_MARKERS = {"$", "US$", "USD", "NT$", "€", "£", "¥"}
EMPTY_MARKERS = {"", "-", "--", "—", "–", "N/A", "NA", "不適用"}


@dataclass(frozen=True)
class LayoutCandidate:
    spec: NormalizationSpec
    score: float


def _text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def _is_empty(value: Any) -> bool:
    return value is None or _text(value) in EMPTY_MARKERS


def _is_numeric(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return True
    if not isinstance(value, str):
        return False
    text = value.strip()
    return bool(
        NUMBER_PATTERN.fullmatch(text)
        or NEGATIVE_PATTERN.fullmatch(text)
    )


def _is_period(value: Any) -> bool:
    if is_period_like_header(value) or isinstance(value, (date, datetime)):
        return True
    if not isinstance(value, str):
        return False
    compact = value.strip().replace("年", "").replace("月", "")
    if not compact.isdigit():
        return False
    try:
        return is_period_like_header(int(compact))
    except ValueError:
        return False


def _period_label(value: Any) -> str:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return _text(value)


def _source(row: int, column: int) -> SourceCoordinateSpec:
    return SourceCoordinateSpec(
        row=row,
        column=column,
        cell=f"{get_column_letter(column)}{row}",
    )


def _region_spec(region: TableRegion) -> TableRegionSpec:
    return TableRegionSpec(
        region_id=region.region_id,
        min_row=region.min_row,
        max_row=region.max_row,
        min_column=region.min_column,
        max_column=region.max_column,
        discovery_method=region.discovery_method,
    )


def _row_values(
    sheet: SheetIR,
    region: TableRegion,
    row: int,
    *,
    resolve_merged: bool = False,
) -> list[Any]:
    return [
        sheet.value_at(
            row,
            column,
            resolve_merged=resolve_merged,
        )
        for column in range(region.min_column, region.max_column + 1)
    ]


def _metadata(
    sheet: SheetIR,
    region: TableRegion,
    first_header_row: int,
    scan_end_row: int | None = None,
) -> tuple[NormalizedMetadataSpec, list[str]]:
    title: str | None = None
    unit: str | None = None
    notes: list[str] = []
    warnings: list[str] = []
    metadata_rows = list(range(region.min_row, first_header_row))
    if scan_end_row is not None:
        metadata_rows.extend(
            range(
                first_header_row + 1,
                min(scan_end_row, region.max_row + 1),
            )
        )

    for row in metadata_rows:
        texts = [
            _text(sheet.value_at(row, column, resolve_merged=False))
            for column in range(region.min_column, region.max_column + 1)
        ]
        texts = [text for text in texts if text]
        if not texts:
            continue
        combined = " ".join(texts)
        match = UNIT_PATTERN.search(combined)
        if match and unit is None:
            unit = match.group(1).strip()
            continue
        if THOUSANDS_PATTERN.search(combined) and unit is None:
            unit = combined.strip("() ")
            if "except" in combined.lower():
                warnings.append(
                    "人工確認：單位說明包含例外項目，已完整保留原始單位說明"
                )
            if combined not in notes:
                notes.append(combined)
            continue
        if title is None:
            title = combined
        else:
            notes.append(combined)

    marker_values = {
        _text(cell.value)
        for (row, column), cell in sheet.cells.items()
        if region.contains(row, column)
    }
    if unit is not None and marker_values.intersection(CURRENCY_MARKERS):
        if not any(marker in unit for marker in CURRENCY_MARKERS):
            unit = f"$; {unit}"

    return (
        NormalizedMetadataSpec(title=title, unit=unit, notes=notes),
        warnings,
    )


def _unit_from_label(label: str) -> str | None:
    match = re.search(r"[（(]\s*([^（）()]+?)\s*[）)]", label)
    if match:
        return match.group(1).strip()
    if "%" in label:
        return "%"
    return None


def _numeric_coverage(
    sheet: SheetIR,
    rows: list[int],
    columns: list[int],
) -> float:
    if not rows:
        return 0.0
    hits = 0
    for row in rows:
        if any(_is_numeric(sheet.value_at(row, column)) for column in columns):
            hits += 1
    return hits / len(rows)


def _candidate_data_rows(
    sheet: SheetIR,
    region: TableRegion,
    start_row: int,
) -> list[int]:
    return [
        row
        for row in range(start_row, region.max_row + 1)
        if any(
            not _is_empty(sheet.value_at(row, column))
            for column in range(region.min_column, region.max_column + 1)
        )
    ]


def _period_axis_candidate(
    sheet: SheetIR,
    region: TableRegion,
) -> LayoutCandidate | None:
    period_rows: list[tuple[int, list[int]]] = []
    scan_end = min(region.max_row, region.min_row + 29)
    for row in range(region.min_row, scan_end + 1):
        anchors = [
            column
            for column in range(region.min_column, region.max_column + 1)
            if _is_period(
                sheet.value_at(row, column, resolve_merged=False)
            )
        ]
        if len(anchors) >= 2:
            labels = [
                _period_label(
                    sheet.value_at(
                        row,
                        column,
                        resolve_merged=False,
                    )
                )
                for column in anchors
            ]
            # Repeated years usually form a hierarchical header such as
            # ``2025 / Revenue, Profit``. Leave those rows to the multi-row
            # strategy so the measure labels are not discarded.
            if len(set(labels)) == len(labels):
                period_rows.append((row, anchors))

    if not period_rows:
        return None

    best: LayoutCandidate | None = None
    for header_row, anchors in period_rows:
        first_anchor = min(anchors)
        label_candidates = range(region.min_column, first_anchor)
        if not label_candidates:
            continue

        raw_rows = _candidate_data_rows(sheet, region, header_row + 1)
        label_column = max(
            label_candidates,
            key=lambda column: sum(
                bool(_text(sheet.value_at(row, column)))
                and not _is_numeric(sheet.value_at(row, column))
                for row in raw_rows
            ),
        )

        groups: list[tuple[int, str, list[int]]] = []
        sorted_anchors = sorted(anchors)
        for index, anchor in enumerate(sorted_anchors):
            next_anchor = (
                sorted_anchors[index + 1]
                if index + 1 < len(sorted_anchors)
                else region.max_column + 1
            )
            candidate_columns = list(range(anchor, next_anchor))
            groups.append(
                (
                    anchor,
                    _period_label(
                        sheet.value_at(
                            header_row,
                            anchor,
                            resolve_merged=False,
                        )
                    ),
                    candidate_columns,
                )
            )

        data_rows = [
            row
            for row in raw_rows
            if _text(sheet.value_at(row, label_column))
            and any(
                any(
                    _is_numeric(sheet.value_at(row, column))
                    for column in columns
                )
                for _, _, columns in groups
            )
        ]
        if len(data_rows) < 2:
            continue

        metadata, warnings = _metadata(
            sheet,
            region,
            header_row,
            scan_end_row=min(data_rows),
        )
        label_header = _text(sheet.value_at(header_row, label_column))
        if not label_header or _is_period(label_header):
            label_header = "項目"

        output_columns = [
            OutputColumnSpec(
                label=label_header,
                source_columns=[label_column],
                header_sources=[_source(header_row, label_column)],
                selection_rule="first_non_empty",
                semantic_role="dimension",
            )
        ]
        value_groups: list[PeriodValueGroup] = []
        marker_pair = False

        for anchor, label, candidate_columns in groups:
            markers = sum(
                _text(sheet.value_at(row, column)) in CURRENCY_MARKERS
                for row in data_rows
                for column in candidate_columns
            )
            if markers:
                marker_pair = True
            output_columns.append(
                OutputColumnSpec(
                    label=label,
                    source_columns=candidate_columns,
                    header_sources=[_source(header_row, anchor)],
                    selection_rule="first_numeric",
                    semantic_role="period_measure",
                    unit=metadata.unit,
                )
            )
            value_groups.append(
                PeriodValueGroup(
                    label=label,
                    candidate_columns=candidate_columns,
                    header_source=_source(header_row, anchor),
                    unit=metadata.unit,
                )
            )

        coverage = sum(
            _numeric_coverage(sheet, data_rows, columns)
            for _, _, columns in groups
        ) / len(groups)
        label_coverage = sum(
            bool(_text(sheet.value_at(row, label_column)))
            for row in data_rows
        ) / len(data_rows)
        score = min(
            0.99,
            0.58
            + min(len(groups), 4) * 0.045
            + coverage * 0.16
            + label_coverage * 0.07,
        )
        strategy = (
            LayoutStrategy.MARKER_VALUE_PAIR
            if marker_pair
            else LayoutStrategy.PERIOD_AXIS
        )
        transformations = ["period_axis_to_columns"]
        if marker_pair:
            transformations.append("marker_value_pair_resolution")

        spec = NormalizationSpec(
            sheet_name=sheet.title,
            region=_region_spec(region),
            strategy=strategy,
            header_rows=[header_row],
            data_rows=data_rows,
            output_columns=output_columns,
            period_value_groups=value_groups,
            metadata=metadata,
            transformations=transformations,
            confidence=round(score, 4),
            confidence_reasons=[
                f"辨識到 {len(groups)} 個期間群組",
                f"期間數值覆蓋率 {coverage:.0%}",
                f"列標籤覆蓋率 {label_coverage:.0%}",
            ],
            warnings=warnings,
        )
        candidate = LayoutCandidate(spec=spec, score=score)
        if best is None or candidate.score > best.score:
            best = candidate

    return best


def _combined_header_label(
    sheet: SheetIR,
    rows: list[int],
    column: int,
) -> tuple[str, list[SourceCoordinateSpec]]:
    labels: list[str] = []
    sources: list[SourceCoordinateSpec] = []
    seen: set[tuple[int, int]] = set()
    for row in rows:
        cell = sheet.cell_at(row, column, resolve_merged=True)
        if cell is None or (cell.row, cell.column) in seen:
            continue
        seen.add((cell.row, cell.column))
        label = _text(cell.value)
        if label and label not in labels:
            labels.append(label)
            sources.append(_source(cell.row, cell.column))
    return " ".join(labels), sources


def _multi_row_candidate(
    sheet: SheetIR,
    region: TableRegion,
) -> LayoutCandidate | None:
    scan_end = min(region.max_row - 2, region.min_row + 28)
    best: LayoutCandidate | None = None
    for top_row in range(region.min_row, scan_end + 1):
        bottom_row = top_row + 1
        data_rows = _candidate_data_rows(sheet, region, bottom_row + 1)
        if len(data_rows) < 2:
            continue

        labels: list[tuple[int, str, list[SourceCoordinateSpec]]] = []
        for column in range(region.min_column, region.max_column + 1):
            label, sources = _combined_header_label(
                sheet,
                [top_row, bottom_row],
                column,
            )
            if label:
                labels.append((column, label, sources))
        if len(labels) < 2:
            continue

        lower_count = sum(
            not _is_empty(sheet.value_at(bottom_row, column))
            for column, _, _ in labels
        )
        upper_count = sum(
            not _is_empty(sheet.value_at(top_row, column))
            for column, _, _ in labels
        )
        has_merged_header = any(
            sheet.merged_anchors.get((top_row, column), (top_row, column))
            != (top_row, column)
            for column, _, _ in labels
        )
        top_period_labels = [
            _period_label(sheet.value_at(top_row, column))
            for column, _, _ in labels
            if _is_period(sheet.value_at(top_row, column))
        ]
        has_repeated_period_header = (
            len(top_period_labels) >= 2
            and len(set(top_period_labels)) < len(top_period_labels)
        )
        if not has_merged_header and (
            upper_count < 1 or lower_count < 2
        ):
            continue

        coverage = sum(
            _numeric_coverage(sheet, data_rows, [column])
            for column, _, _ in labels[1:]
        ) / max(1, len(labels) - 1)
        if coverage < 0.4:
            continue

        metadata, warnings = _metadata(sheet, region, top_row)
        output_columns = [
            OutputColumnSpec(
                label=label,
                source_columns=[column],
                header_sources=sources,
                selection_rule="first_non_empty",
                semantic_role=("dimension" if index == 0 else "measure"),
                unit=(
                    _unit_from_label(label)
                    or (metadata.unit if index > 0 else None)
                ),
            )
            for index, (column, label, sources) in enumerate(labels)
        ]
        score = min(
            0.98,
            0.67
            + coverage * 0.2
            + (0.10 if has_merged_header else 0.0)
            + (0.10 if has_repeated_period_header else 0.0),
        )
        spec = NormalizationSpec(
            sheet_name=sheet.title,
            region=_region_spec(region),
            strategy=LayoutStrategy.MULTI_ROW_HEADER,
            header_rows=[top_row, bottom_row],
            data_rows=data_rows,
            output_columns=output_columns,
            metadata=metadata,
            transformations=["merge_header_rows"],
            confidence=round(score, 4),
            confidence_reasons=[
                "相鄰兩列表頭具有互補欄名",
                f"數值欄覆蓋率 {coverage:.0%}",
            ],
            warnings=warnings,
        )
        candidate = LayoutCandidate(spec=spec, score=score)
        if best is None or candidate.score > best.score:
            best = candidate
    return best


def _flat_candidate(
    sheet: SheetIR,
    region: TableRegion,
) -> LayoutCandidate | None:
    scan_end = min(region.max_row - 2, region.min_row + 29)
    best: LayoutCandidate | None = None
    for header_row in range(region.min_row, scan_end + 1):
        header_cells = [
            (column, sheet.value_at(header_row, column, resolve_merged=False))
            for column in range(region.min_column, region.max_column + 1)
        ]
        header_cells = [
            (column, value)
            for column, value in header_cells
            if not _is_empty(value)
        ]
        if len(header_cells) < 2:
            continue

        period_count = sum(
            _is_period(value)
            for _, value in header_cells
        )
        # A row containing multiple periods is a cross-tab axis, not a flat
        # header. Let the period strategy preserve the label column and map
        # sparse marker/value groups instead of treating the periods as the
        # only columns.
        if period_count >= 2:
            continue

        plausible = sum(
            isinstance(value, str) or _is_period(value)
            for _, value in header_cells
        )
        plausible_ratio = plausible / len(header_cells)
        if plausible_ratio < 0.5:
            continue

        data_rows = _candidate_data_rows(sheet, region, header_row + 1)
        if len(data_rows) < 2:
            continue
        minimum_cells = max(1, (len(header_cells) + 1) // 2)
        valid_rows = [
            row
            for row in data_rows
            if sum(
                not _is_empty(sheet.value_at(row, column))
                for column, _ in header_cells
            )
            >= minimum_cells
        ]
        if len(valid_rows) < 2:
            continue

        metadata, warnings = _metadata(sheet, region, header_row)
        output_columns: list[OutputColumnSpec] = []
        for index, (column, raw_label) in enumerate(header_cells):
            label = _text(raw_label) or f"未命名欄位 {index + 1}"
            output_columns.append(
                OutputColumnSpec(
                    label=label,
                    source_columns=[column],
                    header_sources=[_source(header_row, column)],
                    selection_rule="first_non_empty",
                    semantic_role=("dimension" if index == 0 else "measure"),
                    unit=_unit_from_label(label) or metadata.unit,
                )
            )

        row_coverage = len(valid_rows) / len(data_rows)
        numeric_data_cells = sum(
            _is_numeric(sheet.value_at(row, column))
            for row in valid_rows
            for column, _ in header_cells
        )
        score = min(
            0.97,
            0.55
            + plausible_ratio * 0.22
            + row_coverage * 0.13
            + min(len(valid_rows), 5) * 0.014,
        )
        if numeric_data_cells == 0:
            # Text-only lists are still extractable, but lack enough
            # structural evidence for unattended numeric presentation.
            score = min(score, 0.89)
        spec = NormalizationSpec(
            sheet_name=sheet.title,
            region=_region_spec(region),
            strategy=LayoutStrategy.FLAT_TABLE,
            header_rows=[header_row],
            data_rows=valid_rows,
            output_columns=output_columns,
            metadata=metadata,
            transformations=["trim_empty_rows_and_columns"],
            confidence=round(score, 4),
            confidence_reasons=[
                f"表頭可解釋比例 {plausible_ratio:.0%}",
                f"資料列覆蓋率 {row_coverage:.0%}",
            ],
            warnings=warnings,
        )
        candidate = LayoutCandidate(spec=spec, score=score)
        if best is None or candidate.score > best.score:
            best = candidate
    return best


STRATEGIES: tuple[
    Callable[[SheetIR, TableRegion], LayoutCandidate | None], ...
] = (
    _period_axis_candidate,
    _multi_row_candidate,
    _flat_candidate,
)


def analyze_sheet_layouts(sheet: SheetIR) -> list[NormalizationSpec]:
    """Optimize worksheet-level hypotheses together with regional plans."""
    _, hypotheses = build_logical_table_hypotheses(sheet)
    regional_specs: list[NormalizationSpec] = []
    for region in sheet.regions:
        candidates = [
            candidate
            for strategy in STRATEGIES
            if (candidate := strategy(sheet, region)) is not None
        ]
        if not candidates:
            continue
        priority = {
            LayoutStrategy.MARKER_VALUE_PAIR: 3,
            LayoutStrategy.PERIOD_AXIS: 3,
            LayoutStrategy.MULTI_ROW_HEADER: 2,
            LayoutStrategy.CROSS_TAB: 2,
            LayoutStrategy.FLAT_TABLE: 1,
        }
        period_candidates = [
            candidate
            for candidate in candidates
            if candidate.spec.strategy
            in {
                LayoutStrategy.PERIOD_AXIS,
                LayoutStrategy.MARKER_VALUE_PAIR,
            }
        ]
        selectable = period_candidates or candidates
        selected = max(
            selectable,
            key=lambda candidate: (
                candidate.score,
                priority[candidate.spec.strategy],
            ),
        )
        regional_specs.append(selected.spec)

    optimized = optimize_logical_plan(
        hypotheses,
        regional_specs,
    )
    validated, _ = validate_normalization_plan(
        sheet,
        optimized,
    )
    return validated
