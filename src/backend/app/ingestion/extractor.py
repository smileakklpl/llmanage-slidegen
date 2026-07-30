import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.ingestion.classifier import (
    inspect_excel_content,
    is_period_like_header,
)
from app.ingestion.schemas import (
    ColumnDataType,
    ExtractedCell,
    ExtractedTableRow,
    SheetContentInspection,
    SheetContentType,
    SourceCell,
    TableColumnSpec,
    TableDatasetSpec,
    TableMetadata,
    WorkbookTableExtraction,
)


MAX_EXTRACT_ROWS = 10_000
MAX_EXTRACT_COLUMNS = 200
HEADER_SCAN_ROWS = 30
BLANK_ROWS_TO_STOP = 3


UNIT_PATTERN = re.compile(
    r"(?:單位|unit)\s*[:：]\s*(.+)",
    flags=re.IGNORECASE,
)


NUMBER_PATTERN = re.compile(
    r"^[+-]?\d[\d,]*(?:\.\d+)?$"
)


NEGATIVE_PARENTHESES_PATTERN = re.compile(
    r"^\(([+-]?\d[\d,]*(?:\.\d+)?)\)$"
)


def _is_empty(value: Any) -> bool:
    """判斷儲存格是否為空。"""
    return value is None or (
        isinstance(value, str)
        and not value.strip()
    )


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""

    return re.sub(
        r"\s+",
        " ",
        str(value).strip(),
    )


def _normalize_scalar(value: Any) -> Any:
    """
    將常見的文字數字轉為真正數字。

    例如：
    "1,234"   → 1234
    "(1,234)" → -1234
    """
    if not isinstance(value, str):
        return value

    text = value.strip()

    if not text:
        return None

    negative_match = NEGATIVE_PARENTHESES_PATTERN.fullmatch(
        text
    )

    if negative_match:
        numeric_text = (
            negative_match
            .group(1)
            .replace(",", "")
        )

        numeric_value = float(numeric_text)

        if numeric_value.is_integer():
            return -int(numeric_value)

        return -numeric_value

    if NUMBER_PATTERN.fullmatch(text):
        numeric_text = text.replace(",", "")
        numeric_value = float(numeric_text)

        if numeric_value.is_integer():
            return int(numeric_value)

        return numeric_value

    return text


def _row_values(
    worksheet: Worksheet,
    row_number: int,
    min_column: int,
    max_column: int,
) -> list[Any]:
    return [
        worksheet.cell(
            row=row_number,
            column=column_number,
        ).value
        for column_number in range(
            min_column,
            max_column + 1,
        )
    ]


def _count_non_empty(values: list[Any]) -> int:
    return sum(
        not _is_empty(value)
        for value in values
    )


def _detect_actual_bounds(
    worksheet: Worksheet,
) -> tuple[int, int, int, int] | None:
    """
    找到工作表真正有內容的範圍。

    回傳：
    min_row, max_row, min_column, max_column
    """
    max_row = min(
        worksheet.max_row,
        MAX_EXTRACT_ROWS,
    )

    max_column = min(
        worksheet.max_column,
        MAX_EXTRACT_COLUMNS,
    )

    occupied_rows: list[int] = []
    occupied_columns: list[int] = []

    for row in worksheet.iter_rows(
        min_row=1,
        max_row=max_row,
        min_col=1,
        max_col=max_column,
    ):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue

            if _is_empty(cell.value):
                continue

            occupied_rows.append(cell.row)
            occupied_columns.append(cell.column)

    if not occupied_rows or not occupied_columns:
        return None

    return (
        min(occupied_rows),
        max(occupied_rows),
        min(occupied_columns),
        max(occupied_columns),
    )


def _header_value_is_plausible(value: Any) -> bool:
    """
    判斷某個值是否可能是表頭文字。

    表頭也可能是年份、日期或月份，
    所以不限制只能是字串。

    期間型欄名的判斷與 classifier 共用
    :func:`~app.ingestion.classifier.is_period_like_header`，
    避免兩層對「什麼算表頭」有不同看法——分類器認得、抽取器卻不認，
    工作表就會被分類成有表格卻抽不出東西。
    """
    if _is_empty(value):
        return False

    if isinstance(value, str):
        return not value.startswith("=")

    return is_period_like_header(value)


def _detect_header_row(
    worksheet: Worksheet,
    bounds: tuple[int, int, int, int],
) -> int | None:
    """
    尋找最可能的表頭列。

    依據：
    1. 該列至少有兩個非空值。
    2. 多數值看起來像欄位名稱。
    3. 後面至少有兩列資料。
    """
    (
        min_row,
        max_row,
        min_column,
        max_column,
    ) = bounds

    scan_end_row = min(
        max_row,
        min_row + HEADER_SCAN_ROWS - 1,
    )

    candidates: list[tuple[int, float]] = []

    for row_number in range(
        min_row,
        scan_end_row + 1,
    ):
        values = _row_values(
            worksheet,
            row_number,
            min_column,
            max_column,
        )

        non_empty_values = [
            value
            for value in values
            if not _is_empty(value)
        ]

        if len(non_empty_values) < 2:
            continue

        plausible_headers = sum(
            _header_value_is_plausible(value)
            for value in non_empty_values
        )

        plausible_ratio = (
            plausible_headers
            / len(non_empty_values)
        )

        unique_labels = {
            _normalize_text(value)
            for value in non_empty_values
        }

        unique_ratio = (
            len(unique_labels)
            / len(non_empty_values)
        )

        following_rows = 0

        # 依表頭實際欄位數動態決定資料列最低非空值數量。
        # 例如兩欄表格允許其中一欄缺失，仍可視為有效資料列。
        minimum_data_cells = max(
            1,
            (len(non_empty_values) + 1) // 2,
        )

        for data_row in range(
            row_number + 1,
            min(row_number + 5, max_row + 1),
        ):
            data_values = _row_values(
                worksheet,
                data_row,
                min_column,
                max_column,
            )

            if (
                _count_non_empty(data_values)
                >= minimum_data_cells
            ):
                following_rows += 1

        if following_rows < 2:
            continue

        score = (
            min(
                len(non_empty_values) / 5,
                1,
            ) * 0.30
            + plausible_ratio * 0.30
            + unique_ratio * 0.15
            + min(
                following_rows / 3,
                1,
            ) * 0.25
        )

        candidates.append(
            (row_number, score)
        )

    if not candidates:
        return None

    candidates.sort(
        key=lambda item: item[1],
        reverse=True,
    )

    return candidates[0][0]


def _detect_column_bounds(
    worksheet: Worksheet,
    header_row: int,
    overall_bounds: tuple[int, int, int, int],
) -> tuple[int, int] | None:
    """
    根據表頭列找出表格左右邊界。
    """
    (
        _,
        _,
        overall_min_column,
        overall_max_column,
    ) = overall_bounds

    header_columns = [
        column_number
        for column_number in range(
            overall_min_column,
            overall_max_column + 1,
        )
        if not _is_empty(
            worksheet.cell(
                row=header_row,
                column=column_number,
            ).value
        )
    ]

    if not header_columns:
        return None

    return (
        min(header_columns),
        max(header_columns),
    )


def _detect_data_end_row(
    worksheet: Worksheet,
    header_row: int,
    max_row: int,
    min_column: int,
    max_column: int,
) -> int:
    """
    找出資料最後一列。

    遇到連續三列完全空白時停止，
    避免把遠端備註錯當成資料。
    """
    last_non_empty_row = header_row
    consecutive_blank_rows = 0

    for row_number in range(
        header_row + 1,
        max_row + 1,
    ):
        values = _row_values(
            worksheet,
            row_number,
            min_column,
            max_column,
        )

        if _count_non_empty(values) == 0:
            consecutive_blank_rows += 1

            if (
                consecutive_blank_rows
                >= BLANK_ROWS_TO_STOP
            ):
                break

            continue

        consecutive_blank_rows = 0
        last_non_empty_row = row_number

    return last_non_empty_row


def _extract_unit_from_header(
    label: str,
) -> str | None:
    """
    嘗試從欄名擷取單位。

    例如：
    市占率（%） → %
    營收（千元） → 千元
    """
    bracket_match = re.search(
        r"[（(]\s*([^（）()]+?)\s*[）)]",
        label,
    )

    if bracket_match:
        return bracket_match.group(1).strip()

    if "%" in label:
        return "%"

    return None


def _extract_metadata(
    worksheet: Worksheet,
    header_row: int,
    min_column: int,
    max_column: int,
) -> TableMetadata:
    """
    從表頭前面的列擷取：

    - 標題
    - 單位
    - 備註
    """
    title: str | None = None
    unit: str | None = None
    notes: list[str] = []

    for row_number in range(
        1,
        header_row,
    ):
        row_texts = []

        for column_number in range(
            min_column,
            max_column + 1,
        ):
            value = worksheet.cell(
                row=row_number,
                column=column_number,
            ).value

            if _is_empty(value):
                continue

            if isinstance(value, str):
                text = _normalize_text(value)

                if text:
                    row_texts.append(text)

        if not row_texts:
            continue

        combined_text = " ".join(row_texts)

        unit_match = UNIT_PATTERN.search(
            combined_text
        )

        if unit_match and unit is None:
            unit = (
                unit_match
                .group(1)
                .strip()
            )
            continue

        if title is None:
            title = combined_text
        else:
            notes.append(combined_text)

    return TableMetadata(
        title=title,
        entity=None,
        unit=unit,
        notes=notes,
    )


def _create_column_key(
    label: str,
    index: int,
    existing_keys: set[str],
) -> str:
    """
    將欄位名稱轉成穩定 key。

    中文會被保留，例如：
    北區營收 → 北區營收

    重複欄位會變成：
    營收
    營收_2
    """
    normalized = re.sub(
        r"[^\w]+",
        "_",
        label.strip().lower(),
        flags=re.UNICODE,
    ).strip("_")

    if not normalized:
        normalized = f"column_{index + 1}"

    candidate = normalized
    duplicate_index = 2

    while candidate in existing_keys:
        candidate = (
            f"{normalized}_{duplicate_index}"
        )
        duplicate_index += 1

    return candidate


def _infer_data_type(
    values: list[Any],
) -> ColumnDataType:
    non_empty_values = [
        value
        for value in values
        if value is not None
    ]

    if not non_empty_values:
        return ColumnDataType.EMPTY

    detected_types: set[ColumnDataType] = set()

    for value in non_empty_values:
        if isinstance(value, bool):
            detected_types.add(
                ColumnDataType.BOOLEAN
            )

        elif isinstance(value, datetime):
            detected_types.add(
                ColumnDataType.DATETIME
            )

        elif isinstance(value, date):
            detected_types.add(
                ColumnDataType.DATE
            )

        elif (
            isinstance(value, int)
            and not isinstance(value, bool)
        ):
            detected_types.add(
                ColumnDataType.INTEGER
            )

        elif isinstance(value, float):
            detected_types.add(
                ColumnDataType.NUMBER
            )

        else:
            detected_types.add(
                ColumnDataType.STRING
            )

    if detected_types == {
        ColumnDataType.INTEGER
    }:
        return ColumnDataType.INTEGER

    if detected_types.issubset({
        ColumnDataType.INTEGER,
        ColumnDataType.NUMBER,
    }):
        return ColumnDataType.NUMBER

    if len(detected_types) == 1:
        return next(iter(detected_types))

    return ColumnDataType.MIXED


def _extract_single_sheet(
    filename: str,
    formula_worksheet: Worksheet,
    value_worksheet: Worksheet,
    classification: SheetContentInspection,
) -> TableDatasetSpec:
    warnings: list[str] = []

    bounds = _detect_actual_bounds(
        formula_worksheet
    )

    if bounds is None:
        raise ValueError(
            f"工作表 {formula_worksheet.title} 沒有資料"
        )

    (
        _,
        overall_max_row,
        _,
        _,
    ) = bounds

    header_row = _detect_header_row(
        formula_worksheet,
        bounds,
    )

    if header_row is None:
        raise ValueError(
            f"工作表 {formula_worksheet.title} "
            "找不到可信的表頭列"
        )

    column_bounds = _detect_column_bounds(
        formula_worksheet,
        header_row,
        bounds,
    )

    if column_bounds is None:
        raise ValueError(
            f"工作表 {formula_worksheet.title} "
            "找不到表格欄位"
        )

    min_column, max_column = column_bounds

    data_end_row = _detect_data_end_row(
        worksheet=formula_worksheet,
        header_row=header_row,
        max_row=overall_max_row,
        min_column=min_column,
        max_column=max_column,
    )

    metadata = _extract_metadata(
        worksheet=formula_worksheet,
        header_row=header_row,
        min_column=min_column,
        max_column=max_column,
    )

    existing_keys: set[str] = set()
    header_definitions: list[
        tuple[int, str, str, str | None]
    ] = []

    for index, column_number in enumerate(
        range(
            min_column,
            max_column + 1,
        )
    ):
        raw_label = formula_worksheet.cell(
            row=header_row,
            column=column_number,
        ).value

        label = _normalize_text(raw_label)

        if not label:
            label = f"未命名欄位 {index + 1}"
            warnings.append(
                f"{get_column_letter(column_number)}"
                f"{header_row} 沒有欄位名稱"
            )

        key = _create_column_key(
            label=label,
            index=index,
            existing_keys=existing_keys,
        )

        existing_keys.add(key)

        column_unit = (
            _extract_unit_from_header(label)
            or metadata.unit
        )

        header_definitions.append(
            (
                column_number,
                key,
                label,
                column_unit,
            )
        )

    extracted_rows: list[
        ExtractedTableRow
    ] = []

    column_values: dict[str, list[Any]] = {
        key: []
        for _, key, _, _ in header_definitions
    }

    column_empty_counts: dict[str, int] = {
        key: 0
        for _, key, _, _ in header_definitions
    }

    missing_formula_cache_cells: list[str] = []

    for row_number in range(
        header_row + 1,
        data_end_row + 1,
    ):
        row_cells: dict[str, ExtractedCell] = {}
        row_has_value = False

        for (
            column_number,
            key,
            _,
            _,
        ) in header_definitions:
            formula_cell = formula_worksheet.cell(
                row=row_number,
                column=column_number,
            )

            value_cell = value_worksheet.cell(
                row=row_number,
                column=column_number,
            )

            formula: str | None = None

            if (
                isinstance(formula_cell.value, str)
                and formula_cell.value.startswith("=")
            ):
                formula = formula_cell.value
                raw_value = value_cell.value

                if raw_value is None:
                    missing_formula_cache_cells.append(
                        formula_cell.coordinate
                    )

            else:
                raw_value = formula_cell.value

            normalized_value = _normalize_scalar(
                raw_value
            )

            if normalized_value is not None:
                row_has_value = True
            else:
                column_empty_counts[key] += 1

            column_values[key].append(
                normalized_value
            )

            row_cells[key] = ExtractedCell(
                raw_value=raw_value,
                value=normalized_value,
                formula=formula,
                number_format=(
                    formula_cell.number_format
                ),
                source=SourceCell(
                    sheet=formula_worksheet.title,
                    cell=formula_cell.coordinate,
                    row=row_number,
                    column=column_number,
                ),
            )

        if not row_has_value:
            continue

        extracted_rows.append(
            ExtractedTableRow(
                excel_row=row_number,
                cells=row_cells,
            )
        )

    columns: list[TableColumnSpec] = []

    for index, (
        column_number,
        key,
        label,
        unit,
    ) in enumerate(header_definitions):
        columns.append(
            TableColumnSpec(
                key=key,
                label=label,
                index=index,
                data_type=_infer_data_type(
                    column_values[key]
                ),
                unit=unit,
                nullable=(
                    column_empty_counts[key] > 0
                ),
                header_source=SourceCell(
                    sheet=formula_worksheet.title,
                    cell=(
                        f"{get_column_letter(column_number)}"
                        f"{header_row}"
                    ),
                    row=header_row,
                    column=column_number,
                ),
            )
        )

    if missing_formula_cache_cells:
        warnings.append(
            "部分公式沒有儲存的計算結果："
            + "、".join(
                missing_formula_cache_cells[:10]
            )
        )

    header_start = (
        f"{get_column_letter(min_column)}"
        f"{header_row}"
    )

    header_end = (
        f"{get_column_letter(max_column)}"
        f"{header_row}"
    )

    full_start = header_start
    full_end = (
        f"{get_column_letter(max_column)}"
        f"{data_end_row}"
    )

    if data_end_row > header_row:
        data_range = (
            f"{formula_worksheet.title}!"
            f"{get_column_letter(min_column)}"
            f"{header_row + 1}:"
            f"{get_column_letter(max_column)}"
            f"{data_end_row}"
        )
    else:
        data_range = None

    return TableDatasetSpec(
        filename=filename,
        sheet_name=formula_worksheet.title,
        table_kind=(
            classification.primary_content_type
        ),
        financial_statement_subtype=(
            classification
            .financial_statement_subtype
        ),
        metadata=metadata,
        header_row=header_row,
        header_range=(
            f"{formula_worksheet.title}!"
            f"{header_start}:{header_end}"
        ),
        data_range=data_range,
        full_range=(
            f"{formula_worksheet.title}!"
            f"{full_start}:{full_end}"
        ),
        row_count=len(extracted_rows),
        column_count=len(columns),
        columns=columns,
        rows=extracted_rows,
        warnings=warnings,
    )


def _sheet_contains_extractable_table(
    classification: SheetContentInspection,
) -> bool:
    """
    判斷該工作表是否具有可抽取表格。
    """
    if classification.primary_content_type in {
        SheetContentType.STRUCTURED_TABLE,
        SheetContentType.FINANCIAL_STATEMENT,
    }:
        return True

    if (
        classification.primary_content_type
        == SheetContentType.MIXED_CONTENT
    ):
        return any(
            component in {
                SheetContentType.STRUCTURED_TABLE,
                SheetContentType.FINANCIAL_STATEMENT,
            }
            for component in classification.components
        )

    return False


def extract_excel_tables(
    file_path: str | Path,
    sheet_name: str | None = None,
) -> WorkbookTableExtraction:
    """
    抽取 Excel 中所有可辨識的表格。

    sheet_name 有值時，只抽取指定工作表。
    """
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(
            f"找不到 Excel 檔案：{path}"
        )

    content_inspection = inspect_excel_content(
        path
    )

    classification_by_sheet = {
        sheet.sheet_name: sheet
        for sheet in content_inspection.sheets
    }

    formula_workbook = load_workbook(
        filename=path,
        read_only=False,
        data_only=False,
    )

    value_workbook = load_workbook(
        filename=path,
        read_only=False,
        data_only=True,
    )

    tables: list[TableDatasetSpec] = []
    skipped_sheets: list[str] = []
    warnings: list[str] = []

    try:
        available_sheet_names = (
            formula_workbook.sheetnames
        )

        if (
            sheet_name is not None
            and sheet_name
            not in available_sheet_names
        ):
            raise ValueError(
                f"找不到工作表：{sheet_name}"
            )

        selected_sheet_names = (
            [sheet_name]
            if sheet_name
            else available_sheet_names
        )

        for current_sheet_name in (
            selected_sheet_names
        ):
            classification = (
                classification_by_sheet[
                    current_sheet_name
                ]
            )

            if not _sheet_contains_extractable_table(
                classification
            ):
                skipped_sheets.append(
                    current_sheet_name
                )
                continue

            formula_worksheet = (
                formula_workbook[
                    current_sheet_name
                ]
            )

            value_worksheet = (
                value_workbook[
                    current_sheet_name
                ]
            )

            try:
                table = _extract_single_sheet(
                    filename=path.name,
                    formula_worksheet=(
                        formula_worksheet
                    ),
                    value_worksheet=(
                        value_worksheet
                    ),
                    classification=classification,
                )

                tables.append(table)

            except ValueError as error:
                skipped_sheets.append(
                    current_sheet_name
                )

                warnings.append(str(error))

    finally:
        formula_workbook.close()
        value_workbook.close()

    return WorkbookTableExtraction(
        filename=path.name,
        table_count=len(tables),
        tables=tables,
        skipped_sheets=skipped_sheets,
        warnings=warnings,
    )