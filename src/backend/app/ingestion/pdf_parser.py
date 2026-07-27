import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pdfplumber
from openpyxl.utils import get_column_letter
from pypdf import PdfReader

from app.ingestion.schemas import (
    ColumnDataType,
    ExtractedCell,
    ExtractedTableRow,
    FinancialStatementSubtype,
    PdfDocumentContent,
    PdfPageContent,
    SheetContentInspection,
    SheetContentType,
    SourceCell,
    TableColumnSpec,
    TableDatasetSpec,
    TableMetadata,
    WorkbookContentInspection,
    WorkbookTableExtraction,
)


MAX_PDF_PAGES = 100
MAX_PAGE_TEXT_CHARS = 100_000
MAX_DOCUMENT_TEXT_CHARS = 1_000_000
MAX_TABLES_PER_PAGE = 20

MIN_TEXT_LAYER_CHARS = 20
BLANK_ROWS_TO_STOP = 3


LINE_TABLE_SETTINGS = {
    "vertical_strategy": "lines",
    "horizontal_strategy": "lines",
    "snap_tolerance": 3,
    "join_tolerance": 3,
    "intersection_tolerance": 5,
}

TEXT_TABLE_SETTINGS = {
    "vertical_strategy": "text",
    "horizontal_strategy": "text",
    "min_words_vertical": 2,
    "min_words_horizontal": 1,
    "intersection_tolerance": 5,
}


FINANCIAL_KEYWORDS: dict[
    FinancialStatementSubtype,
    list[str],
] = {
    FinancialStatementSubtype.BALANCE_SHEET: [
        "資產負債表",
        "流動資產",
        "非流動資產",
        "資產總計",
        "流動負債",
        "非流動負債",
        "負債總計",
        "權益總計",
        "負債及權益總計",
        "balance sheet",
        "current assets",
        "total assets",
        "total liabilities",
        "total equity",
    ],
    FinancialStatementSubtype.INCOME_STATEMENT: [
        "損益表",
        "綜合損益表",
        "營業收入",
        "營業成本",
        "營業毛利",
        "營業利益",
        "稅前淨利",
        "本期淨利",
        "每股盈餘",
        "income statement",
        "revenue",
        "gross profit",
        "operating income",
        "net income",
    ],
    FinancialStatementSubtype.CASH_FLOW_STATEMENT: [
        "現金流量表",
        "營業活動之現金流量",
        "投資活動之現金流量",
        "籌資活動之現金流量",
        "期初現金",
        "期末現金",
        "cash flow statement",
        "operating activities",
        "investing activities",
        "financing activities",
    ],
}


FINANCIAL_MINIMUM_HITS = {
    FinancialStatementSubtype.BALANCE_SHEET: 2,
    FinancialStatementSubtype.INCOME_STATEMENT: 2,
    FinancialStatementSubtype.CASH_FLOW_STATEMENT: 2,
}


NUMBER_PATTERN = re.compile(
    r"^[+-]?\d[\d,]*(?:\.\d+)?$"
)

NEGATIVE_NUMBER_PATTERN = re.compile(
    r"^\(([+-]?\d[\d,]*(?:\.\d+)?)\)$"
)

PERCENT_PATTERN = re.compile(
    r"^([+-]?\d[\d,]*(?:\.\d+)?)%$"
)

DATE_PATTERN = re.compile(
    r"^\d{4}-\d{1,2}-\d{1,2}$"
)

UNIT_PATTERN = re.compile(
    r"(?:單位|unit)\s*[:：]\s*([^\n]+)",
    flags=re.IGNORECASE,
)


def _is_empty(value: Any) -> bool:
    return value is None or (
        isinstance(value, str)
        and not value.strip()
    )


def _clean_text(value: Any) -> str:
    if value is None:
        return ""

    return re.sub(
        r"\s+",
        " ",
        str(value).strip(),
    )


def _normalize_for_match(value: Any) -> str:
    return re.sub(
        r"[\s\u3000:：\-_/()（）]+",
        "",
        str(value or "").strip().lower(),
    )


def _normalize_scalar(value: Any) -> Any:
    """將 PDF 表格文字轉成常見 Python 型態。"""
    if not isinstance(value, str):
        return value

    text = _clean_text(value)

    if not text:
        return None

    lowered = text.lower()

    if lowered in {"true", "yes", "是"}:
        return True

    if lowered in {"false", "no", "否"}:
        return False

    percent_match = PERCENT_PATTERN.fullmatch(
        text
    )

    if percent_match:
        numeric_text = (
            percent_match
            .group(1)
            .replace(",", "")
        )

        number = float(numeric_text)

        if number.is_integer():
            return int(number)

        return number

    negative_match = (
        NEGATIVE_NUMBER_PATTERN.fullmatch(text)
    )

    if negative_match:
        numeric_text = (
            negative_match
            .group(1)
            .replace(",", "")
        )

        number = float(numeric_text)

        if number.is_integer():
            return -int(number)

        return -number

    if NUMBER_PATTERN.fullmatch(text):
        number = float(
            text.replace(",", "")
        )

        if number.is_integer():
            return int(number)

        return number

    if DATE_PATTERN.fullmatch(text):
        try:
            return date.fromisoformat(text)
        except ValueError:
            pass

    return text


def _check_pdf_access(
    file_path: str | Path,
) -> int:
    """
    檢查 PDF 是否可開啟與是否需要密碼。

    回傳總頁數。
    """
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(
            f"找不到 PDF：{path}"
        )

    try:
        reader = PdfReader(str(path))
    except Exception as error:
        raise ValueError(
            f"PDF 無法開啟：{error}"
        ) from error

    if reader.is_encrypted:
        try:
            decrypt_result = reader.decrypt("")
        except Exception as error:
            raise ValueError(
                "PDF 受到密碼或加密保護"
            ) from error

        if not decrypt_result:
            raise ValueError(
                "PDF 受到密碼保護，"
                "目前不接受需要密碼的 PDF"
            )

    return len(reader.pages)


def _detect_financial_statement(
    text: str,
) -> tuple[
    FinancialStatementSubtype | None,
    float,
    list[str],
]:
    normalized_document = (
        _normalize_for_match(text)
    )

    candidates: list[
        tuple[
            FinancialStatementSubtype,
            float,
            list[str],
        ]
    ] = []

    for subtype, keywords in (
        FINANCIAL_KEYWORDS.items()
    ):
        hits = [
            keyword
            for keyword in keywords
            if (
                _normalize_for_match(keyword)
                in normalized_document
            )
        ]

        minimum_hits = (
            FINANCIAL_MINIMUM_HITS[subtype]
        )

        if len(hits) < minimum_hits:
            continue

        score = min(
            0.99,
            0.55 + len(hits) * 0.08,
        )

        candidates.append(
            (subtype, score, hits)
        )

    if not candidates:
        return None, 0.0, []

    candidates.sort(
        key=lambda item: (
            item[1],
            len(item[2]),
        ),
        reverse=True,
    )

    return candidates[0]


def _is_plausible_table(
    table_object: Any,
    minimum_rows: int,
    minimum_consistency: float,
) -> bool:
    """
    排除 pdfplumber 將一般段落文字誤判成表格的情況。

    判斷條件：
    1. 至少有指定數量的有效列。
    2. 至少有兩欄。
    3. 多數資料列的欄位數具有一定一致性。
    """
    try:
        raw_rows = table_object.extract()
    except Exception:
        return False

    if not raw_rows:
        return False

    rows = _clean_table_rows(raw_rows)

    if len(rows) < minimum_rows:
        return False

    non_empty_counts = [
        sum(
            not _is_empty(value)
            for value in row
        )
        for row in rows
    ]

    meaningful_counts = [
        count
        for count in non_empty_counts
        if count >= 2
    ]

    # 至少要有數列真正包含兩個以上欄位。
    if len(meaningful_counts) < minimum_rows:
        return False

    maximum_column_count = max(
        meaningful_counts,
        default=0,
    )

    if maximum_column_count < 2:
        return False

    # 計算最常出現的欄位數比例。
    dominant_count = max(
        meaningful_counts.count(count)
        for count in set(meaningful_counts)
    )

    consistency = (
        dominant_count
        / len(meaningful_counts)
    )

    if consistency < minimum_consistency:
        return False

    return True


def _find_page_tables(
    page: Any,
) -> list[Any]:
    """
    先以 PDF 線條尋找表格。

    如果找不到，再以文字位置推測；但文字模式容易將
    標題與段落誤判為表格，因此使用更嚴格的過濾條件。
    """
    try:
        line_tables = page.find_tables(
            table_settings=LINE_TABLE_SETTINGS
        )
    except Exception:
        line_tables = []

    plausible_line_tables = [
        table
        for table in line_tables
        if _is_plausible_table(
            table_object=table,
            minimum_rows=2,
            minimum_consistency=0.5,
        )
    ]

    if plausible_line_tables:
        return plausible_line_tables[
            :MAX_TABLES_PER_PAGE
        ]

    try:
        text_tables = page.find_tables(
            table_settings=TEXT_TABLE_SETTINGS
        )
    except Exception:
        text_tables = []

    # 文字模式較容易出現假陽性，因此要求：
    # 至少三列，且 60% 以上的有效列具有一致欄位數。
    plausible_text_tables = [
        table
        for table in text_tables
        if _is_plausible_table(
            table_object=table,
            minimum_rows=3,
            minimum_consistency=0.6,
        )
    ]

    return plausible_text_tables[
        :MAX_TABLES_PER_PAGE
    ]

def _calculate_page_type(
    has_text_layer: bool,
    table_count: int,
    image_count: int,
    financial_subtype: (
        FinancialStatementSubtype | None
    ),
) -> tuple[
    SheetContentType,
    list[SheetContentType],
]:
    components: list[SheetContentType] = []

    if has_text_layer:
        components.append(
            SheetContentType.DOCUMENT_TEXT
        )

    if table_count > 0:
        components.append(
            SheetContentType.STRUCTURED_TABLE
        )

    if image_count > 0:
        components.append(
            SheetContentType.EMBEDDED_IMAGE
        )

    if financial_subtype is not None:
        components.append(
            SheetContentType.FINANCIAL_STATEMENT
        )

        return (
            SheetContentType.FINANCIAL_STATEMENT,
            components,
        )

    if (
        has_text_layer
        and table_count > 0
    ):
        return (
            SheetContentType.MIXED_CONTENT,
            components,
        )

    if table_count > 0:
        return (
            SheetContentType.STRUCTURED_TABLE,
            components,
        )

    if has_text_layer:
        return (
            SheetContentType.DOCUMENT_TEXT,
            components,
        )

    if image_count > 0:
        return (
            SheetContentType.EMBEDDED_IMAGE,
            components,
        )

    return (
        SheetContentType.UNKNOWN,
        components,
    )


def inspect_pdf_document(
    file_path: str | Path,
) -> tuple[
    WorkbookContentInspection,
    PdfDocumentContent,
]:
    """
    讀取 PDF 文字層並分類每一頁。

    回傳：
    - 相容既有 Pipeline 的內容分類
    - PDF 完整文字內容
    """
    path = Path(file_path)
    total_page_count = _check_pdf_access(path)

    warnings: list[str] = []

    process_page_count = min(
        total_page_count,
        MAX_PDF_PAGES,
    )

    if total_page_count > MAX_PDF_PAGES:
        warnings.append(
            f"PDF 共 {total_page_count} 頁，"
            f"目前只處理前 {MAX_PDF_PAGES} 頁"
        )

    page_contents: list[PdfPageContent] = []
    sheet_results: list[
        SheetContentInspection
    ] = []

    full_text_parts: list[str] = []
    stored_document_chars = 0

    scanned_page_count = 0
    total_text_char_count = 0
    document_text_truncated = False

    try:
        pdf = pdfplumber.open(str(path))
    except Exception as error:
        raise ValueError(
            f"PDF 文字層解析失敗：{error}"
        ) from error

    try:
        for page_index in range(
            process_page_count
        ):
            page = pdf.pages[page_index]
            page_number = page_index + 1

            try:
                raw_text = (
                    page.extract_text(
                        x_tolerance=2,
                        y_tolerance=3,
                    )
                    or ""
                )
            except Exception:
                raw_text = ""

            raw_text = raw_text.strip()

            text_char_count = len(raw_text)
            total_text_char_count += (
                text_char_count
            )

            word_count = len(
                raw_text.split()
            )

            image_count = len(
                getattr(page, "images", [])
            )

            tables = _find_page_tables(page)
            table_count = len(tables)

            has_text_layer = (
                text_char_count
                >= MIN_TEXT_LAYER_CHARS
            )

            likely_scanned = (
                not has_text_layer
                and image_count > 0
            )

            page_warnings: list[str] = []

            if likely_scanned:
                scanned_page_count += 1

                page_warnings.append(
                    "此頁可能是掃描影像，"
                    "需在第八階段使用 OCR"
                )

            (
                financial_subtype,
                financial_score,
                financial_hits,
            ) = _detect_financial_statement(
                raw_text
            )

            (
                primary_type,
                components,
            ) = _calculate_page_type(
                has_text_layer=has_text_layer,
                table_count=table_count,
                image_count=image_count,
                financial_subtype=(
                    financial_subtype
                ),
            )

            page_confidence = 0.30

            if financial_subtype is not None:
                page_confidence = financial_score

            elif table_count > 0:
                page_confidence = 0.90

            elif has_text_layer:
                page_confidence = 0.95

            elif likely_scanned:
                page_confidence = 0.80

            stored_page_text = raw_text[
                :MAX_PAGE_TEXT_CHARS
            ]

            page_text_truncated = (
                len(raw_text)
                > MAX_PAGE_TEXT_CHARS
            )

            if page_text_truncated:
                page_warnings.append(
                    "此頁文字過長，API 僅保留部分文字"
                )

            page_contents.append(
                PdfPageContent(
                    page_number=page_number,
                    width=float(page.width),
                    height=float(page.height),
                    text=stored_page_text,
                    text_char_count=(
                        text_char_count
                    ),
                    word_count=word_count,
                    text_truncated=(
                        page_text_truncated
                    ),
                    table_count=table_count,
                    image_count=image_count,
                    has_text_layer=(
                        has_text_layer
                    ),
                    likely_scanned=(
                        likely_scanned
                    ),
                    financial_statement_subtype=(
                        financial_subtype
                    ),
                    confidence=round(
                        page_confidence,
                        4,
                    ),
                    warnings=page_warnings,
                )
            )

            evidence = [
                f"文字字元數：{text_char_count}",
                f"偵測表格數：{table_count}",
                f"內嵌圖片數：{image_count}",
            ]

            if financial_hits:
                evidence.append(
                    "財務關鍵字："
                    + "、".join(
                        financial_hits[:8]
                    )
                )

            line_count = (
                len(raw_text.splitlines())
                if raw_text
                else 0
            )

            sheet_results.append(
                SheetContentInspection(
                    sheet_name=(
                        f"page_{page_number}"
                    ),
                    sheet_state="visible",
                    primary_content_type=(
                        primary_type
                    ),
                    components=components,
                    financial_statement_subtype=(
                        financial_subtype
                    ),
                    confidence=round(
                        page_confidence,
                        4,
                    ),
                    max_row=line_count,
                    max_column=1,
                    non_empty_cells=word_count,
                    text_cells=word_count,
                    numeric_cells=0,
                    formula_cells=0,
                    date_cells=0,
                    merged_range_count=0,
                    chart_count=0,
                    image_count=image_count,
                    detected_header_row=None,
                    structured_table_score=(
                        0.90
                        if table_count > 0
                        else 0.0
                    ),
                    financial_statement_score=(
                        round(
                            financial_score,
                            4,
                        )
                    ),
                    evidence=evidence,
                    warnings=page_warnings,
                )
            )

            page_marker = (
                f"\n\n--- Page "
                f"{page_number} ---\n"
            )

            available_chars = (
                MAX_DOCUMENT_TEXT_CHARS
                - stored_document_chars
            )

            if available_chars > 0:
                text_to_append = (
                    page_marker + raw_text
                )[:available_chars]

                full_text_parts.append(
                    text_to_append
                )

                stored_document_chars += len(
                    text_to_append
                )

            if (
                stored_document_chars
                >= MAX_DOCUMENT_TEXT_CHARS
                and page_index
                < process_page_count - 1
            ):
                document_text_truncated = True

    finally:
        pdf.close()

    meaningful_types = [
        result.primary_content_type
        for result in sheet_results
        if result.primary_content_type
        != SheetContentType.EMPTY
    ]

    if not meaningful_types:
        overall_type = (
            SheetContentType.EMPTY
        )

    elif len(set(meaningful_types)) == 1:
        overall_type = meaningful_types[0]

    else:
        overall_type = (
            SheetContentType.MIXED_CONTENT
        )

    if sheet_results:
        confidence = sum(
            result.confidence
            for result in sheet_results
        ) / len(sheet_results)
    else:
        confidence = 0.0

    if document_text_truncated:
        warnings.append(
            "PDF 文字總量過大，"
            "API 的 full_text 已截斷"
        )

    document = PdfDocumentContent(
        filename=path.name,
        page_count=total_page_count,
        processed_page_count=(
            process_page_count
        ),
        has_text_layer=any(
            page.has_text_layer
            for page in page_contents
        ),
        scanned_page_count=(
            scanned_page_count
        ),
        total_text_char_count=(
            total_text_char_count
        ),
        full_text="".join(
            full_text_parts
        ),
        text_truncated=(
            document_text_truncated
        ),
        pages=page_contents,
        warnings=warnings.copy(),
    )

    classification = (
        WorkbookContentInspection(
            filename=path.name,
            sheet_count=len(
                sheet_results
            ),
            overall_content_type=(
                overall_type
            ),
            confidence=round(
                confidence,
                4,
            ),
            sheets=sheet_results,
            warnings=warnings.copy(),
        )
    )

    return classification, document


def _clean_table_rows(
    raw_rows: list[list[Any]],
) -> list[list[str]]:
    rows = [
        [
            _clean_text(value)
            for value in row
        ]
        for row in raw_rows
        if row is not None
    ]

    while rows and not any(rows[0]):
        rows.pop(0)

    while rows and not any(rows[-1]):
        rows.pop()

    if not rows:
        return []

    max_columns = max(
        len(row)
        for row in rows
    )

    for row in rows:
        row.extend(
            [""] * (
                max_columns - len(row)
            )
        )

    non_empty_columns = [
        column_index
        for column_index in range(
            max_columns
        )
        if any(
            not _is_empty(
                row[column_index]
            )
            for row in rows
        )
    ]

    if not non_empty_columns:
        return []

    min_column = min(
        non_empty_columns
    )

    max_column = max(
        non_empty_columns
    )

    return [
        row[
            min_column:
            max_column + 1
        ]
        for row in rows
    ]


def _detect_table_header(
    rows: list[list[str]],
) -> int | None:
    candidates: list[
        tuple[int, float]
    ] = []

    for row_index in range(
        min(len(rows), 10)
    ):
        row = rows[row_index]

        non_empty_values = [
            value
            for value in row
            if not _is_empty(value)
        ]

        if len(non_empty_values) < 2:
            continue

        normalized_values = [
            _normalize_scalar(value)
            for value in non_empty_values
        ]

        text_ratio = sum(
            isinstance(value, str)
            for value in normalized_values
        ) / len(normalized_values)

        unique_ratio = (
            len(set(non_empty_values))
            / len(non_empty_values)
        )

        minimum_data_cells = max(
            1,
            (
                len(non_empty_values)
                + 1
            ) // 2,
        )

        following_rows = sum(
            sum(
                not _is_empty(value)
                for value in data_row
            ) >= minimum_data_cells
            for data_row in rows[
                row_index + 1:
                row_index + 5
            ]
        )

        if following_rows < 1:
            continue

        score = (
            text_ratio * 0.45
            + unique_ratio * 0.20
            + min(
                following_rows / 3,
                1,
            ) * 0.25
            + max(
                0,
                0.10
                - row_index * 0.01,
            )
        )

        candidates.append(
            (row_index, score)
        )

    if not candidates:
        return None

    candidates.sort(
        key=lambda item: item[1],
        reverse=True,
    )

    return candidates[0][0]


def _create_column_key(
    label: str,
    index: int,
    existing_keys: set[str],
) -> str:
    normalized = re.sub(
        r"[^\w]+",
        "_",
        label.strip().lower(),
        flags=re.UNICODE,
    ).strip("_")

    if not normalized:
        normalized = (
            f"column_{index + 1}"
        )

    candidate = normalized
    duplicate_index = 2

    while candidate in existing_keys:
        candidate = (
            f"{normalized}_"
            f"{duplicate_index}"
        )

        duplicate_index += 1

    return candidate


def _infer_data_type(
    values: list[Any],
) -> ColumnDataType:
    non_empty_values = [
        value
        for value in values
        if value is not None
    ]

    if not non_empty_values:
        return ColumnDataType.EMPTY

    detected_types: set[
        ColumnDataType
    ] = set()

    for value in non_empty_values:
        if isinstance(value, bool):
            detected_types.add(
                ColumnDataType.BOOLEAN
            )

        elif isinstance(value, datetime):
            detected_types.add(
                ColumnDataType.DATETIME
            )

        elif isinstance(value, date):
            detected_types.add(
                ColumnDataType.DATE
            )

        elif (
            isinstance(value, int)
            and not isinstance(value, bool)
        ):
            detected_types.add(
                ColumnDataType.INTEGER
            )

        elif isinstance(value, float):
            detected_types.add(
                ColumnDataType.NUMBER
            )

        else:
            detected_types.add(
                ColumnDataType.STRING
            )

    if detected_types == {
        ColumnDataType.INTEGER
    }:
        return ColumnDataType.INTEGER

    if detected_types.issubset({
        ColumnDataType.INTEGER,
        ColumnDataType.NUMBER,
    }):
        return ColumnDataType.NUMBER

    if len(detected_types) == 1:
        return next(iter(detected_types))

    return ColumnDataType.MIXED


def _extract_table_metadata(
    page: Any,
    table_bbox: tuple[
        float,
        float,
        float,
        float,
    ],
) -> TableMetadata:
    """
    取表格上方最後幾行文字，
    嘗試判斷標題與單位。
    """
    top = max(
        0,
        float(table_bbox[1]) - 150,
    )

    try:
        area = page.crop(
            (
                0,
                top,
                float(page.width),
                float(table_bbox[1]),
            )
        )

        text_above = (
            area.extract_text() or ""
        )
    except Exception:
        text_above = ""

    lines = [
        _clean_text(line)
        for line in text_above.splitlines()
        if _clean_text(line)
    ]

    unit: str | None = None
    title: str | None = None

    for line in lines:
        unit_match = (
            UNIT_PATTERN.search(line)
        )

        if unit_match:
            unit = (
                unit_match.group(1).strip()
            )

    non_unit_lines = [
        line
        for line in lines
        if not UNIT_PATTERN.search(line)
    ]

    if non_unit_lines:
        title = non_unit_lines[-1]

    return TableMetadata(
        title=title,
        entity=None,
        unit=unit,
        notes=[],
    )


def _table_to_dataset(
    filename: str,
    page: Any,
    page_number: int,
    table_number: int,
    table_object: Any,
    page_financial_subtype: (
        FinancialStatementSubtype | None
    ),
) -> TableDatasetSpec | None:
    try:
        raw_rows = table_object.extract()
    except Exception:
        return None

    if not raw_rows:
        return None

    rows = _clean_table_rows(
        raw_rows
    )

    if not rows:
        return None

    header_index = _detect_table_header(
        rows
    )

    if header_index is None:
        return None

    header_row = rows[header_index]
    max_column = len(header_row)

    virtual_sheet = (
        f"page_{page_number}_"
        f"table_{table_number}"
    )

    metadata = _extract_table_metadata(
        page,
        table_object.bbox,
    )

    existing_keys: set[str] = set()

    header_definitions: list[
        tuple[int, str, str, str | None]
    ] = []

    for column_index, raw_label in enumerate(
        header_row
    ):
        label = _clean_text(raw_label)

        if not label:
            label = (
                f"未命名欄位 "
                f"{column_index + 1}"
            )

        key = _create_column_key(
            label=label,
            index=column_index,
            existing_keys=existing_keys,
        )

        existing_keys.add(key)

        unit_match = re.search(
            r"[（(]\s*([^（）()]+)"
            r"\s*[）)]",
            label,
        )

        column_unit = (
            unit_match.group(1).strip()
            if unit_match
            else metadata.unit
        )

        header_definitions.append(
            (
                column_index,
                key,
                label,
                column_unit,
            )
        )

    extracted_rows: list[
        ExtractedTableRow
    ] = []

    column_values = {
        key: []
        for _, key, _, _
        in header_definitions
    }

    column_empty_counts = {
        key: 0
        for _, key, _, _
        in header_definitions
    }

    blank_rows = 0

    for row_index in range(
        header_index + 1,
        len(rows),
    ):
        raw_row = rows[row_index]

        if (
            sum(
                not _is_empty(value)
                for value in raw_row
            )
            == 0
        ):
            blank_rows += 1

            if (
                blank_rows
                >= BLANK_ROWS_TO_STOP
            ):
                break

            continue

        blank_rows = 0

        row_cells: dict[
            str,
            ExtractedCell,
        ] = {}

        for (
            column_index,
            key,
            _,
            _,
        ) in header_definitions:
            raw_value = (
                raw_row[column_index]
                if column_index
                < len(raw_row)
                else ""
            )

            normalized_value = (
                _normalize_scalar(
                    raw_value
                )
            )

            if normalized_value is None:
                column_empty_counts[key] += 1

            column_values[key].append(
                normalized_value
            )

            coordinate = (
                f"{get_column_letter(column_index + 1)}"
                f"{row_index + 1}"
            )

            row_cells[key] = ExtractedCell(
                raw_value=raw_value,
                value=normalized_value,
                formula=None,
                number_format=None,
                source=SourceCell(
                    sheet=virtual_sheet,
                    cell=coordinate,
                    row=row_index + 1,
                    column=column_index + 1,
                ),
            )

        extracted_rows.append(
            ExtractedTableRow(
                excel_row=row_index + 1,
                cells=row_cells,
            )
        )

    columns: list[
        TableColumnSpec
    ] = []

    for column_index, (
        _,
        key,
        label,
        unit,
    ) in enumerate(header_definitions):
        columns.append(
            TableColumnSpec(
                key=key,
                label=label,
                index=column_index,
                data_type=_infer_data_type(
                    column_values[key]
                ),
                unit=unit,
                nullable=(
                    column_empty_counts[key] > 0
                ),
                header_source=SourceCell(
                    sheet=virtual_sheet,
                    cell=(
                        f"{get_column_letter(column_index + 1)}"
                        f"{header_index + 1}"
                    ),
                    row=header_index + 1,
                    column=column_index + 1,
                ),
            )
        )

    last_column_letter = (
        get_column_letter(
            max_column
        )
    )

    header_excel_row = (
        header_index + 1
    )

    if extracted_rows:
        last_data_row = (
            extracted_rows[-1].excel_row
        )

        data_range = (
            f"{virtual_sheet}!A"
            f"{header_excel_row + 1}:"
            f"{last_column_letter}"
            f"{last_data_row}"
        )

        full_end_row = last_data_row
    else:
        data_range = None
        full_end_row = header_excel_row

    table_kind = (
        SheetContentType
        .FINANCIAL_STATEMENT
        if page_financial_subtype
        is not None
        else SheetContentType
        .STRUCTURED_TABLE
    )

    return TableDatasetSpec(
        filename=filename,
        sheet_name=virtual_sheet,
        table_kind=table_kind,
        financial_statement_subtype=(
            page_financial_subtype
        ),
        metadata=metadata,
        header_row=header_excel_row,
        header_range=(
            f"{virtual_sheet}!A"
            f"{header_excel_row}:"
            f"{last_column_letter}"
            f"{header_excel_row}"
        ),
        data_range=data_range,
        full_range=(
            f"{virtual_sheet}!A"
            f"{header_excel_row}:"
            f"{last_column_letter}"
            f"{full_end_row}"
        ),
        row_count=len(extracted_rows),
        column_count=len(columns),
        columns=columns,
        rows=extracted_rows,
        warnings=[
            "此表格由 PDF 視覺結構推估，"
            "建議保留人工確認機制"
        ],
    )


def extract_pdf_tables(
    file_path: str | Path,
    sheet_name: str | None = None,
) -> WorkbookTableExtraction:
    """
    從 PDF 文字層抽取表格。

    sheet_name 可輸入：
    page_1
    page_2

    留空代表處理全部頁面。
    """
    path = Path(file_path)
    total_page_count = _check_pdf_access(path)

    requested_page: int | None = None

    if sheet_name:
        page_match = re.fullmatch(
            r"page_(\d+)",
            sheet_name,
        )

        if not page_match:
            raise ValueError(
                "PDF sheet_name 格式應為 page_1"
            )

        requested_page = int(
            page_match.group(1)
        )

        if not (
            1
            <= requested_page
            <= total_page_count
        ):
            raise ValueError(
                f"PDF 沒有第 "
                f"{requested_page} 頁"
            )

    process_page_count = min(
        total_page_count,
        MAX_PDF_PAGES,
    )

    tables: list[TableDatasetSpec] = []
    skipped_sheets: list[str] = []
    warnings: list[str] = []

    if total_page_count > MAX_PDF_PAGES:
        warnings.append(
            f"PDF 超過 {MAX_PDF_PAGES} 頁，"
            "目前只處理前段頁面"
        )

    pdf = pdfplumber.open(str(path))

    try:
        for page_index in range(
            process_page_count
        ):
            page_number = page_index + 1

            if (
                requested_page is not None
                and page_number
                != requested_page
            ):
                continue

            page = pdf.pages[page_index]

            try:
                page_text = (
                    page.extract_text() or ""
                )
            except Exception:
                page_text = ""

            (
                financial_subtype,
                _,
                _,
            ) = _detect_financial_statement(
                page_text
            )

            page_tables = (
                _find_page_tables(page)
            )

            extracted_on_page = 0

            for table_index, table_object in enumerate(
                page_tables,
                start=1,
            ):
                dataset = _table_to_dataset(
                    filename=path.name,
                    page=page,
                    page_number=page_number,
                    table_number=table_index,
                    table_object=table_object,
                    page_financial_subtype=(
                        financial_subtype
                    ),
                )

                if dataset is None:
                    continue

                tables.append(dataset)
                extracted_on_page += 1

            if extracted_on_page == 0:
                skipped_sheets.append(
                    f"page_{page_number}"
                )

                if (
                    len(page_text.strip())
                    < MIN_TEXT_LAYER_CHARS
                    and len(
                        getattr(
                            page,
                            "images",
                            [],
                        )
                    ) > 0
                ):
                    warnings.append(
                        f"第 {page_number} 頁"
                        "可能是掃描影像，"
                        "需使用 OCR"
                    )

    finally:
        pdf.close()

    return WorkbookTableExtraction(
        filename=path.name,
        table_count=len(tables),
        tables=tables,
        skipped_sheets=skipped_sheets,
        warnings=warnings,
    )