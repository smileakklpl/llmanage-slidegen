"""
FR-3 外部稽核 Excel 輸出
=========================
對應 docs/圖表原生性與資料同步設計.md Stage 6 與 §5 的第三份資料副本。

輸出一份與簡報圖表一一對應的 `.xlsx`，讓主管能獨立核對每個數字，
不必打開 PPT 或右鍵編輯資料。

工作表命名規則：``P.{頁碼}_{指標名稱}``（對齊附件四慣例）。

一致性保證：本模組與 PPT 圖表**吃同一份 ChartSpec**，因此三份副本
（chart XML 快取、PPT 內嵌 workbook、本檔案）數值必然相同。這裡刻意
不重新計算任何數值，只做搬運與排版。
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..core import config
from ..charts.chart_builder import ChartSpec, ScatterSpec
from ..charts.chart_planner import ResolvedChart


#: Excel 工作表名稱長度上限（Excel 硬限制）。
MAX_SHEET_NAME_LENGTH = 31

#: Excel 工作表名稱不允許的字元。
_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TITLE_FONT = Font(bold=True, size=13)
_NOTE_FONT = Font(italic=True, size=9, color="595959")
_THIN_BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)


@dataclass
class ExportReport:
    """匯出結果摘要。"""

    output_path: Path | None = None
    sheet_names: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def sanitize_sheet_name(raw: str, used: Sequence[str] = ()) -> str:
    """
    產生合法且不重複的工作表名稱。

    Excel 限制：長度 ≤ 31、不得含 ``[]:*?/\\``、不得重複。
    超長時從中間截斷並保留頁碼前綴（頁碼是對照 PPT 的關鍵資訊）。
    """
    cleaned = _INVALID_SHEET_CHARS.sub("_", raw).strip() or "Sheet"

    if len(cleaned) > MAX_SHEET_NAME_LENGTH:
        cleaned = cleaned[:MAX_SHEET_NAME_LENGTH]

    if cleaned not in used:
        return cleaned

    # 重名時加序號，並確保加了序號仍不超長。
    for index in range(2, 100):
        suffix = f"_{index}"
        candidate = cleaned[: MAX_SHEET_NAME_LENGTH - len(suffix)] + suffix

        if candidate not in used:
            return candidate

    raise ValueError(f"無法為 {raw!r} 產生不重複的工作表名稱")


def sheet_name_for(chart: ResolvedChart, used: Sequence[str] = ()) -> str:
    """依 ``P.{頁碼}_{指標名稱}`` 規則產生工作表名稱。"""
    page = chart.plan.page_number
    prefix = f"P.{page}_" if page is not None else ""
    return sanitize_sheet_name(f"{prefix}{chart.metric.name}", used)


# ---------------------------------------------------------------------------
# 單張工作表
# ---------------------------------------------------------------------------
def _write_header_block(worksheet, chart: ResolvedChart) -> int:
    """
    寫入表頭資訊區，回傳資料表起始列號。

    表頭刻意包含計算公式與來源檔案，滿足「任一簡報數字可回溯至來源
    儲存格與計算公式」的可追溯性要求。
    """
    metric = chart.metric
    rows = [
        ("簡報頁碼", chart.plan.page_number),
        ("圖表標題", chart.plan.chart_title),
        ("圖表類型", chart.skill_name),
        ("指標名稱", metric.name),
        ("指標鍵", metric.metric_key),
        ("單位", metric.unit or "未標示"),
        ("計算方式", metric.formula or "原始值，未經計算"),
    ]

    worksheet["A1"] = chart.plan.chart_title or metric.name
    worksheet["A1"].font = _TITLE_FONT

    row = 3

    for label, value in rows:
        worksheet.cell(row=row, column=1, value=label).font = Font(bold=True)
        worksheet.cell(row=row, column=2, value=value)
        row += 1

    if metric.notes:
        worksheet.cell(row=row, column=1, value="資料注意事項").font = Font(bold=True)

        for note in metric.notes:
            worksheet.cell(row=row, column=2, value=note).font = _NOTE_FONT
            row += 1
    else:
        row += 0

    return row + 1


def _write_category_table(
    worksheet,
    chart: ResolvedChart,
    start_row: int,
) -> None:
    """寫入類別型圖表（column/bar/line/pie）的資料表與來源欄。"""
    spec = chart.spec
    assert isinstance(spec, ChartSpec)

    metric = chart.metric
    series_names = list(spec.series)

    headers = ["類別", *series_names, "來源（檔案 / 工作表 / 儲存格）"]

    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=start_row, column=column_index, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER

    for offset, category in enumerate(spec.categories):
        row = start_row + 1 + offset

        worksheet.cell(row=row, column=1, value=category).border = _THIN_BORDER

        sources: list[str] = []

        for series_offset, series_name in enumerate(series_names, start=2):
            values = spec.series[series_name]
            value = values[offset] if offset < len(values) else None

            cell = worksheet.cell(row=row, column=series_offset, value=value)
            cell.border = _THIN_BORDER
            cell.number_format = _number_format_for(metric.unit)

            source = metric.source_of(series_name, category)

            if source is not None:
                sources.append(f"{series_name}: {source.describe()}")

        worksheet.cell(
            row=row,
            column=len(headers),
            # 衍生指標沒有直接來源儲存格，明確標示計算來源而非留空。
            value="；".join(sources) or "衍生計算值（見上方計算方式）",
        ).border = _THIN_BORDER

    _autofit(worksheet, len(headers))


def _write_scatter_table(
    worksheet,
    chart: ResolvedChart,
    start_row: int,
) -> None:
    """寫入散點圖的資料表。"""
    spec = chart.spec
    assert isinstance(spec, ScatterSpec)

    x_name, _, y_name = spec.series_name.partition(" vs ")
    headers = ["標籤", x_name or "X", y_name or "Y"]

    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=start_row, column=column_index, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER

    labels = list(spec.labels or [])

    for offset, (x_value, y_value) in enumerate(spec.points):
        row = start_row + 1 + offset
        label = labels[offset] if offset < len(labels) else f"點 {offset + 1}"

        for column_index, value in enumerate((label, x_value, y_value), start=1):
            cell = worksheet.cell(row=row, column=column_index, value=value)
            cell.border = _THIN_BORDER

    _autofit(worksheet, len(headers))


def _number_format_for(unit: str | None) -> str:
    if unit in {"%", "％"}:
        return "0.0"

    if unit == "名":
        return "0"

    return "#,##0.0"


def _autofit(worksheet, column_count: int) -> None:
    """粗略估算欄寬。中文字元以兩倍寬計算。"""
    for column_index in range(1, column_count + 1):
        letter = get_column_letter(column_index)
        widest = 8

        for cell in worksheet[letter]:
            if cell.value is None:
                continue

            text = str(cell.value)
            width = sum(2 if ord(char) > 0x2E80 else 1 for char in text)
            widest = max(widest, min(width + 2, 60))

        worksheet.column_dimensions[letter].width = widest


# ---------------------------------------------------------------------------
# 主入口
# ---------------------------------------------------------------------------
def export_audit_workbook(
    charts: Sequence[ResolvedChart],
    *,
    output_path: str | Path | None = None,
    include_index: bool = True,
) -> ExportReport:
    """
    輸出稽核用 Excel，每張圖表一個工作表。

    Args:
        charts: 已查表完成的圖表（與 PPT 用的是同一批物件）。
        output_path: 輸出路徑，預設 ``outputs/deck_data.xlsx``。
        include_index: 是否加一張索引頁，列出頁碼與工作表對照。

    Returns:
        :class:`ExportReport`。
    """
    if not charts:
        raise ValueError("沒有任何圖表可匯出")

    workbook = Workbook()
    # 移除 openpyxl 預設建立的空白工作表，改由下方依需求新增。
    workbook.remove(workbook.active)

    report = ExportReport()

    if include_index:
        _write_index_sheet(workbook, charts)
        report.sheet_names.append("索引")

    for chart in charts:
        name = sheet_name_for(chart, report.sheet_names)
        worksheet = workbook.create_sheet(title=name)
        report.sheet_names.append(name)

        data_start = _write_header_block(worksheet, chart)

        if isinstance(chart.spec, ScatterSpec):
            _write_scatter_table(worksheet, chart, data_start)
        else:
            _write_category_table(worksheet, chart, data_start)

        if chart.metric.requires_human_review:
            report.warnings.append(
                f"工作表 {name} 的來源資料標記為需人工確認"
            )

    target = Path(output_path or (config.OUTPUT_DIR / "deck_data.xlsx"))
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(target))

    report.output_path = target
    return report


def _write_index_sheet(workbook: Workbook, charts: Sequence[ResolvedChart]) -> None:
    """索引頁：頁碼 ↔ 工作表 ↔ 指標對照，方便主管快速定位。"""
    worksheet = workbook.create_sheet(title="索引")

    worksheet["A1"] = "簡報資料對照索引"
    worksheet["A1"].font = _TITLE_FONT

    headers = ["簡報頁碼", "工作表名稱", "圖表標題", "指標鍵", "單位", "來源檔案"]

    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=3, column=column_index, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _THIN_BORDER

    used: list[str] = ["索引"]

    for offset, chart in enumerate(charts):
        row = 4 + offset
        name = sheet_name_for(chart, used)
        used.append(name)

        source_files = sorted(
            {
                source.filename
                for source in chart.metric.evidence.values()
            }
        )

        values = [
            chart.plan.page_number,
            name,
            chart.plan.chart_title,
            chart.metric.metric_key,
            chart.metric.unit or "未標示",
            "、".join(source_files) or "衍生計算",
        ]

        for column_index, value in enumerate(values, start=1):
            worksheet.cell(row=row, column=column_index, value=value).border = (
                _THIN_BORDER
            )

    _autofit(worksheet, len(headers))
