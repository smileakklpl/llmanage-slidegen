"""SheetMap + xlsx → 表格資料。

一支 reader 吃下所有工作表，靠的是 SheetMap 描述的形狀參數。

最關鍵的一行：`rows` 一律排除 `total_row`。合計列混進一般機構會佔據第一名
並使其後所有名次位移。
"""

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from contracts.sheet_map import SheetSpec


@dataclass
class SheetData:
    """一張工作表讀出來的內容。期間值以 {期間: 數值} 表示。"""

    sheet_name: str
    periods: List[str]
    rows: Dict[str, Dict[str, float]]  # {機構名稱: {期間: 值}}，**不含合計列**
    total: Dict[str, float]  # 合計列的 {期間: 值}
    total_label: str = ""

    def entity_sum(self, entity: str) -> Optional[float]:
        """該機構全期間加總。市佔率的分子。"""
        vals = self.rows.get(entity)
        return sum(vals.values()) if vals else None

    def total_sum(self) -> float:
        """合計列全期間加總。市佔率的分母。"""
        return sum(self.total.values())


def _num(v) -> Optional[float]:
    return float(v) if isinstance(v, (int, float)) else None


def read_sheet(xlsx: str | Path, spec: SheetSpec) -> SheetData:
    ws = load_workbook(xlsx, data_only=True)[spec.sheet_name]

    entity_col = spec.entity_col
    if entity_col is None:
        raise ValueError(f"「{spec.sheet_name}」的 SheetMap 沒有標出 entity 欄，無法讀取")

    ecol = column_index_from_string(entity_col)
    period_cols = [(c.period_key or c.header, column_index_from_string(c.col_letter))
                   for c in spec.columns if c.role == "period"]
    periods = [p for p, _ in period_cols]

    rows: Dict[str, Dict[str, float]] = {}
    total: Dict[str, float] = {}

    for r in range(spec.first_data_row, spec.last_data_row + 1):
        name = ws.cell(r, ecol).value
        if name is None or str(name).strip() == "":
            continue
        name = str(name).strip()
        vals = {p: v for p, ci in period_cols if (v := _num(ws.cell(r, ci).value)) is not None}

        # 合計列絕不進 rows。排除它是 reader 的責任，不是下游的。
        if spec.total_row is not None and r == spec.total_row:
            total = vals
            continue
        rows[name] = vals

    if spec.total_row is not None and not total:
        raise ValueError(
            f"「{spec.sheet_name}」的 SheetMap 指出合計列在第 {spec.total_row} 列，"
            f"但該列讀不到數值。結構定位可能有誤，拒絕繼續——"
            f"合計列若判斷錯誤，排名會整個位移。"
        )

    return SheetData(
        sheet_name=spec.sheet_name,
        periods=periods,
        rows=rows,
        total=total,
        total_label=spec.total_label or "",
    )
