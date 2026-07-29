"""來源格式辨識 — 認得的走確定性快路徑，認不得的才問模型。

    認得的格式  →  確定性建 SheetMap（零延遲、零 token）
    認不得的    →  profiler + LLM locator（通用性）

模型在 total_row / row_order 這類確定性可算的欄位上會出錯，已知格式不該
承擔那個風險。LLM 路徑仍是處理未知格式的唯一手段，由 spike_a 量它的命中率。

目前認得兩種：

  fsc_monthly       月報原始檔（entity_by_metric，標題列第 4 列，第 38 列後是註釋）
  entity_by_period  機構 × 期間寬表

辨識一律保守：指紋不完全符合就回 unknown 交給模型。認錯的代價比認不出來高。
"""

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Set

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from contracts.sheet_map import ColumnSpec, SheetSpec
from engine.profiler import trend

FSC_MONTHLY = "fsc_monthly"
ENTITY_BY_PERIOD = "entity_by_period"
UNKNOWN = "unknown"

TOTAL_LABELS = ("總計", "合計", "小計", "Total")
_PERIOD_HEADER = re.compile(r"^1\d{4}$")  # 民國年月六碼，如 11412


def norm(v) -> str:
    return re.sub(r"[\s　]+", "", str(v or ""))


@dataclass
class Recognition:
    kind: str
    sheets: List[SheetSpec] = field(default_factory=list)
    notes: List[str] = field(default_factory=list)
    periods: Set[str] = field(default_factory=set)

    @property
    def recognized(self) -> bool:
        return self.kind != UNKNOWN and bool(self.sheets)


def _find_label_row(ws, col: int, lo: int, hi: int) -> Optional[int]:
    """在指定範圍內找出實體欄文字是合計標籤的那一列。純詞彙比對，不涉語意。"""
    for r in range(lo, hi + 1):
        if norm(ws.cell(r, col).value) in TOTAL_LABELS:
            return r
    return None


def _last_row_with_entity(ws, col: int, start: int, limit: int = 400) -> int:
    """實體欄最後一個非空列。金管會檔第 38 列之後是註釋，靠這個切掉。"""
    last = start
    blanks = 0
    for r in range(start, start + limit):
        if norm(ws.cell(r, col).value):
            last, blanks = r, 0
        else:
            blanks += 1
            if blanks >= 3:  # 連續空列視為資料結束
                break
    return last


# --- 指紋 1：金管會月報原始檔 ------------------------------------------------
FSC_HEADER_ROW = 4
FSC_FIRST_DATA_ROW = 5


def _match_fsc_monthly(ws) -> Optional[SheetSpec]:
    if norm(ws.cell(FSC_HEADER_ROW, 1).value) != "金融機構名稱":
        return None
    total_row = _find_label_row(ws, 1, FSC_FIRST_DATA_ROW, FSC_FIRST_DATA_ROW + 60)
    if total_row is None:
        return None

    cols = [ColumnSpec(header="金融機構名稱", col_letter="A", role="entity")]
    for c in range(2, (ws.max_column or 2) + 1):
        h = norm(ws.cell(FSC_HEADER_ROW, c).value)
        if not h:
            continue
        # 這種檔一欄一指標，沒有期間欄；期間寫在第 3 列的「資料月份」。
        cols.append(ColumnSpec(header=h, col_letter=get_column_letter(c), role="derived"))

    return SheetSpec(
        sheet_name=ws.title,
        archetype="entity_by_metric",
        header_row=FSC_HEADER_ROW,
        first_data_row=FSC_FIRST_DATA_ROW,
        last_data_row=total_row,
        columns=cols,
        total_row=total_row,
        total_label=norm(ws.cell(total_row, 1).value),
        row_order="source",
        notes=[f"第 {total_row + 1} 列之後為揭露項目註釋，非資料"],
    )


# --- 指紋 2：機構 × 期間寬表 --------------------------------------------------
def _match_entity_by_period(ws) -> Optional[SheetSpec]:
    header_row = None
    for r in range(1, 8):
        heads = [norm(ws.cell(r, c).value) for c in range(2, min(ws.max_column or 2, 40) + 1)]
        if sum(1 for h in heads if _PERIOD_HEADER.match(h)) >= 3:
            header_row = r
            break
    if header_row is None:
        return None

    first = header_row + 1
    last = _last_row_with_entity(ws, 1, first)
    if last <= first:
        return None

    total_row = _find_label_row(ws, 1, first, last)

    cols = [ColumnSpec(header=norm(ws.cell(header_row, 1).value) or "實體",
                       col_letter="A", role="entity")]
    period_idx: List[int] = []
    for c in range(2, (ws.max_column or 2) + 1):
        h = norm(ws.cell(header_row, c).value)
        if not h:
            continue
        letter = get_column_letter(c)
        if _PERIOD_HEADER.match(h):
            cols.append(ColumnSpec(header=h, col_letter=letter, role="period", period_key=h))
            period_idx.append(c)
        else:
            # 非期間的數值欄一律當衍生欄——下游會重算而不採用來源值
            cols.append(ColumnSpec(header=h, col_letter=letter, role="derived"))

    # row_order 用算術判定，不猜。
    #
    # 必須掃所有數值欄，不能只看最後一個期間：來源可能依衍生欄
    # 降序排的，而市佔率＝全年加總佔比，並不蘊含最新月份也單調——
    # 只探測 11412 會得到 none，把 sorted_desc 誤判成 source。
    # 衍生欄先掃，因為排序鍵通常是衍生值（市佔率、排名）。
    derived_idx = [c for c in range(2, (ws.max_column or 2) + 1)
                   if c not in period_idx and norm(ws.cell(header_row, c).value)]
    order, sorted_by = "source", None
    for probe in derived_idx + period_idx:
        vals = [ws.cell(r, probe).value for r in range(first, last + 1) if r != total_row]
        t = trend(vals)
        if t in ("desc", "asc"):
            order, sorted_by = f"sorted_{t}", norm(ws.cell(header_row, probe).value)
            break

    return SheetSpec(
        sheet_name=ws.title,
        archetype="entity_by_period",
        header_row=header_row,
        first_data_row=first,
        last_data_row=last,
        columns=cols,
        total_row=total_row,
        total_label=norm(ws.cell(total_row, 1).value) if total_row else None,
        row_order=order,
        sorted_by=sorted_by,
    )


_MATCHERS = ((FSC_MONTHLY, _match_fsc_monthly), (ENTITY_BY_PERIOD, _match_entity_by_period))


def recognize_workbook(path: str | Path) -> Recognition:
    """辨識單一活頁簿。所有工作表必須同屬一種格式，否則回 unknown。"""
    wb = load_workbook(path, data_only=True)
    for kind, matcher in _MATCHERS:
        specs = [matcher(ws) for ws in wb.worksheets]
        if all(s is not None for s in specs):
            periods = {c.period_key for s in specs for c in s.columns
                       if c.role == "period" and c.period_key}
            return Recognition(kind=kind, sheets=specs, periods=periods,
                               notes=[f"{Path(path).name}：辨識為 {kind}"])
    return Recognition(kind=UNKNOWN, notes=[f"{Path(path).name}：無法辨識，需交給模型定位"])


def recognize_dataset(target: str | Path) -> Dict[Path, Recognition]:
    """辨識一個檔案或一整個目錄（一指標一檔的資料集）。"""
    target = Path(target)
    files = sorted(target.glob("*.xlsx")) if target.is_dir() else [target]
    if not files:
        raise ValueError(f"{target} 底下沒有 xlsx")
    return {f: recognize_workbook(f) for f in files}
