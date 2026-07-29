"""Workbook profiler — 把 xlsx 壓成幾 KB 的純文字結構描述。

這支程式**不做任何語意判斷**，只負責如實描述。判斷交給模型。
輸出刻意保持純文字而非 JSON：省 token，且模型對表格式文字的理解比巢狀 JSON 好。

三個設計決定：

1. **頭尾各數列都要給。** 合計列可能在首列也可能在末列，只給開頭會漏掉。
2. **不信 openpyxl 的 max_column。** 尾端可能有格式沒內容，以實際有值的最右欄為準。
3. **樣本值截斷。** 樣本供模型判斷欄位型別與位置，不是給它推論用的資料。
"""

from dataclasses import dataclass
from pathlib import Path
from typing import Any, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

MAX_CELL_CHARS = 18


def _fmt(v: Any) -> str:
    if v is None:
        return ""
    s = str(v)
    if isinstance(v, float) and v == int(v):
        s = str(int(v))
    return s if len(s) <= MAX_CELL_CHARS else s[: MAX_CELL_CHARS - 1] + "…"


def _true_extent(ws) -> tuple[int, int]:
    """回傳實際有值的 (最大列, 最大欄)。不使用 ws.max_row / max_column。"""
    max_r = max_c = 0
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        for c_idx, v in enumerate(row, 1):
            if v is not None and str(v).strip() != "":
                max_r = max(max_r, r_idx)
                max_c = max(max_c, c_idx)
    return max_r, max_c


def trend(values: List[Any]) -> str:
    """回傳 'desc' / 'asc' / 'none'。純算術，不涉語意。"""
    nums = [v for v in values if isinstance(v, (int, float))]
    if len(nums) < 3:
        return "none"
    if all(a > b for a, b in zip(nums, nums[1:])):
        return "desc"
    if all(a < b for a, b in zip(nums, nums[1:])):
        return "asc"
    return "none"


def _trend_report(body: List[List[Any]], headers: List[str], kinds: List[str]) -> List[str]:
    """逐欄陳述數值走勢，並額外試算「排除首列」「排除末列」兩種情形。

    為什麼要做這件事：單調性是算術，不是語意判斷，本來就該由程式算。
    實測 qwen2.5:14b 讓它自己比大小時，四張工作表的 row_order 全部答反——
    它看的是合計列在頭還是在尾，不是看數值。合計列通常落在首列或末列，
    所以三種切法一起報，模型不必自己猜要不要排除它。
    """
    lines: List[str] = []
    for i, (h, k) in enumerate(zip(headers, kinds)):
        if k not in ("number", "ratio?"):
            continue
        col = [r[i] for r in body]
        whole, drop_first, drop_last = trend(col), trend(col[1:]), trend(col[:-1])
        if "none" == whole == drop_first == drop_last:
            continue
        lines.append(
            f"  {get_column_letter(i + 1)} 欄 {h!r}："
            f"全部資料列={whole}；排除首列={drop_first}；排除末列={drop_last}"
        )
    return lines


def _col_kind(values: List[Any]) -> str:
    vals = [v for v in values if v is not None and str(v).strip() != ""]
    if not vals:
        return "empty"
    nums = sum(1 for v in vals if isinstance(v, (int, float)))
    if nums / len(vals) >= 0.9:
        frac = sum(1 for v in vals if isinstance(v, float) and 0 <= v <= 1)
        return "ratio?" if frac / len(vals) >= 0.9 else "number"
    return "text"


@dataclass
class SheetProfile:
    name: str
    n_rows: int
    n_cols: int
    reported_cols: int
    merged: List[str]
    head: List[List[str]]
    tail: List[List[str]]
    col_kinds: List[str]
    headers: List[str]
    trends: List[str]

    def to_text(self) -> str:
        lines = [f"## 工作表「{self.name}」"]
        lines.append(f"實際資料範圍：{self.n_rows} 列 × {self.n_cols} 欄")
        if self.reported_cols != self.n_cols:
            lines.append(
                f"（注意：檔案回報 {self.reported_cols} 欄，"
                f"但第 {self.n_cols + 1} 欄之後只有格式沒有內容）"
            )
        lines.append(f"合併儲存格：{', '.join(self.merged) if self.merged else '無'}")

        lines.append("\n欄位一覽：")
        for i, (h, k) in enumerate(zip(self.headers, self.col_kinds), 1):
            lines.append(f"  {get_column_letter(i)}  值型別={k:7s}  第1列文字={h!r}")

        lines.append("\n數值走勢（由程式算出的算術事實，desc=嚴格遞減 asc=嚴格遞增 none=非單調）：")
        lines.extend(self.trends or ["  所有數值欄皆非單調"])

        lines.append(f"\n前 {len(self.head)} 列：")
        for r_no, row in self.head:
            lines.append(f"  第{r_no:>3}列 | " + " | ".join(row))
        lines.append(f"\n後 {len(self.tail)} 列：")
        for r_no, row in self.tail:
            lines.append(f"  第{r_no:>3}列 | " + " | ".join(row))
        return "\n".join(lines)


def profile_sheet(ws, head_n: int = 5, tail_n: int = 3) -> SheetProfile:
    reported_cols = ws.max_column or 0
    n_rows, n_cols = _true_extent(ws)

    grid = [
        [row[i] if i < len(row) else None for i in range(n_cols)]
        for row in ws.iter_rows(min_row=1, max_row=n_rows, values_only=True)
    ]

    headers = [_fmt(v) for v in grid[0]] if grid else []
    body = grid[1:] if len(grid) > 1 else []
    col_kinds = [_col_kind([r[i] for r in body]) for i in range(n_cols)]

    head = [(i + 1, [_fmt(v) for v in grid[i]]) for i in range(min(head_n, n_rows))]
    tail_start = max(head_n, n_rows - tail_n)
    tail = [(i + 1, [_fmt(v) for v in grid[i]]) for i in range(tail_start, n_rows)]

    return SheetProfile(
        name=ws.title,
        n_rows=n_rows,
        n_cols=n_cols,
        reported_cols=reported_cols,
        merged=[str(r) for r in ws.merged_cells.ranges][:12],
        head=head,
        tail=tail,
        col_kinds=col_kinds,
        headers=headers,
        trends=_trend_report(body, headers, col_kinds),
    )


def profile_workbook(
    path: str | Path, head_n: int = 5, tail_n: int = 3
) -> dict[str, str]:
    """回傳 {工作表名稱: 該張的結構描述}。

    刻意逐張分開而非合成一大段：整份一次餵給模型有兩個代價——
    輸入長度隨工作表數線性成長（弱模型的 context 撐不住，且會被靜默截斷），
    而且模型漏掉後半段時，你只會看到「整張漏掉」，分不清是沒看到還是判斷錯。
    逐張呼叫則每次輸入固定約 550 tokens，且失敗可歸因到單一張表。

    每段都保留活頁簿標頭，模型才知道 SheetMap.workbook 要填什麼。
    """
    path = Path(path)
    wb = load_workbook(path, data_only=True)
    header = (
        f"# 活頁簿：{path.name}\n"
        f"工作表共 {len(wb.sheetnames)} 張：{wb.sheetnames}\n"
    )
    return {
        name: header + "\n" + profile_sheet(wb[name], head_n, tail_n).to_text()
        for name in wb.sheetnames
    }


if __name__ == "__main__":
    import sys

    for name, text in profile_workbook(sys.argv[1]).items():
        print(text)
        print()
