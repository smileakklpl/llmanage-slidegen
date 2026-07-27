import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from app.ingestion.schemas import (
    FinancialStatementSubtype,
    SheetContentInspection,
    SheetContentType,
    WorkbookContentInspection,
)


FINANCIAL_KEYWORDS: dict[
    FinancialStatementSubtype,
    list[str],
] = {
    FinancialStatementSubtype.BALANCE_SHEET: [
        "資產負債表",
        "資產總計",
        "流動資產",
        "非流動資產",
        "負債總計",
        "流動負債",
        "非流動負債",
        "權益總計",
        "負債及權益總計",
        "balancesheet",
        "totalassets",
        "totalliabilities",
        "totalequity",
    ],
    FinancialStatementSubtype.INCOME_STATEMENT: [
        "損益表",
        "綜合損益表",
        "營業收入",
        "營業成本",
        "營業毛利",
        "營業利益",
        "稅前淨利",
        "本期淨利",
        "每股盈餘",
        "incomestatement",
        "revenue",
        "grossprofit",
        "netincome",
    ],
    FinancialStatementSubtype.CASH_FLOW_STATEMENT: [
        "現金流量表",
        "營業活動之現金流量",
        "投資活動之現金流量",
        "籌資活動之現金流量",
        "期初現金",
        "期末現金",
        "cashflowstatement",
        "operatingactivities",
        "investingactivities",
        "financingactivities",
    ],
}


FINANCIAL_MINIMUM_HITS: dict[
    FinancialStatementSubtype,
    int,
] = {
    FinancialStatementSubtype.BALANCE_SHEET: 4,
    FinancialStatementSubtype.INCOME_STATEMENT: 3,
    FinancialStatementSubtype.CASH_FLOW_STATEMENT: 3,
}


def _normalize_text(value: Any) -> str:
    """
    將文字正規化，方便做關鍵字比對。

    例如：
    「負債及 權益總計」會轉成「負債及權益總計」。
    """
    text = str(value).strip().lower()

    return re.sub(
        r"[\s\u3000:：\-_/()（）]+",
        "",
        text,
    )


def _is_number(value: Any) -> bool:
    """排除 bool，避免 True 被視為數字 1。"""
    return isinstance(value, (int, float)) and not isinstance(
        value,
        bool,
    )


def _count_non_empty_in_row(
    worksheet: Worksheet,
    row_number: int,
    max_column: int,
    min_column: int = 1,
) -> int:
    """
    計算指定列在欄位範圍內的非空值數量。

    min_column 預設為 1，以維持原本呼叫方式；
    表頭判斷時則只掃描候選表格的欄位範圍，
    避免工作表其他區域的內容干擾分類。
    """
    count = 0

    for column_number in range(
        min_column,
        max_column + 1,
    ):
        value = worksheet.cell(
            row=row_number,
            column=column_number,
        ).value

        if value not in (None, ""):
            count += 1

    return count


def _detect_header_row(
    worksheet: Worksheet,
    scan_rows: int = 20,
    scan_columns: int = 30,
) -> int | None:
    """
    尋找可能的表頭列。

    判斷原則：
    1. 該列至少有兩個非空值。
    2. 其中至少一半是文字。
    3. 後續至少有兩列具備足夠資料，並允許部分欄位缺失。
    """
    max_row = min(worksheet.max_row, scan_rows)
    max_column = min(worksheet.max_column, scan_columns)

    if max_row < 2 or max_column < 2:
        return None

    candidates: list[tuple[int, float]] = []

    for row_number in range(1, max_row + 1):
        values = [
            worksheet.cell(
                row=row_number,
                column=column_number,
            ).value
            for column_number in range(1, max_column + 1)
        ]

        non_empty_values = [
            value
            for value in values
            if value not in (None, "")
        ]

        if len(non_empty_values) < 2:
            continue

        text_count = sum(
            isinstance(value, str)
            and not value.startswith("=")
            for value in non_empty_values
        )

        text_ratio = text_count / len(non_empty_values)

        if text_ratio < 0.5:
            continue

        header_columns = [
            column_number
            for column_number, value in enumerate(
                values,
                start=1,
            )
            if value not in (None, "")
        ]

        header_min_column = min(header_columns)
        header_max_column = max(header_columns)

        # 允許資料列有部分欄位缺失。
        # 例如兩欄表格「名稱｜數值」中，數值欄可能暫時為空；
        # 只要至少一半的欄位有值，仍視為可能的資料列。
        minimum_data_cells = max(
            1,
            (len(non_empty_values) + 1) // 2,
        )

        following_row_counts = []

        for following_row in range(
            row_number + 1,
            min(row_number + 5, max_row + 1),
        ):
            following_row_counts.append(
                _count_non_empty_in_row(
                    worksheet,
                    following_row,
                    header_max_column,
                    header_min_column,
                )
            )

        valid_following_rows = sum(
            count >= minimum_data_cells
            for count in following_row_counts
        )

        if valid_following_rows < 2:
            continue

        score = (
            text_ratio
            + min(len(non_empty_values) / 10, 1)
            + valid_following_rows / 4
        )

        candidates.append((row_number, score))

    if not candidates:
        return None

    candidates.sort(
        key=lambda item: item[1],
        reverse=True,
    )

    return candidates[0][0]


def _calculate_structured_table_score(
    worksheet: Worksheet,
    header_row: int | None,
    non_empty_cells: int,
    numeric_cells: int,
    merged_range_count: int,
) -> tuple[float, int]:
    """計算工作表像結構化表格的程度。"""
    score = 0.0
    data_like_rows = 0

    max_scan_row = min(worksheet.max_row, 200)
    max_scan_column = min(worksheet.max_column, 50)

    if header_row is not None:
        score += 0.35

        header_columns = [
            column_number
            for column_number in range(
                1,
                max_scan_column + 1,
            )
            if worksheet.cell(
                row=header_row,
                column=column_number,
            ).value not in (None, "")
        ]

        if header_columns:
            header_min_column = min(header_columns)
            header_max_column = max(header_columns)
            header_value_count = len(header_columns)

            # 與表頭偵測採相同標準：
            # 允許部分欄位缺失，但至少一半欄位要有值。
            minimum_data_cells = max(
                1,
                (header_value_count + 1) // 2,
            )

            for row_number in range(
                header_row + 1,
                max_scan_row + 1,
            ):
                non_empty_count = _count_non_empty_in_row(
                    worksheet,
                    row_number,
                    header_max_column,
                    header_min_column,
                )

                if non_empty_count >= minimum_data_cells:
                    data_like_rows += 1

    if data_like_rows >= 2:
        score += 0.25
    elif data_like_rows == 1:
        score += 0.1

    if worksheet.max_column >= 2:
        score += 0.1

    if non_empty_cells >= 6:
        score += 0.1

    if numeric_cells >= 2:
        score += 0.15

    # 一般資料表通常不會大量使用合併儲存格。
    reasonable_merged_limit = max(
        2,
        int(non_empty_cells * 0.05),
    )

    if merged_range_count <= reasonable_merged_limit:
        score += 0.05

    return min(score, 1.0), data_like_rows


def _detect_financial_statement(
    text_values: list[str],
) -> tuple[
    FinancialStatementSubtype | None,
    float,
    list[str],
]:
    """
    使用財務關鍵字判斷報表類型。

    回傳：
    - 財務報表類型
    - 信心分數
    - 命中的關鍵字
    """
    normalized_document = "\n".join(
        _normalize_text(value)
        for value in text_values
    )

    candidate_results: list[
        tuple[
            FinancialStatementSubtype,
            float,
            list[str],
        ]
    ] = []

    for subtype, keywords in FINANCIAL_KEYWORDS.items():
        hits = [
            keyword
            for keyword in keywords
            if _normalize_text(keyword) in normalized_document
        ]

        minimum_hits = FINANCIAL_MINIMUM_HITS[subtype]

        if len(hits) < minimum_hits:
            continue

        confidence = min(
            0.99,
            0.55 + len(hits) * 0.08,
        )

        candidate_results.append(
            (
                subtype,
                confidence,
                hits,
            )
        )

    if not candidate_results:
        return None, 0.0, []

    candidate_results.sort(
        key=lambda item: (
            item[1],
            len(item[2]),
        ),
        reverse=True,
    )

    return candidate_results[0]


def _inspect_sheet(
    worksheet: Worksheet,
) -> SheetContentInspection:
    non_empty_cells = 0
    text_cells = 0
    numeric_cells = 0
    formula_cells = 0
    date_cells = 0

    text_values: list[str] = []
    warnings: list[str] = []
    evidence: list[str] = []

    # 避免格式異常的 Excel 宣告過大使用範圍，
    # 第一版最多掃描前 1000 列、100 欄。
    scan_max_row = min(worksheet.max_row, 1000)
    scan_max_column = min(worksheet.max_column, 100)

    if worksheet.max_row > 1000:
        warnings.append(
            "工作表超過 1000 列，分類階段只掃描前 1000 列"
        )

    if worksheet.max_column > 100:
        warnings.append(
            "工作表超過 100 欄，分類階段只掃描前 100 欄"
        )

    for row in worksheet.iter_rows(
        min_row=1,
        max_row=scan_max_row,
        min_col=1,
        max_col=scan_max_column,
    ):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue

            value = cell.value

            if value in (None, ""):
                continue

            non_empty_cells += 1

            if isinstance(value, str):
                if value.startswith("="):
                    formula_cells += 1
                else:
                    text_cells += 1
                    text_values.append(value)

            elif isinstance(value, (date, datetime)):
                date_cells += 1

            elif _is_number(value):
                numeric_cells += 1

            else:
                text_cells += 1
                text_values.append(str(value))

    chart_count = len(getattr(worksheet, "_charts", []))
    image_count = len(getattr(worksheet, "_images", []))
    merged_range_count = len(
        worksheet.merged_cells.ranges
    )

    header_row = _detect_header_row(worksheet)

    (
        structured_table_score,
        data_like_rows,
    ) = _calculate_structured_table_score(
        worksheet=worksheet,
        header_row=header_row,
        non_empty_cells=non_empty_cells,
        numeric_cells=numeric_cells,
        merged_range_count=merged_range_count,
    )

    (
        financial_subtype,
        financial_score,
        financial_hits,
    ) = _detect_financial_statement(text_values)

    components: list[SheetContentType] = []

    if financial_subtype is not None:
        components.append(
            SheetContentType.FINANCIAL_STATEMENT
        )

        evidence.append(
            "辨識到財務報表關鍵字："
            + "、".join(financial_hits[:8])
        )

    elif structured_table_score >= 0.6:
        components.append(
            SheetContentType.STRUCTURED_TABLE
        )

        evidence.append(
            f"偵測到約 {data_like_rows} 列類表格資料"
        )

    if chart_count > 0:
        components.append(
            SheetContentType.NATIVE_CHART
        )
        evidence.append(
            f"工作表包含 {chart_count} 個 Excel 原生圖表"
        )

    if image_count > 0:
        components.append(
            SheetContentType.EMBEDDED_IMAGE
        )
        evidence.append(
            f"工作表包含 {image_count} 張內嵌圖片"
        )

    if header_row is not None:
        evidence.append(
            f"第 {header_row} 列可能是表頭"
        )

    if (
        non_empty_cells == 0
        and chart_count == 0
        and image_count == 0
    ):
        primary_content_type = SheetContentType.EMPTY
        confidence = 1.0

    elif len(components) >= 2:
        primary_content_type = (
            SheetContentType.MIXED_CONTENT
        )
        confidence = max(
            structured_table_score,
            financial_score,
            0.9,
        )

    elif len(components) == 1:
        primary_content_type = components[0]

        if (
            primary_content_type
            == SheetContentType.FINANCIAL_STATEMENT
        ):
            confidence = financial_score

        elif (
            primary_content_type
            == SheetContentType.STRUCTURED_TABLE
        ):
            confidence = structured_table_score

        else:
            # 原生圖表與內嵌圖片由檔案結構直接確認。
            confidence = 1.0

    else:
        primary_content_type = SheetContentType.UNKNOWN
        confidence = 0.3

        warnings.append(
            "工作表有內容，但目前無法判斷其資料結構"
        )

    return SheetContentInspection(
        sheet_name=worksheet.title,
        sheet_state=worksheet.sheet_state,
        primary_content_type=primary_content_type,
        components=components,
        financial_statement_subtype=financial_subtype,
        confidence=round(confidence, 4),
        max_row=worksheet.max_row,
        max_column=worksheet.max_column,
        non_empty_cells=non_empty_cells,
        text_cells=text_cells,
        numeric_cells=numeric_cells,
        formula_cells=formula_cells,
        date_cells=date_cells,
        merged_range_count=merged_range_count,
        chart_count=chart_count,
        image_count=image_count,
        detected_header_row=header_row,
        structured_table_score=round(
            structured_table_score,
            4,
        ),
        financial_statement_score=round(
            financial_score,
            4,
        ),
        evidence=evidence,
        warnings=warnings,
    )


def _calculate_overall_content_type(
    sheet_results: list[SheetContentInspection],
) -> SheetContentType:
    meaningful_types = [
        result.primary_content_type
        for result in sheet_results
        if result.primary_content_type
        != SheetContentType.EMPTY
    ]

    if not meaningful_types:
        return SheetContentType.EMPTY

    if len(set(meaningful_types)) == 1:
        return meaningful_types[0]

    return SheetContentType.MIXED_CONTENT


def inspect_excel_content(
    file_path: str | Path,
) -> WorkbookContentInspection:
    """
    掃描 Excel 每一張工作表並判斷內容類型。

    此函式只接受 .xlsx。
    """
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(
            f"找不到 Excel 檔案：{path}"
        )

    warnings: list[str] = []

    try:
        workbook = load_workbook(
            filename=path,
            read_only=False,
            data_only=False,
        )
    except Exception as error:
        raise ValueError(
            f"Excel 檔案無法開啟：{error}"
        ) from error

    try:
        sheet_results = [
            _inspect_sheet(worksheet)
            for worksheet in workbook.worksheets
        ]
    finally:
        workbook.close()

    overall_content_type = (
        _calculate_overall_content_type(
            sheet_results
        )
    )

    if sheet_results:
        confidence = sum(
            result.confidence
            for result in sheet_results
        ) / len(sheet_results)
    else:
        confidence = 0.0
        warnings.append("Excel 中沒有任何工作表")

    hidden_sheets = [
        result.sheet_name
        for result in sheet_results
        if result.sheet_state != "visible"
    ]

    if hidden_sheets:
        warnings.append(
            "發現隱藏工作表："
            + "、".join(hidden_sheets)
        )

    return WorkbookContentInspection(
        filename=path.name,
        sheet_count=len(sheet_results),
        overall_content_type=overall_content_type,
        confidence=round(confidence, 4),
        sheets=sheet_results,
        warnings=warnings,
    )