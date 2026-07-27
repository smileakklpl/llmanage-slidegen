import json
import os
import re
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pymupdf
from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter
from PIL import Image

from app.ingestion.schemas import (
    ColumnDataType,
    ExtractedCell,
    ExtractedTableRow,
    SheetContentInspection,
    SheetContentType,
    SourceCell,
    TableColumnSpec,
    TableDatasetSpec,
    TableMetadata,
    VisualDocumentContent,
    VisualPageContent,
    VisualTableContent,
    VisualTextBlock,
    WorkbookContentInspection,
    WorkbookTableExtraction,
)


MAX_VISUAL_PAGES = 20
PDF_RENDER_DPI = 200
HUMAN_REVIEW_CONFIDENCE = 0.80

_ENGINE_CACHE: dict[bool, Any] = {}


def _env_flag(
    name: str,
    default: bool = False,
) -> bool:
    value = os.getenv(
        name,
        "1" if default else "0",
    )

    return value.strip().lower() in {
        "1",
        "true",
        "yes",
        "on",
    }


def _get_engine(
    enable_chart_recognition: bool,
) -> Any:
    """
    延遲載入 OCR 模型。

    避免只啟動 FastAPI 時就立刻下載與載入模型。
    """
    if enable_chart_recognition in _ENGINE_CACHE:
        return _ENGINE_CACHE[
            enable_chart_recognition
        ]

    try:
        from paddleocr import PPStructureV3
    except ImportError as error:
        raise RuntimeError(
            "尚未安裝 PaddleOCR，請先執行 "
            'python -m pip install "paddleocr[all]"'
        ) from error

    engine = PPStructureV3(
        use_doc_orientation_classify=True,
        use_doc_unwarping=False,
        use_textline_orientation=True,
        use_formula_recognition=False,
        use_chart_recognition=(
            enable_chart_recognition
        ),
        format_block_content=True,
        device=os.getenv(
            "OCR_DEVICE",
            "cpu",
        ),
    )

    _ENGINE_CACHE[
        enable_chart_recognition
    ] = engine

    return engine


def _result_to_dict(
    result: Any,
) -> dict[str, Any]:
    data = getattr(result, "json", {})

    if callable(data):
        data = data()

    if isinstance(data, str):
        data = json.loads(data)

    if not isinstance(data, dict):
        raise ValueError(
            "PaddleOCR 回傳格式不是 JSON object"
        )

    return data


def _bbox_from_polygon(
    polygon: Any,
) -> list[float]:
    if polygon is None:
        return []

    try:
        points = list(polygon)

        xs = [
            float(point[0])
            for point in points
        ]

        ys = [
            float(point[1])
            for point in points
        ]

        return [
            min(xs),
            min(ys),
            max(xs),
            max(ys),
        ]

    except Exception:
        return []


def _normalize_scalar(
    value: Any,
) -> Any:
    if not isinstance(value, str):
        return value

    text = re.sub(
        r"\s+",
        " ",
        value.strip(),
    )

    if not text:
        return None

    if re.fullmatch(
        r"\([+-]?\d[\d,]*(?:\.\d+)?\)",
        text,
    ):
        number = float(
            text[1:-1].replace(",", "")
        )

        return (
            -int(number)
            if number.is_integer()
            else -number
        )

    if re.fullmatch(
        r"[+-]?\d[\d,]*(?:\.\d+)?",
        text,
    ):
        number = float(
            text.replace(",", "")
        )

        return (
            int(number)
            if number.is_integer()
            else number
        )

    if re.fullmatch(
        r"\d{4}-\d{1,2}-\d{1,2}",
        text,
    ):
        try:
            return date.fromisoformat(text)
        except ValueError:
            pass

    return text


def _infer_data_type(
    values: list[Any],
) -> ColumnDataType:
    non_empty_values = [
        value
        for value in values
        if value is not None
    ]

    if not non_empty_values:
        return ColumnDataType.EMPTY

    types: set[ColumnDataType] = set()

    for value in non_empty_values:
        if isinstance(value, bool):
            types.add(ColumnDataType.BOOLEAN)

        elif isinstance(value, datetime):
            types.add(ColumnDataType.DATETIME)

        elif isinstance(value, date):
            types.add(ColumnDataType.DATE)

        elif (
            isinstance(value, int)
            and not isinstance(value, bool)
        ):
            types.add(ColumnDataType.INTEGER)

        elif isinstance(value, float):
            types.add(ColumnDataType.NUMBER)

        else:
            types.add(ColumnDataType.STRING)

    if types == {ColumnDataType.INTEGER}:
        return ColumnDataType.INTEGER

    if types.issubset({
        ColumnDataType.INTEGER,
        ColumnDataType.NUMBER,
    }):
        return ColumnDataType.NUMBER

    if len(types) == 1:
        return next(iter(types))

    return ColumnDataType.MIXED


def _create_column_key(
    label: str,
    index: int,
    existing_keys: set[str],
) -> str:
    key = re.sub(
        r"[^\w]+",
        "_",
        label.strip().lower(),
        flags=re.UNICODE,
    ).strip("_")

    if not key:
        key = f"column_{index + 1}"

    candidate = key
    duplicate_index = 2

    while candidate in existing_keys:
        candidate = (
            f"{key}_{duplicate_index}"
        )

        duplicate_index += 1

    return candidate


def _html_table_rows(
    html: str,
) -> list[list[str]]:
    soup = BeautifulSoup(
        html,
        "html.parser",
    )

    table = soup.find("table")

    if table is None:
        return []

    rows: list[list[str]] = []

    for row_element in table.find_all("tr"):
        cells = row_element.find_all(
            ["th", "td"],
            recursive=False,
        )

        row = [
            re.sub(
                r"\s+",
                " ",
                cell.get_text(
                    " ",
                    strip=True,
                ),
            )
            for cell in cells
        ]

        if row:
            rows.append(row)

    if not rows:
        return []

    width = max(
        len(row)
        for row in rows
    )

    for row in rows:
        row.extend(
            [""] * (
                width - len(row)
            )
        )

    return rows


def _html_to_dataset(
    filename: str,
    page_number: int,
    table_index: int,
    html: str,
    confidence: float,
) -> TableDatasetSpec | None:
    rows = _html_table_rows(html)

    if len(rows) < 2:
        return None

    header = rows[0]
    data_rows = rows[1:]

    virtual_sheet = (
        f"visual_page_{page_number}_"
        f"table_{table_index}"
    )

    existing_keys: set[str] = set()
    definitions: list[
        tuple[int, str, str]
    ] = []

    for column_index, raw_label in enumerate(
        header
    ):
        label = raw_label.strip()

        if not label:
            label = (
                f"未命名欄位 "
                f"{column_index + 1}"
            )

        key = _create_column_key(
            label,
            column_index,
            existing_keys,
        )

        existing_keys.add(key)

        definitions.append(
            (
                column_index,
                key,
                label,
            )
        )

    values_by_column: dict[
        str,
        list[Any],
    ] = {
        key: []
        for _, key, _ in definitions
    }

    empty_counts = {
        key: 0
        for _, key, _ in definitions
    }

    extracted_rows: list[
        ExtractedTableRow
    ] = []

    for row_index, raw_row in enumerate(
        data_rows,
        start=2,
    ):
        row_cells: dict[
            str,
            ExtractedCell,
        ] = {}

        row_has_value = False

        for (
            column_index,
            key,
            _,
        ) in definitions:
            raw_value = (
                raw_row[column_index]
                if column_index < len(raw_row)
                else ""
            )

            value = _normalize_scalar(
                raw_value
            )

            if value is None:
                empty_counts[key] += 1
            else:
                row_has_value = True

            values_by_column[key].append(
                value
            )

            coordinate = (
                f"{get_column_letter(column_index + 1)}"
                f"{row_index}"
            )

            row_cells[key] = ExtractedCell(
                raw_value=raw_value,
                value=value,
                formula=None,
                number_format=None,
                source=SourceCell(
                    sheet=virtual_sheet,
                    cell=coordinate,
                    row=row_index,
                    column=column_index + 1,
                ),
            )

        if row_has_value:
            extracted_rows.append(
                ExtractedTableRow(
                    excel_row=row_index,
                    cells=row_cells,
                )
            )

    columns: list[
        TableColumnSpec
    ] = []

    for column_index, key, label in definitions:
        coordinate = (
            f"{get_column_letter(column_index + 1)}1"
        )

        columns.append(
            TableColumnSpec(
                key=key,
                label=label,
                index=column_index,
                data_type=_infer_data_type(
                    values_by_column[key]
                ),
                unit=(
                    "%"
                    if "%" in label
                    else None
                ),
                nullable=(
                    empty_counts[key] > 0
                ),
                header_source=SourceCell(
                    sheet=virtual_sheet,
                    cell=coordinate,
                    row=1,
                    column=column_index + 1,
                ),
            )
        )

    last_column = get_column_letter(
        len(columns)
    )

    last_row = max(
        1,
        len(extracted_rows) + 1,
    )

    return TableDatasetSpec(
        filename=filename,
        sheet_name=virtual_sheet,
        table_kind=(
            SheetContentType
            .STRUCTURED_TABLE
        ),
        financial_statement_subtype=None,
        metadata=TableMetadata(
            title=None,
            entity=None,
            unit=None,
            notes=[],
        ),
        header_row=1,
        header_range=(
            f"{virtual_sheet}!A1:"
            f"{last_column}1"
        ),
        data_range=(
            f"{virtual_sheet}!A2:"
            f"{last_column}{last_row}"
            if extracted_rows
            else None
        ),
        full_range=(
            f"{virtual_sheet}!A1:"
            f"{last_column}{last_row}"
        ),
        row_count=len(extracted_rows),
        column_count=len(columns),
        columns=columns,
        rows=extracted_rows,
        warnings=[
            "此表格由圖片 OCR 與版面模型產生",
            (
                "視覺表格平均信心分數："
                f"{confidence:.2f}"
            ),
            "合併儲存格可能需要人工確認",
        ],
    )


def _prepare_page_images(
    file_path: Path,
    output_directory: Path,
) -> tuple[
    list[tuple[int, Path]],
    int,
    list[str],
]:
    warnings: list[str] = []

    if file_path.suffix.lower() == ".pdf":
        document = pymupdf.open(
            str(file_path)
        )

        try:
            total_pages = len(document)

            processed_pages = min(
                total_pages,
                MAX_VISUAL_PAGES,
            )

            if total_pages > MAX_VISUAL_PAGES:
                warnings.append(
                    f"掃描 PDF 共 {total_pages} 頁，"
                    f"只進行前 {MAX_VISUAL_PAGES} 頁"
                    "的視覺辨識"
                )

            page_images = []

            for page_index in range(
                processed_pages
            ):
                page = document[
                    page_index
                ]

                pixmap = page.get_pixmap(
                    dpi=PDF_RENDER_DPI,
                    alpha=False,
                )

                image_path = (
                    output_directory
                    / f"page_{page_index + 1}.png"
                )

                pixmap.save(
                    str(image_path)
                )

                page_images.append(
                    (
                        page_index + 1,
                        image_path,
                    )
                )

            return (
                page_images,
                total_pages,
                warnings,
            )

        finally:
            document.close()

    try:
        with Image.open(file_path) as image:
            image.verify()
    except Exception as error:
        raise ValueError(
            f"圖片無法開啟：{error}"
        ) from error

    return (
        [(1, file_path)],
        1,
        warnings,
    )


def _classify_page(
    has_text: bool,
    has_table: bool,
    has_chart: bool,
    has_image: bool,
) -> tuple[
    SheetContentType,
    list[SheetContentType],
]:
    components: list[
        SheetContentType
    ] = []

    if has_text:
        components.append(
            SheetContentType.DOCUMENT_TEXT
        )

    if has_table:
        components.append(
            SheetContentType.STRUCTURED_TABLE
        )

    if has_chart:
        components.append(
            SheetContentType.CHART_IMAGE
        )

    if has_image:
        components.append(
            SheetContentType.EMBEDDED_IMAGE
        )

    if has_chart and has_table:
        primary = (
            SheetContentType.MIXED_CONTENT
        )

    elif has_chart:
        primary = (
            SheetContentType.CHART_IMAGE
        )

    elif has_table:
        primary = (
            SheetContentType
            .STRUCTURED_TABLE
        )

    elif has_text:
        primary = (
            SheetContentType.DOCUMENT_TEXT
        )

    elif has_image:
        primary = (
            SheetContentType.EMBEDDED_IMAGE
        )

    else:
        primary = SheetContentType.UNKNOWN

    return primary, components


def inspect_visual_input(
    file_path: str | Path,
    original_filename: str | None = None,
    enable_chart_recognition: (
        bool | None
    ) = None,
    engine: Any | None = None,
) -> tuple[
    WorkbookContentInspection,
    VisualDocumentContent,
    WorkbookTableExtraction,
]:
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(
            f"找不到視覺檔案：{path}"
        )

    filename = (
        original_filename or path.name
    )

    if enable_chart_recognition is None:
        enable_chart_recognition = _env_flag(
            "ENABLE_CHART_RECOGNITION",
            default=False,
        )

    if engine is None:
        engine = _get_engine(
            enable_chart_recognition
        )

    visual_pages: list[
        VisualPageContent
    ] = []

    classifications: list[
        SheetContentInspection
    ] = []

    extracted_tables: list[
        TableDatasetSpec
    ] = []

    skipped_pages: list[str] = []
    warnings: list[str] = []

    with tempfile.TemporaryDirectory() as temp:
        page_images, total_pages, render_warnings = (
            _prepare_page_images(
                path,
                Path(temp),
            )
        )

        warnings.extend(render_warnings)

        for page_number, image_path in page_images:
            prediction = engine.predict(
                str(image_path)
            )

            result_items = list(
                prediction
            )

            if not result_items:
                warnings.append(
                    f"第 {page_number} 頁"
                    "沒有取得視覺辨識結果"
                )

                skipped_pages.append(
                    f"page_{page_number}"
                )

                continue

            payload = _result_to_dict(
                result_items[0]
            )

            width = int(
                payload.get("width", 0) or 0
            )

            height = int(
                payload.get("height", 0) or 0
            )

            ocr_result = (
                payload.get(
                    "overall_ocr_res",
                    {},
                )
                or {}
            )

            texts = list(
                ocr_result.get(
                    "rec_texts",
                    [],
                )
                or []
            )

            scores = [
                float(score)
                for score in (
                    ocr_result.get(
                        "rec_scores",
                        [],
                    )
                    or []
                )
            ]

            polygons = list(
                ocr_result.get(
                    "rec_polys",
                    [],
                )
                or []
            )

            text_blocks: list[
                VisualTextBlock
            ] = []

            for index, text in enumerate(texts):
                confidence = (
                    scores[index]
                    if index < len(scores)
                    else 0.0
                )

                polygon = (
                    polygons[index]
                    if index < len(polygons)
                    else None
                )

                text_blocks.append(
                    VisualTextBlock(
                        text=str(text),
                        confidence=confidence,
                        bbox=_bbox_from_polygon(
                            polygon
                        ),
                        label="text",
                    )
                )

            average_confidence = (
                sum(scores) / len(scores)
                if scores
                else 0.0
            )

            parsing_blocks = list(
                payload.get(
                    "parsing_res_list",
                    [],
                )
                or []
            )

            labels = [
                str(
                    block.get(
                        "block_label",
                        "",
                    )
                ).lower()
                for block in parsing_blocks
            ]

            chart_blocks = [
                block
                for block in parsing_blocks
                if "chart" in str(
                    block.get(
                        "block_label",
                        "",
                    )
                ).lower()
            ]

            chart_contents = [
                str(
                    block.get(
                        "block_content",
                        "",
                    )
                ).strip()
                for block in chart_blocks
                if str(
                    block.get(
                        "block_content",
                        "",
                    )
                ).strip()
            ]

            table_results = list(
                payload.get(
                    "table_res_list",
                    [],
                )
                or []
            )

            visual_tables: list[
                VisualTableContent
            ] = []

            extracted_on_page = 0

            for table_index, table_result in enumerate(
                table_results,
                start=1,
            ):
                html = table_result.get(
                    "pred_html"
                )

                table_ocr = (
                    table_result.get(
                        "table_ocr_pred",
                        {},
                    )
                    or {}
                )

                table_scores = [
                    float(score)
                    for score in (
                        table_ocr.get(
                            "rec_scores",
                            [],
                        )
                        or []
                    )
                ]

                table_confidence = (
                    sum(table_scores)
                    / len(table_scores)
                    if table_scores
                    else average_confidence
                )

                table_rows = (
                    _html_table_rows(html)
                    if html
                    else []
                )

                visual_tables.append(
                    VisualTableContent(
                        page_number=page_number,
                        table_index=table_index,
                        html=html,
                        row_count=len(
                            table_rows
                        ),
                        column_count=(
                            max(
                                (
                                    len(row)
                                    for row
                                    in table_rows
                                ),
                                default=0,
                            )
                        ),
                        confidence=round(
                            table_confidence,
                            4,
                        ),
                        bbox=[],
                    )
                )

                if html:
                    dataset = _html_to_dataset(
                        filename=filename,
                        page_number=page_number,
                        table_index=table_index,
                        html=html,
                        confidence=(
                            table_confidence
                        ),
                    )

                    if dataset is not None:
                        extracted_tables.append(
                            dataset
                        )

                        extracted_on_page += 1

            has_text = bool(
                "\n".join(texts).strip()
            )

            has_table = (
                bool(table_results)
                or "table" in labels
            )

            has_chart = bool(chart_blocks)

            has_image = any(
                label in {
                    "image",
                    "figure",
                    "picture",
                }
                for label in labels
            )

            (
                primary_type,
                components,
            ) = _classify_page(
                has_text=has_text,
                has_table=has_table,
                has_chart=has_chart,
                has_image=has_image,
            )

            page_warnings: list[str] = []

            requires_human_review = (
                average_confidence
                < HUMAN_REVIEW_CONFIDENCE
                or has_chart
                or (
                    has_table
                    and extracted_on_page == 0
                )
            )

            if (
                average_confidence
                < HUMAN_REVIEW_CONFIDENCE
            ):
                page_warnings.append(
                    "OCR 平均信心分數低於 0.80"
                )

            if has_chart:
                page_warnings.append(
                    "圖表截圖的數值需人工確認，"
                    "不得只依長條高度或折線位置"
                    "直接視為正確數值"
                )

            if (
                has_table
                and extracted_on_page == 0
            ):
                page_warnings.append(
                    "偵測到表格區域，"
                    "但無法建立結構化資料表"
                )

            visual_pages.append(
                VisualPageContent(
                    page_number=page_number,
                    width=width,
                    height=height,
                    primary_content_type=(
                        primary_type
                    ),
                    components=components,
                    ocr_text="\n".join(texts),
                    average_ocr_confidence=round(
                        average_confidence,
                        4,
                    ),
                    text_blocks=text_blocks,
                    tables=visual_tables,
                    chart_count=len(
                        chart_blocks
                    ),
                    chart_contents=(
                        chart_contents
                    ),
                    likely_scanned=True,
                    requires_human_review=(
                        requires_human_review
                    ),
                    warnings=page_warnings,
                )
            )

            classifications.append(
                SheetContentInspection(
                    sheet_name=(
                        f"page_{page_number}"
                    ),
                    sheet_state="visible",
                    primary_content_type=(
                        primary_type
                    ),
                    components=components,
                    financial_statement_subtype=None,
                    confidence=round(
                        average_confidence,
                        4,
                    ),
                    max_row=len(texts),
                    max_column=1,
                    non_empty_cells=len(texts),
                    text_cells=len(texts),
                    numeric_cells=0,
                    formula_cells=0,
                    date_cells=0,
                    merged_range_count=0,
                    chart_count=len(
                        chart_blocks
                    ),
                    image_count=(
                        1 if has_image else 0
                    ),
                    detected_header_row=None,
                    structured_table_score=(
                        0.9
                        if has_table
                        else 0.0
                    ),
                    financial_statement_score=0.0,
                    evidence=[
                        f"OCR 文字區塊：{len(texts)}",
                        (
                            "OCR 平均信心："
                            f"{average_confidence:.2f}"
                        ),
                        (
                            "視覺表格數："
                            f"{len(table_results)}"
                        ),
                        (
                            "圖表區塊數："
                            f"{len(chart_blocks)}"
                        ),
                    ],
                    warnings=page_warnings,
                )
            )

            if extracted_on_page == 0:
                skipped_pages.append(
                    f"page_{page_number}"
                )

    meaningful_types = [
        result.primary_content_type
        for result in classifications
    ]

    if not meaningful_types:
        overall_type = (
            SheetContentType.UNKNOWN
        )

    elif len(set(meaningful_types)) == 1:
        overall_type = meaningful_types[0]

    else:
        overall_type = (
            SheetContentType.MIXED_CONTENT
        )

    confidence_values = [
        page.average_ocr_confidence
        for page in visual_pages
    ]

    document_confidence = (
        sum(confidence_values)
        / len(confidence_values)
        if confidence_values
        else 0.0
    )

    requires_review = any(
        page.requires_human_review
        for page in visual_pages
    )

    visual_document = VisualDocumentContent(
        filename=filename,
        engine="PaddleOCR PPStructureV3",
        page_count=total_pages,
        processed_page_count=len(
            visual_pages
        ),
        average_ocr_confidence=round(
            document_confidence,
            4,
        ),
        chart_recognition_enabled=(
            enable_chart_recognition
        ),
        requires_human_review=(
            requires_review
        ),
        pages=visual_pages,
        warnings=warnings.copy(),
    )

    classification = (
        WorkbookContentInspection(
            filename=filename,
            sheet_count=len(
                classifications
            ),
            overall_content_type=(
                overall_type
            ),
            confidence=round(
                document_confidence,
                4,
            ),
            sheets=classifications,
            warnings=warnings.copy(),
        )
    )

    extraction = WorkbookTableExtraction(
        filename=filename,
        table_count=len(
            extracted_tables
        ),
        tables=extracted_tables,
        skipped_sheets=skipped_pages,
        warnings=warnings.copy(),
    )

    return (
        classification,
        visual_document,
        extraction,
    )